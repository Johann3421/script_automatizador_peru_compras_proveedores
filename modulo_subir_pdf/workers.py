import sys, os, time, threading, queue, json, re
from io import BytesIO
from pathlib import Path
from utils.logger import LogWriter
from automation.browser import init_browser, close_browser
from automation.login import do_login

def _make_stock_log(append_fn):
    class _StockLog:
        def info(self, msg): append_fn(str(msg))
        def warning(self, msg): append_fn(f"⚠ {msg}")
        def error(self, msg): append_fn(f"❌ {msg}")
        def success(self, msg): append_fn(f"✅ {msg}")
        def ok(self, msg): append_fn(f"✅ {msg}")
    return _StockLog()

def execute_stock(app, usuario, password, acuerdo, catalogo, categoria, pausa):
    from automation.browser import init_browser, close_browser
    from automation.login import do_login
    from automation_otro_bot.stock import (
        ejecutar_stock, paso2_navegacion_stock, paso3_filtros_stock,
        paso4_actualizar_stock, analizar_excel_stock, generar_reporte_excel,
        RESULTADOS, STOP_EVENT,
    )
    # Estado compartido
    global STOP_EVENT
    STOP_EVENT = app._stock_stop_event
    # LogWriter puente → reenvía a la UI (pestaña 2)
    log = _make_stock_log(app._append_stock_log)
    pw = browser = page = None
    app._stock_pw = None
    app._stock_browser = None
    try:
        app._append_stock_log("🚀 Iniciando navegador...")
        headless = not bool(app.check_stock_visible.get())
        pw, browser, page = init_browser(headless=headless)
        app._stock_pw = pw
        app._stock_browser = browser
        try:
            page.set_viewport_size({"width": 1920, "height": 1080})
        except Exception:
            pass
        app._append_stock_log("🔐 Login...")
        ok = do_login(page, usuario, password, "", log, app._stock_stop_event,
                      app.captcha_bridge)
        if not ok or app._stock_stop_event.is_set():
            app._append_stock_log("❌ Login falló")
            return
        app._append_stock_log("📍 Navegando a MejoraBasica...")
        paso2_navegacion_stock(page)
        app._append_stock_log(f"📋 Filtros: {acuerdo} > {catalogo} > {categoria}")
        if not paso3_filtros_stock(page, acuerdo, catalogo, categoria):
            app._append_stock_log("❌ No se pudieron aplicar filtros")
            return
        app._stock_total = len(app._stock_excel_df)
        app.lbl_stock_stat_total.configure(text=str(app._stock_total))
        app.progress_stock.set(0)
        def stock_log(msg):
            app._append_stock_log(msg)
            if "✅" in msg:
                app._stock_ok += 1
                app.lbl_stock_stat_ok.configure(text=str(app._stock_ok))
            elif "❌" in msg:
                app._stock_errors += 1
                app.lbl_stock_stat_fail.configure(text=str(app._stock_errors))
            app.progress_stock.set(
                (app._stock_ok + app._stock_errors) / max(1, app._stock_total)
            )
        exitos = paso4_actualizar_stock(
            page, app._stock_excel_df, pausa, stock_log,
            usuario, password, app.captcha_bridge,
            acuerdo, catalogo, categoria
        )
        stock_log(f"✅ {exitos}/{app._stock_total} actualizados")
    except Exception as e:
        if "has been closed" in str(e).lower() or "connection closed" in str(e).lower():
            app._append_stock_log("⏹ Proceso detenido por el usuario")
        else:
            app._append_stock_log(f"❌ Error fatal: {e}")
            import traceback
            app._append_stock_log(traceback.format_exc())
    finally:
        if browser and pw:
            try:
                close_browser(pw, browser)
            except Exception:
                pass
        # Generar reporte SIEMPRE (incluso si se detuvo o falló)
        try:
            output_dir = os.path.dirname(app._stock_excel_path) or "."
            ts = time.strftime("%Y%m%d_%H%M%S")
            report_path = os.path.join(output_dir, f"reporte_stock_{ts}.xlsx")
            generar_reporte_excel(report_path, acuerdo, catalogo, categoria)
            app._stock_report_path = report_path
            app.lbl_stock_report.configure(text=os.path.basename(report_path),
                                            text_color="#5dade2")
            app._append_stock_log(f"📊 Reporte: {report_path}")
        except Exception as rep_err:
            app._append_stock_log(f"⚠ No se pudo generar reporte: {rep_err}")
        app._stock_pw = None
        app._stock_browser = None
        app._stock_running = False
        app.btn_stock_start.configure(state="normal")
        app.btn_stock_stop.configure(state="disabled")
        status_text = "Detenido" if app._stock_stop_event.is_set() else "Completado"
        status_color = "#e74c3c" if app._stock_stop_event.is_set() else "#5dade2"
        app.lbl_stock_status.configure(text=status_text, text_color=status_color)
def execute_extract(app, usuario, password, headless):
    import re, json as _json, urllib.parse
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        log.info("📊 Iniciando navegador...")
        if stop.is_set(): return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass
        log.info("📊 Navegador listo")
        if stop.is_set(): return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok: log.error("📊 Login fallido.")
            return
        log.ok("📊 Login exitoso.")
        # Navegar a reportes, seleccionar Acuerdo 249, Catalogo 252, Categoria 11735
        # para cargar el HTML con el endpoint real y las cookies de sesión
        REPORTES_URL = "https://www.catalogos.perucompras.gob.pe/Reportes/ProductoOfertadoIndex"
        page.goto(REPORTES_URL, wait_until="networkidle", timeout=60_000)
        time.sleep(2)
        # El endpoint real (del form action)
        ENDPOINT = "https://www.catalogos.perucompras.gob.pe/Reportes/_detProductoOfertadoIndex"
        def _read_options(selector):
            opts = []
            try:
                for o in page.locator(selector).first.locator("option").all():
                    v = o.get_attribute("value")
                    if v and v.strip() and v.strip() != "0":
                        opts.append({"value": v.strip(), "text": o.inner_text().strip()})
            except Exception: pass
            return opts
        # Seleccionar Acuerdo 249 para cargar catálogos
        log.info("📊 Cargando menú de catálogos...")
        page.locator("#ajaxAcuerdo").first.select_option(value="249")
        time.sleep(3)
        acuerdos = _read_options("#ajaxAcuerdo")
        output_dir = os.path.join(_THIS_DIR, "output_extract")
        os.makedirs(output_dir, exist_ok=True)
        index_data = {"acuerdo": "249", "combinaciones": [], "extracted_at": time.strftime("%Y-%m-%d %H:%M:%S")}
        total = 0
        for ac in acuerdos:
            if stop.is_set(): break
            log.info(f"📊 Acuerdo: {ac['text']} ({ac['value']})")
            page.locator("#ajaxAcuerdo").first.select_option(value=ac["value"])
            time.sleep(3)
            catalogos = _read_options("#ajaxCatalogo")
            for cat in catalogos:
                if stop.is_set(): break
                log.info(f"📊   Catalogo: {cat['text']} ({cat['value']})")
                page.locator("#ajaxCatalogo").first.select_option(value=cat["value"])
                time.sleep(3)
                categorias = _read_options("#ajaxCategoria")
                for cg in categorias:
                    if stop.is_set(): break
                    log.info(f"📊     Categoria: {cg['text']} ({cg['value']})")
                    try:
                        # GET como hace ListarProductosOfertados: devuelve HTML con TODOS los <tr>
                        qs = f"N_Acuerdo={ac['value']}&N_Catalogo={cat['value']}&N_Categoria={cg['value']}&C_Descripcion="
                        url = f"https://www.catalogos.perucompras.gob.pe/Reportes/_detProductoOfertadoIndex?{qs}"
                        html = page.evaluate("""
                            async (url) => {
                                const resp = await fetch(url, {
                                    method: 'GET',
                                    headers: { 'X-Requested-With': 'XMLHttpRequest' }
                                });
                                return await resp.text();
                            }
                        """, url)
                        # Parsear HTML: <tbody id="rOferta"> → <tr class="gradeA"> → <td>
                        records = []
                        tbody = re.search(r'<tbody\s+id="rOferta"[^>]*>(.*?)</tbody>', html, re.DOTALL)
                        if tbody:
                            for tr in re.finditer(r'<tr\b[^>]*>(.*?)</tr>', tbody.group(1), re.DOTALL):
                                row_html = tr.group(1)
                                id_m = re.search(r'value="(\d+)"', row_html)
                                tds = [re.sub(r'<[^>]+>', ' ', td).replace('&amp;', '&').strip()
                                       for td in re.findall(r'<td\b[^>]*>(.*?)</td>', row_html, re.DOTALL)]
                                tds = [re.sub(r'\s+', ' ', t).strip() for t in tds]
                                if len(tds) >= 10:
                                    records.append({
                                        "ID_ProductoOfertado": id_m.group(1) if id_m else "",
                                        "descripcion": tds[1], "moneda": tds[3], "precio": tds[4],
                                        "fecha_registro": tds[5], "estado_ficha": tds[6],
                                        "estado_oferta": tds[7], "fecha_adjudicacion": tds[8],
                                        "fecha_publicacion": tds[9], "motivo": tds[10],
                                        "justificacion": tds[11] if len(tds) > 11 else "",
                                        "puntaje": tds[12] if len(tds) > 12 else "",
                                    })
                        safe = lambda s: re.sub(r'[<>:"/\\|?*]', '_', s[:40])
                        ac_dir = f"{ac['value']}_{safe(ac['text'])}"
                        cat_dir = f"{cat['value']}_{safe(cat['text'])}"
                        cat_path = os.path.join(output_dir, ac_dir, cat_dir)
                        os.makedirs(cat_path, exist_ok=True)
                        cg_file = f"{cg['value']}_{safe(cg['text'])}.json"
                        file_path = os.path.join(cat_path, cg_file)
                        with open(file_path, "w", encoding="utf-8") as f:
                            _json.dump({"meta": {"acuerdo": ac["value"], "catalogo": cat["value"],
                                "categoria": cg["value"], "total": len(records)}, "records": records},
                                f, ensure_ascii=False, indent=2)
                        total += len(records)
                        log.ok(f"📊       +{len(records)} registros → {os.path.basename(cg_file)}")
                        index_data["combinaciones"].append({
                            "acuerdo": ac["value"], "acuerdo_label": ac["text"],
                            "catalogo": cat["value"], "catalogo_label": cat["text"],
                            "categoria": cg["value"], "categoria_label": cg["text"],
                            "records": len(records), "file": file_path,
                        })
                    except Exception as e:
                        log.warn(f"📊       Error: {e}")
                    time.sleep(0.3)
        index_data["total_records"] = total
        index_path = os.path.join(output_dir, "index.json")
        with open(index_path, "w", encoding="utf-8") as f:
            _json.dump(index_data, f, ensure_ascii=False, indent=2)
        log.ok(f"📊 Index: {index_path} ({total} registros)")
    except Exception as e:
        log.error(f"📊 Error fatal: {e}")
        import traceback; log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try: close_browser(pw, browser); log.info("📊 Navegador cerrado")
            except Exception: pass
        app.after(0, lambda: app.btn_extract.configure(state="normal"))
        app.after(0, lambda: app.btn_compare.configure(state="normal"))
        app.after(0, lambda: app.btn_discovery.configure(state="normal"))
        app.after(0, lambda: app.btn_discovery2.configure(state="normal"))
def execute_certs_only(app, usuario, password, headless, pre_selected):
    """Ejecuta SOLO la corrección de certificaciones ISO 9001/14001 para todas las fichas."""
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        from automation_mod.navegacion_productos import (
            GESTION_URL, guardar_cambios,
            leer_certificaciones_pagina, agregar_certificaciones_faltantes,
        )
        from automation_mod.bulk_subir_pdf import buscar_producto_api, URL_EDIT
        import time as _time
        log.info("🏅 Solo Certificaciones: iniciando navegador...")
        if stop.is_set():
            return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try:
                page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception:
                pass
        log.info("Navegador listo")
        if stop.is_set():
            return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok:
                log.error("Login fallido.")
            return
        log.ok("Login exitoso.")
        page.goto(GESTION_URL, wait_until="networkidle", timeout=60_000)
        _time.sleep(2)
        log.info(f"En {page.url}")
        total = len(app._excel_rows)
        ok_count = 0
        skip_count = 0
        err_count = 0
        for idx, row in enumerate(app._excel_rows):
            if stop.is_set():
                break
            parte = row.get("parte", "")
            ficha = row.get("ficha", "")
            certs_esp = row.get("certs_esperadas", [])
            log.info(f"🏅 [{idx+1}/{total}] {parte} (ficha={ficha})")
            # Navegar directo a la ficha
            estado = pre_selected.get("estado", "OBSERVADO")
            edit_url = f"{URL_EDIT}?ID_CatalogoProducto={ficha}&C_EstadoNav={estado}&C_Moneda=USD"
            try:
                page.goto(edit_url, wait_until="networkidle", timeout=60_000)
                _time.sleep(1.5)
            except Exception as e:
                log.warn(f"  Error navegando a ficha: {e}")
                err_count += 1
                continue
            # Verificar si ya tiene las ISO
            page_certs = leer_certificaciones_pagina(page)
            iso_presentes = {c["valor"].upper() for c in page_certs if "ISO" in c["valor"].upper()}
            iso_faltantes = [c for c in certs_esp if "ISO" in c.upper() and c.upper() not in iso_presentes]
            if not iso_faltantes:
                log.ok(f"  Ya tiene todas las ISO, saltando")
                skip_count += 1
                continue
            # Agregar certs faltantes
            try:
                result = agregar_certificaciones_faltantes(page, certs_esp, log, stop)
                added = result.get("added", [])
                if added:
                    log.ok(f"  ISO agregadas: {added}")
                    ok_count += 1
                else:
                    log.warn(f"  No se agregaron ISO (result={result})")
                    err_count += 1
            except Exception as e:
                log.warn(f"  Error en certs: {e}")
                err_count += 1
        log.ok(f"🏅 Completado: {ok_count} OK | {skip_count} ya tenían ISO | {err_count} errores")
    except Exception as e:
        log.error(f"Error fatal: {e}")
        import traceback
        log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try:
                close_browser(pw, browser)
                log.info("Navegador cerrado")
            except Exception:
                pass
        app.after(0, lambda: app.btn_certs.configure(state="normal"))
def execute_nro_parte(app, usuario, password, headless, pre_selected):
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        from automation_mod.navegacion_productos import (
            GESTION_URL, guardar_cambios, leer_caracteristicas_pagina,
            eliminar_caracteristica, agregar_caracteristica_texto,
        )
        from automation_mod.bulk_subir_pdf import URL_EDIT
        import time as _time
        log.info("🏷️ Iniciando navegador...")
        if stop.is_set(): return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass
        log.info("Navegador listo")
        if stop.is_set(): return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok: log.error("Login fallido.")
            return
        log.ok("Login exitoso. Iniciando N° de Parte...")
        total = len(app._excel_rows)
        ok_count = 0; err_count = 0
        estado = pre_selected.get("estado", "OBSERVADO")
        for idx, row in enumerate(app._excel_rows):
            if stop.is_set(): break
            parte = row.get("parte", "")
            ficha = row.get("ficha", "")
            log.info(f"🏷️ [{idx+1}/{total}] {parte} (ficha={ficha})")
            edit_url = f"{URL_EDIT}?ID_CatalogoProducto={ficha}&C_EstadoNav={estado}&C_Moneda=USD"
            try:
                page.goto(edit_url, wait_until="networkidle", timeout=60_000)
                _time.sleep(1.5)
                page_chars = leer_caracteristicas_pagina(page)
                nro_row = next((c for c in page_chars
                                if c.get("nombre", "").upper().strip() in ("N° DE PARTE", "NRO PARTE", "NRO_PARTE", "N° PARTE")), None)
                if nro_row and nro_row.get("id"):
                    eliminar_caracteristica(page, nro_row["id"], log, stop)
                    _time.sleep(1.5)
                agregar_caracteristica_texto(page, "NRO_PARTE", parte, log, stop)
                guardar_cambios(page, log, stop)
                ok_count += 1
                log.ok(f"  N° de Parte actualizado: {parte}")
            except Exception as e:
                log.warn(f"  Error: {e}")
                err_count += 1
        log.ok(f"🏷️ Completado: {ok_count} OK | {err_count} errores")
    except Exception as e:
        log.error(f"Error fatal: {e}")
        import traceback; log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try: close_browser(pw, browser); log.info("Navegador cerrado")
            except Exception: pass
        app.after(0, lambda: app.btn_nro.configure(state="normal"))
def execute_compare(app, usuario, password, headless, pre_selected):
    import re, json as _json
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        log.info("🔍 Iniciando navegador...")
        if stop.is_set(): return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass
        log.info("Navegador listo")
        if stop.is_set(): return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok: log.error("Login fallido.")
            return
        log.ok("Login exitoso.")
        acuerdo = pre_selected.get("acuerdo", "249")
        catalogo = pre_selected.get("catalogo", "252")
        categoria = pre_selected.get("categoria", "11735")
        estado = pre_selected.get("estado", "OBSERVADO")
        url = f"https://www.catalogos.perucompras.gob.pe/t_CatalogoProductoMarca?N_Acuerdo={acuerdo}&N_Catalogo={catalogo}&N_Categoria={categoria}&C_EstadoNav={estado}"
        log.info(f"🔍 URL: Acuerdo={acuerdo} Catálogo={catalogo} Categoría={categoria} Estado={estado}")
        # Navegar a la página objetivo (para tener cookies de sesión)
        page.goto("https://www.catalogos.perucompras.gob.pe/t_CatalogoProductoMarca", wait_until="networkidle", timeout=60_000)
        time.sleep(2)
        # Endpoint real (del navegador): GET con N_Catalogo, N_Categoria, C_EstadoNav, C_Descripcion
        # Devuelve HTML con tabla (NO JSON), parsear para extraer fichas
        api_url = f"https://www.catalogos.perucompras.gob.pe/t_CatalogoProductoMarca/_CatalogoProductoIndex?N_Catalogo={catalogo}&N_Categoria={categoria}&C_Descripcion=&C_EstadoNav={estado}&_={int(time.time()*1000)}"
        html = page.evaluate("""
            async (url) => {
                const resp = await fetch(url, {
                    method: 'GET',
                    headers: { 'X-Requested-With': 'XMLHttpRequest' }
                });
                return await resp.text();
            }
        """, api_url)
        # Parsear HTML: buscar <tr> con <input ID_CatalogoProducto> o links de edición
        # La estructura es: <tr><td><input name="ID_CatalogoProducto" value="2267958"/></td>...<td>OBSERVADO</td>...</tr>
        page_data = {}
        for tr in re.finditer(r'<tr\b[^>]*>(.*?)</tr>', html, re.DOTALL):
            row_html = tr.group(1)
            # Extraer ficha del input hidden
            id_m = re.search(r'ID_CatalogoProducto[^>]*value="(\d+)"', row_html)
            ficha_num = id_m.group(1) if id_m else None
            # Extraer todas las celdas para detectar estado
            tds = [re.sub(r'<[^>]+>', ' ', td).replace('&amp;', '&').strip()
                   for td in re.findall(r'<td\b[^>]*>(.*?)</td>', row_html, re.DOTALL)]
            tds = [re.sub(r'\s+', ' ', t).strip() for t in tds]
            # Si no se encontró ficha en el input, buscar en las celdas
            if not ficha_num:
                for cell in tds:
                    if re.fullmatch(r'\d{6,8}', cell):
                        ficha_num = cell; break
            # Detectar estado en las celdas
            estado_val = "DESCONOCIDO"
            for cell in tds:
                cu = cell.upper().strip()
                if cu in ("OBSERVADO", "ADJUDICADA", "EXCLUIDA", "OFERTADA",
                          "NO ADJUDICADA", "NO_ADJUDICADA", "DESIERTA", "CANCELADA",
                          "BAJA", "VIGENTE", "NO VIGENTE", "NO_VIGENTE"):
                    estado_val = cu; break
            if ficha_num:
                page_data[ficha_num] = estado_val
        # Si no se obtuvo nada del HTML, fallback al DOM de la página objetivo
        if not page_data:
            log.warn("GET no devolvió datos, fallback al DOM de la página")
            page.goto(url, wait_until="networkidle", timeout=60_000)
            time.sleep(5)
            page_data = page.evaluate("""
                () => {
                    const result = {};
                    const rows = document.querySelectorAll('table tbody tr, .dataTable tbody tr');
                    for (const row of rows) {
                        const idEl = row.querySelector('input[name="ID_CatalogoProducto"]');
                        const ficha = idEl ? idEl.value : null;
                        if (!ficha) continue;
                        const cells = Array.from(row.querySelectorAll('td')).map(c => c.textContent.trim());
                        let estado = 'DESCONOCIDO';
                        for (const c of cells) {
                            const cu = c.toUpperCase();
                            if (['OBSERVADO','ADJUDICADA','EXCLUIDA','OFERTADA','VIGENTE',
                                 'NO_ADJUDICADA','DESIERTA','CANCELADA','BAJA'].includes(cu)) {
                                estado = cu; break;
                            }
                        }
                        result[ficha] = estado;
                    }
                    return result;
                }
            """)
        log.info(f"🔍 Extraídas {len(page_data)} fichas de la página")
        # Comparar con el Excel: para cada ficha, navegar a la página de edición
        # y comparar TODAS las características con las del Excel
        from automation_mod.navegacion_productos import (
            leer_caracteristicas_pagina, comparar_caracteristicas,
        )
        from automation_mod.bulk_subir_pdf import URL_EDIT
        results = []
        ok = amarillo = rojo = 0
        for idx, row in enumerate(app._excel_rows):
            if stop.is_set(): break
            ficha = str(row.get("ficha", "")).strip()
            parte = row.get("parte", "")
            if not ficha:
                results.append({"index": idx, "status": "not_found"})
                rojo += 1
                continue
            if ficha not in page_data:
                results.append({"index": idx, "status": "not_found"})
                rojo += 1
                log.info(f"  ❌ {ficha}: no encontrada en la lista")
                continue
            # Navegar a la página de edición de la ficha para leer las características
            try:
                edit_url = f"{URL_EDIT}?ID_CatalogoProducto={ficha}&C_EstadoNav={estado}&C_Moneda=USD"
                page.goto(edit_url, wait_until="networkidle", timeout=60_000)
                time.sleep(1.5)
                page_chars = leer_caracteristicas_pagina(page)
                excel_chars = row.get("caracteristicas", {})
                if not excel_chars:
                    # Sin características en Excel: comparar solo existencia
                    results.append({"index": idx, "status": "ok"})
                    ok += 1
                    log.ok(f"  ✅ {ficha}: existe (sin chars en Excel)")
                    continue
                comp = comparar_caracteristicas(page_chars, excel_chars, log)
                n_iguales = comp["iguales"]
                n_diferentes = len(comp["diferentes"])
                n_excel = len(excel_chars)
                n_pagina = len(page_chars)
                # Si no hay diferencias en los chars que existen en AMBAS → verde.
                # Los chars del Excel que no están en la página se ignoran
                # (la página puede no tener todos los chars del Excel).
                if n_diferentes == 0 and n_iguales > 0:
                    results.append({"index": idx, "status": "ok"})
                    ok += 1
                    log.ok(f"  ✅ {ficha}: {n_iguales} chars match (de {n_excel} en Excel, {n_pagina} en página)")
                else:
                    results.append({"index": idx, "status": "differ"})
                    amarillo += 1
                    log.warn(f"  ⚠️ {ficha}: {n_iguales} match, {n_diferentes} difieren")
            except Exception as e:
                results.append({"index": idx, "status": "not_found"})
                rojo += 1
                log.warn(f"  ❌ {ficha}: error leyendo página: {e}")
        # Colorear Excel
        try:
            from utils_mod.excel_writer_mod import write_colored_results
            sheet = app.combo_sheet.get() if hasattr(self, "combo_sheet") else "Hoja1"
            out = write_colored_results(app._excel_path, sheet, results)
            log.ok(f"🔍 Excel coloreado: {os.path.basename(out)}")
            log.ok(f"🔍 Resultado: {ok} verde | {amarillo} amarillo | {rojo} rojo")
        except Exception as e:
            log.warn(f"No se pudo colorear: {e}")
    except Exception as e:
        log.error(f"Error fatal: {e}")
        import traceback; log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try: close_browser(pw, browser); log.info("Navegador cerrado")
            except Exception: pass
        app.after(0, lambda: app.btn_compare.configure(state="normal"))
        app.after(0, lambda: app.btn_discovery.configure(state="normal"))
        app.after(0, lambda: app.btn_discovery2.configure(state="normal"))
def execute_discovery(app, usuario, password, headless):
    """Ejecuta el discovery de endpoints directamente en la app (mismo browser)."""
    import re as _re
    from bs4 import BeautifulSoup
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        log.info("🕵️ Iniciando navegador...")
        if stop.is_set(): return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass
        log.info("Navegador listo")
        if stop.is_set(): return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok: log.error("Login fallido.")
            return
        log.ok("Login exitoso.")
        # Importar funciones del discovery
        from urllib.parse import urljoin
        sys.path.insert(0, _THIS_DIR)
        try:
            import discovery_perucompras as disc
        except Exception as e:
            log.error(f"No se pudo importar discovery_perucompras: {e}")
            return
        output_dir = os.path.join(_THIS_DIR, "discovery_output")
        os.makedirs(output_dir, exist_ok=True)
        # Capa 1: JS estático
        log.info("🕵️ Capa 1: Analizando JS estático...")
        js_info = disc.analizar_js(page)
        with open(os.path.join(output_dir, "js_analisis.json"), "w", encoding="utf-8") as f:
            json.dump(js_info, f, ensure_ascii=False, indent=2)
        log.info(f"  → {js_info['total']} rutas únicas encontradas en JS")
        # Combinar candidatos
        todos_candidatos = sorted(set(disc.CANDIDATOS + js_info["rutas_descubiertas"]))
        log.info(f"🕵️ Capa 2+3: Fuzzing {len(todos_candidatos)} endpoints...")
        # Fuzzing con clasificación y reporte en vivo
        resultados = {
            "accessible": [], "requires_auth": [], "forbidden": [],
            "error_500": [], "redirect": [], "not_found": [],
        }
        # Función JS para hacer fetch desde la página (con cookies de sesión)
        def _fetch(url):
            return page.evaluate("""
                async ([u]) => {
                    try {
                        const r = await fetch(u, { method: 'GET', redirect: 'manual',
                            headers: { 'X-Requested-With': 'XMLHttpRequest' } });
                        return { ok: true, status: r.status, url: r.url,
                                 type: r.type, text: (await r.text()).substring(0, 2000) };
                    } catch (e) {
                        return { ok: false, error: String(e) };
                    }
                }
            """, [url])
        for i, path in enumerate(todos_candidatos, 1):
            if stop.is_set(): break
            url = disc.BASE + path
            try:
                resp = _fetch(url)
                if not resp.get("ok"):
                    # Fallback: usar page.request con timeout en milisegundos (30s)
                    try:
                        r = page.context.request.get(url, headers=disc.HEADERS,
                                                      timeout=30_000, max_redirects=0)
                        resp = {"ok": True, "status": r.status, "url": str(r.url),
                                "text": r.text()[:2000] if r.status == 200 else "",
                                "size": len(r.body())}
                    except Exception as e2:
                        log.warn(f"  [{i}/{len(todos_candidatos)}] ✗ {path} → {e2}")
                        continue
                status = resp.get("status", 0)
                url_final = resp.get("url", url)
                html_text = resp.get("text", "")
                cat = disc.clasificar_respuesta(status, url_final, html_text)
                entry = {
                    "path": path, "url": url, "status": status,
                    "url_final": url_final,
                    "size_bytes": resp.get("size", len(html_text)),
                    "categoria": cat,
                }
                if cat == "accessible":
                    soup = BeautifulSoup(html_text, "lxml")
                    title = soup.find("title")
                    h5 = soup.find("h5")
                    entry["titulo"] = (title.get_text(strip=True) if title
                                       else h5.get_text(strip=True) if h5 else "")
                    sub = disc.extract_urls_from_html(html_text)
                    entry["sub_rutas"] = sorted([s for s in sub
                        if not any(ext in s for ext in ['.js', '.css', '.png'])])
                    log.ok(f"  [{i}/{len(todos_candidatos)}] ✅ {path} → {entry['titulo'][:40]}")
                elif cat in ("requires_auth", "forbidden", "error_500"):
                    log.warn(f"  [{i}/{len(todos_candidatos)}] 🔒 {path} → HTTP {status} ({cat})")
                else:
                    if i % 20 == 0:
                        log.info(f"  [{i}/{len(todos_candidatos)}] ... procesando ...")
                resultados[cat].append(entry)
            except Exception as e:
                log.warn(f"  [{i}/{len(todos_candidatos)}] ✗ {path} → {e}")
            time.sleep(disc.DELAY)
        # Capa 4: crawling
        log.info(f"🕵️ Capa 4: Crawling de {len(resultados['accessible'])} páginas accesibles...")
        visitados = set(todos_candidatos)
        nuevas = set()
        for entry in resultados["accessible"]:
            for sub in entry.get("sub_rutas", []):
                if sub not in visitados and sub.startswith("/"):
                    nuevas.add(sub)
        nuevas -= visitados
        log.info(f"  {len(nuevas)} rutas nuevas a probar")
        for j, path in enumerate(list(nuevas), 1):
            if stop.is_set(): break
            url = disc.BASE + path
            try:
                resp = _fetch(url)
                if resp.get("ok"):
                    status = resp.get("status", 0)
                    cat = disc.clasificar_respuesta(status, resp.get("url", url),
                                                    resp.get("text", ""))
                    if cat == "accessible":
                        log.ok(f"  [{j}/{len(nuevas)}] ✅ {path} (crawl)")
                    elif cat in ("requires_auth", "forbidden", "error_500"):
                        log.warn(f"  [{j}/{len(nuevas)}] 🔒 {path} (crawl)")
                    resultados[cat].append({
                        "path": path, "url": url, "status": status, "categoria": cat
                    })
            except Exception as e:
                log.warn(f"  [{j}/{len(nuevas)}] ✗ {path} → {e}")
            time.sleep(disc.DELAY)
        # Guardar resultados
        with open(os.path.join(output_dir, "endpoints_accesibles.json"), "w", encoding="utf-8") as f:
            json.dump(resultados["accessible"], f, ensure_ascii=False, indent=2)
        with open(os.path.join(output_dir, "endpoints_otro_rol.json"), "w", encoding="utf-8") as f:
            json.dump(resultados["requires_auth"] + resultados["forbidden"],
                      f, ensure_ascii=False, indent=2)
        with open(os.path.join(output_dir, "endpoints_error.json"), "w", encoding="utf-8") as f:
            json.dump(resultados["error_500"], f, ensure_ascii=False, indent=2)
        with open(os.path.join(output_dir, "endpoints_todos.json"), "w", encoding="utf-8") as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2)
        disc.generar_reporte(js_info, resultados, Path(output_dir))
        log.ok(f"🕵️ Completado: {len(resultados['accessible'])} ✅ | "
               f"{len(resultados['requires_auth'])+len(resultados['forbidden'])} 🔒 | "
               f"{len(resultados['error_500'])} 💥")
        log.ok(f"🕵️ Salida en: {output_dir}/")
    except Exception as e:
        log.error(f"Error fatal: {e}")
        import traceback; log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try: close_browser(pw, browser); log.info("Navegador cerrado")
            except Exception: pass
        app.after(0, lambda: app.btn_discovery.configure(state="normal"))
        app.after(0, lambda: app.btn_discovery2.configure(state="normal"))
def execute_discovery2(app, usuario, password, headless):
    """Ejecuta discovery_v2_perucompras.py: 8 técnicas de scraping profundo."""
    from bs4 import BeautifulSoup
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        log.info("🕵️v2 Iniciando navegador...")
        if stop.is_set(): return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass
        log.info("Navegador listo")
        if stop.is_set(): return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok: log.error("Login fallido.")
            return
        log.ok("Login exitoso.")
        sys.path.insert(0, _THIS_DIR)
        import importlib
        try:
            import discovery_v2_perucompras as disc2
            importlib.reload(disc2)
        except Exception as e:
            log.error(f"No se pudo importar discovery_v2_perucompras: {e}")
            return
        disc2.USUARIO = usuario
        disc2.PASSWORD = password
        output_dir = os.path.join(_THIS_DIR, "discovery_v2_output")
        os.makedirs(output_dir, exist_ok=True)
        all_results = {}
        # T1: Recon files
        log.info("🕵️v2 T1: Archivos de reconocimiento...")
        all_results["recon"] = disc2.recon_files(page)
        # T2: JS profundo
        log.info("🕵️v2 T2: Análisis JS profundo...")
        js_rutas, js_detalles = disc2.analizar_js_profundo(page)
        all_results["js_rutas"]    = js_rutas
        all_results["js_detalles"] = js_detalles
        # T3: Enumerar acciones
        log.info(f"🕵️v2 T3: {len(disc2.CONTROLADORES_BASE)}×{len(disc2.ACCIONES_MVC)} acciones...")
        all_results["acciones"] = disc2.enumerar_acciones(page)
        # T4: POST
        log.info(f"🕵️v2 T4: POST a {len(disc2.POST_CANDIDATOS)} endpoints...")
        all_results["post_results"] = disc2.probar_post(page)
        # T5: Brute IDs
        log.info(f"🕵️v2 T5: IDs en /Reportes/Index/N...")
        all_results["brute_ids"] = disc2.bruteforce_ids(page)
        # T6: Headers
        log.info("🕵️v2 T6: Headers de tecnología...")
        all_results["tech_headers"] = disc2.analizar_headers_tecnologia(page)
        # T7: Crawling
        seeds = [r["path"] for r in (all_results["recon"] + all_results["acciones"]
                                      + all_results["post_results"])
                 if r.get("categoria") == "accessible"]
        seeds += ["/Home", "/t_CatalogoProductoMarca", "/ConsultaValoresCreados",
                  "/ProformaSinOrdenCompra"]
        seeds = list(dict.fromkeys(seeds))
        log.info(f"🕵️v2 T7: Crawling recursivo ({len(seeds)} seeds)...")
        all_results["crawl"] = disc2.crawl_recursivo(page, seeds, max_depth=2)
        # T8: Rutas JS nuevas
        ya_probadas = set(r["path"] for results in all_results.values()
                           if isinstance(results, list) for r in results
                           if isinstance(r, dict) and "path" in r)
        js_nuevas = [p for p in js_rutas if p not in ya_probadas and p.startswith("/")]
        if js_nuevas:
            log.info(f"🕵️v2 T8: {len(js_nuevas)} rutas nuevas del JS...")
            js_probe_results = []
            for path in js_nuevas:
                if stop.is_set(): break
                r = disc2.probe(page, path)
                if r["categoria"] not in ("not_found", "silently_redirected"):
                    log.ok(f"  {'✅' if r['status']==200 else '⚠'} {path} → {r['status']}")
                    js_probe_results.append(r)
                time.sleep(disc2.DELAY)
            all_results["js_probed"] = js_probe_results
        # Guardar
        for key, data in all_results.items():
            path = os.path.join(output_dir, f"{key}.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        disc2.generar_reporte_v2(all_results, Path(output_dir))
        total = sum(
            1 for items in all_results.values()
            if isinstance(items, list)
            for it in items
            if isinstance(it, dict) and "categoria" in it
        )
        log.ok(f"🕵️v2 Completado: {total} endpoints no-404")
        log.ok(f"🕵️v2 Salida: {output_dir}/")
    except Exception as e:
        log.error(f"Error fatal: {e}")
        import traceback; log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try: close_browser(pw, browser); log.info("Navegador cerrado")
            except Exception: pass
        app.after(0, lambda: app.btn_discovery2.configure(state="normal"))
def execute_test(app, usuario, password, headless, pre_selected):
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        log.info("🧪 Test Flow: iniciando navegador...")
        if stop.is_set(): return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try:
                page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception:
                pass
        log.info("🧪 Navegador listo (viewport 1920x1080)")
        if stop.is_set(): return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok: log.error("🧪 Login fallido.")
            return
        log.ok("🧪 Login exitoso")
        from automation_mod.navegacion_productos import (
            GESTION_URL,
            subir_pdf_en_edicion, subir_imagen_en_edicion, cambiar_precio_en_edicion,
            guardar_cambios,
            ensure_logged_in_and_ready,
            leer_caracteristicas_pagina, leer_certificaciones_pagina,
            comparar_caracteristicas, corregir_caracteristica,
            agregar_certificaciones_faltantes,
            eliminar_caracteristica, agregar_caracteristica_texto,
        )
        from automation_mod.bulk_subir_pdf import buscar_producto_api, URL_EDIT
        log.info("🧪 Navegando a t_CatalogoProductoMarca...")
        try:
            page.goto(GESTION_URL, wait_until="networkidle", timeout=60_000)
        except Exception as e:
            log.warn(f"🧪 Timeout en goto ({e})")
        # NO aplicamos dropdowns+Buscar (la tabla con miles de fichas tarda 5+ min).
        # Usamos API search por N° de Parte y navegamos directo a la ficha.
        if not app._excel_rows:
            log.warn("🧪 No hay filas en el Excel")
            return
        # 2. Iterar productos (API search + navegación directa — sin tabla)
        pdf_dir = r"D:\SISTEMAS 02\Downloads\COMPUTADORAS\COMPUTADORAS"
        stats = {"ok": 0, "not_found": 0, "failed": 0}
        all_results = []
        catalogo_v = pre_selected.get("catalogo", "252")
        categoria_v = pre_selected.get("categoria", "11735")
        estado_v = pre_selected.get("estado", "OBSERVADO")
        # Test Flow: solo 1 ficha (la primera)
        for idx, row in enumerate(app._excel_rows[:1]):
            if stop.is_set():
                break
            parte = row["parte"]
            ruta_pdf = os.path.join(pdf_dir, f"{parte}.pdf")
            log.info(f"🧪 [TEST 1/{len(app._excel_rows)}] Procesando: {parte}")
            # Verificar sesión antes de procesar
            if not ensure_logged_in_and_ready(
                page, usuario, password, pre_selected, log, stop, app.captcha_bridge
            ):
                log.error(f"🧪 [{idx+1}] No se pudo restablecer sesión, deteniendo")
                all_results.append({"index": idx, "parte": parte, "status": "session_lost"})
                break
            # ── Buscar vía API (sin tabla) ──
            ficha = row.get("ficha", "")
            prod = buscar_producto_api(page, parte, catalogo_v, categoria_v, estado_v, log, ficha=ficha)
            if not prod:
                stats["not_found"] += 1
                app.log_queue.put({"type": "log", "msg": f"🧪 [{idx+1}/{len(app._excel_rows)}] ? {parte} — No encontrado", "level": "notfound"})
                all_results.append({"index": idx, "parte": parte, "status": "not_found"})
                continue
            # ── Navegar directo a la ficha ──
            edit_url = f"{URL_EDIT}?ID_CatalogoProducto={prod['id']}&C_EstadoNav={estado_v}&C_Moneda=USD"
            try:
                page.goto(edit_url, wait_until="networkidle", timeout=60_000)
            except Exception as e:
                log.warn(f"🧪 Timeout navegando a edit: {e}")
            time.sleep(1.5)
            # ── Subir PDF ──
            if not subir_pdf_en_edicion(page, ruta_pdf, log, stop):
                stats["failed"] += 1
                app.log_queue.put({"type": "log", "msg": f"🧪 [TEST 1/{len(app._excel_rows)}] ✗ {parte} — PDF no subido", "level": "error"})
                all_results.append({"index": idx, "parte": parte, "status": "pdf_failed"})
                continue
            # ── Subir imagen (columna IMAGEN (PDF) del Excel) ──
            nombre_imagen = row.get("imagen", "")
            if nombre_imagen:
                subir_imagen_en_edicion(page, str(nombre_imagen), log, stop)
            else:
                log.info("  Sin imagen en el Excel, saltando")
            # ── Cambiar precio (columna PRECIO SUGERIDO del Excel) ──
            precio = row.get("precio", "")
            if precio:
                cambiar_precio_en_edicion(page, str(precio), log, stop)
            else:
                log.info("  Sin precio en el Excel, saltando")
            # ── Guardar ──
            if not guardar_cambios(page, log, stop):
                stats["failed"] += 1
                app.log_queue.put({"type": "log", "msg": f"🧪 [TEST 1/{len(app._excel_rows)}] ✗ {parte} — Error al guardar", "level": "error"})
                all_results.append({"index": idx, "parte": parte, "status": "save_failed"})
                continue
            # Esperar a que la página se estabilice tras el guardado
            try:
                page.wait_for_load_state("networkidle", timeout=30_000)
            except Exception:
                pass
            time.sleep(2)
            # ── Comparar/Correr características contra Excel ──
            chars_ok = 0
            chars_dif = 0
            chars_corr = 0
            try:
                page_chars = leer_caracteristicas_pagina(page)
                excel_chars = row.get("caracteristicas", {})
                comp = comparar_caracteristicas(page_chars, excel_chars, log)
                chars_ok = comp["iguales"]
                chars_dif = len(comp["diferentes"])
                for d in comp["diferentes"]:
                    if stop.is_set():
                        break
                    if corregir_caracteristica(page, d["id"], d["esperado"], log, stop, edit_url=edit_url):
                        chars_corr += 1
            except Exception as e:
                log.warn(f"🧪 Error comparando chars: {e}")
            # ── Agregar certificaciones faltantes ──
            try:
                certs_esp = row.get("certs_esperadas", [])
                if certs_esp:
                    agregar_certificaciones_faltantes(page, certs_esp, log, stop)
            except Exception as e:
                log.warn(f"🧪 Error en certs: {e}")
            # ── Actualizar N° de Parte: si existe eliminar y agregar, si no solo agregar ──
            try:
                page.goto(edit_url, wait_until="networkidle", timeout=60_000)
                time.sleep(1.5)
                page_chars_full = leer_caracteristicas_pagina(page)
                nro_parte_row = next((c for c in page_chars_full
                                      if c.get("nombre", "").upper().strip() in ("N° DE PARTE", "NRO PARTE", "NRO_PARTE", "N° PARTE")), None)
                if nro_parte_row and nro_parte_row.get("id"):
                    # Existe: eliminar y agregar nuevo
                    log.info(f"🧪 N° de Parte existente encontrado, eliminando...")
                    eliminar_caracteristica(page, nro_parte_row["id"], log, stop)
                    time.sleep(2)
                else:
                    # No existe: solo agregar nuevo
                    log.info(f"🧪 N° de Parte no existe, agregando nuevo...")
                agregar_caracteristica_texto(page, "NRO_PARTE", parte, log, stop)
                guardar_cambios(page, log, stop)
            except Exception as e:
                log.warn(f"🧪 Error actualizando N° de Parte: {e}")
            stats["ok"] += 1
            app.log_queue.put({"type": "log", "msg": f"🧪 [{idx+1}/{len(app._excel_rows)}] OK {parte} — {chars_ok} ok, {chars_dif} dif, {chars_corr} corregidas", "level": "complete"})
            all_results.append({"index": idx, "parte": parte, "status": "ok", "chars_ok": chars_ok, "chars_dif": chars_dif, "chars_corr": chars_corr})
        # Resumen final
        log.info(f"🧪 {'='*40}")
        log.info(f"🧪 RESUMEN: {stats['ok']} OK | {stats['not_found']} No encontrados | {stats['failed']} Fallos")
        # Colorear Excel
        if all_results and app._excel_path:
            from utils_mod.excel_writer_mod import write_colored_results
            try:
                sheet = app.combo_sheet.get() if hasattr(self, 'combo_sheet') else "Hoja1"
                out = write_colored_results(app._excel_path, sheet, all_results)
                log.ok(f"🧪 Excel coloreado guardado: {os.path.basename(out)}")
            except Exception as e:
                log.warn(f"🧪 No se pudo colorear el Excel: {e}")
        log.info(f"🧪 {'='*40}")
        if stats["ok"] > 0:
            log.ok(f"🧪 Procesamiento completado exitosamente")
    except Exception as e:
        log.error(f"🧪 Error en test: {e}")
        import traceback
        log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try:
                close_browser(pw, browser)
                log.info("🧪 Navegador cerrado")
            except Exception:
                pass
        app.log_queue.put({"type": "done", "ok": 0, "errors": 0})
        app.after(0, lambda: app.btn_test.configure(state="normal"))
        app.after(0, lambda: app.btn_certs.configure(state="normal"))
# ── Launch ─────────────────────────────────────────────────
def execute(app, usuario, password, headless, rows, pausa, pre_selected=None):
    log = LogWriter(app.log_queue)
    stop = app.stop_event
    pw = browser = None
    try:
        log.info("Iniciando navegador...")
        if stop.is_set():
            return
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try:
                page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception:
                pass
        log.info("Navegador listo")
        if stop.is_set():
            return
        ok = do_login(page, usuario, password, "", log, stop, app.captcha_bridge)
        if not ok or stop.is_set():
            if not ok:
                log.error("Login fallido.")
            return
        log.ok("Login exitoso. Iniciando procesamiento masivo...")
        from automation_mod.bulk_subir_pdf import run_bulk_subir_pdf, URL_MANAGEMENT
        import time as _time
        log.info("Navegando al catálogo de fichas...")
        try:
            page.goto(URL_MANAGEMENT, wait_until="networkidle", timeout=60_000)
            _time.sleep(2)
        except Exception:
            pass
        app._total = len(rows)
        results = run_bulk_subir_pdf(
            page, rows, pre_selected or {}, log, stop, app.captcha_bridge,
            usuario=usuario, password=password,
        )
        all_results = results
        ok_count = sum(1 for r in results if r["status"] == "ok")
        warn_count = sum(1 for r in results if r["status"] in ("not_found",))
        err_count = len(results) - ok_count - warn_count
        # Guardar Excel coloreado
        if all_results and app._excel_path:
            from utils_mod.excel_writer_mod import write_colored_results
            try:
                sheet = app.combo_sheet.get() if hasattr(self, 'combo_sheet') else "Hoja1"
                out = write_colored_results(app._excel_path, sheet, all_results)
                log.ok(f"Excel coloreado guardado: {os.path.basename(out)}")
            except Exception as e:
                log.warn(f"No se pudo colorear el Excel: {e}")
        app.log_queue.put({"type": "stat_ok", "value": ok_count})
        app.log_queue.put({"type": "stat_warn", "value": warn_count})
        app.log_queue.put({"type": "stat_err", "value": err_count})
        log.ok(f"Completado: {ok_count} OK | {warn_count} No encontrados | {err_count} Errores")
    except Exception as e:
        log.error(f"Error fatal: {e}")
        import traceback
        log.error(traceback.format_exc())
    finally:
        if browser and pw:
            try:
                close_browser(pw, browser)
                log.info("Navegador cerrado")
            except Exception:
                pass


# ═══════════════════════════════════════════════════════════════════
#  TRABAJADORES DE LA PESTAÑA: SUBIR PRECIOS JSON
# ═══════════════════════════════════════════════════════════════════

def read_select_options_precios(page, selector):
    try:
        opts = page.evaluate(f"""
            () => {{
                const sel = document.querySelector('{selector}');
                if (!sel) return [];
                return Array.from(sel.options)
                    .map(o => ({{value: o.value, text: o.text}}))
                    .filter(o => o.value && o.value !== '0' && o.text.trim());
            }}
        """)
        return opts
    except Exception:
        return []

def wait_for_options_precios(page, selector, timeout=20):
    try:
        page.wait_for_function(
            f"""() => {{
                const sel = document.querySelector('{selector}');
                if (!sel) return false;
                return Array.from(sel.options).filter(o => o.value && o.value !== '0').length >= 1;
            }}""",
            timeout=timeout * 1000,
        )
        return True
    except Exception:
        return False

def execute_extraer_menu_precios(app, usuario, password, headless, log_func):
    from automation.browser import init_browser, close_browser
    from automation.login import do_login
    import time
    import json
    import os
    
    pw = browser = None
    try:
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass
            
        log_func(app, "✅ Navegador listo, iniciando login...")
        
        # dummy log/stop
        class DummyLog:
            def info(self, m): pass
            def ok(self, m): pass
            def warn(self, m): pass
            def error(self, m): pass
        
        stop = type('S', (), {'is_set': lambda self: False})()
        
        if not do_login(page, usuario, password, "", DummyLog(), stop, app.captcha_bridge):
            log_func(app, "❌ Falló el login.")
            return
            
        log_func(app, "✅ Login exitoso. Navegando a t_ProductoOfertadoAmp...")
        
        url = "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp"
        page.goto(url, wait_until="networkidle", timeout=60000)
        time.sleep(3)
        
        log_func(app, "🔄 Leyendo Acuerdos...")
        acuerdos = read_select_options_precios(page, "#ajaxAcuerdo")
        
        combinaciones = []
        for ac_idx, acuerdo in enumerate(acuerdos):
            log_func(app, f"▶ Acuerdo: {acuerdo['text'][:50]}")
            page.select_option("#ajaxAcuerdo", value=acuerdo["value"])
            time.sleep(2)
            wait_for_options_precios(page, "#ajaxCatalogo", timeout=15)
            
            catalogos = read_select_options_precios(page, "#ajaxCatalogo")
            for cat in catalogos:
                page.select_option("#ajaxCatalogo", value=cat["value"])
                time.sleep(2)
                wait_for_options_precios(page, "#ajaxCategoria", timeout=15)
                
                categorias = read_select_options_precios(page, "#ajaxCategoria")
                
                combinaciones.append({
                    "acuerdo": acuerdo,
                    "catalogo": cat,
                    "categorias": categorias
                })
        
        import utils_mod.config_helper as ch
        output_file = ch.get_writable_path("dropdowns_precios.json", os.path.dirname(__file__))
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(combinaciones, f, ensure_ascii=False, indent=2)
            
        log_func(app, f"✅ Extracción completada. Guardado en dropdowns_precios.json")
        
        # Actualizar UI
        def update_ui():
            app._precios_combinaciones = combinaciones
            acuerdos_str = [f"{c['acuerdo']['value']} - {c['acuerdo']['text']}" for c in combinaciones]
            # Eliminar duplicados manteniendo orden
            acuerdos_str = list(dict.fromkeys(acuerdos_str))
            
            if acuerdos_str:
                app.option_precio_acuerdo.configure(values=acuerdos_str)
                app.option_precio_acuerdo.set(acuerdos_str[0])
                try:
                    import tab_precios_json
                    tab_precios_json._on_precio_acuerdo_changed(app, acuerdos_str[0])
                except Exception:
                    pass
            
            app.btn_iniciar_precios.configure(state="normal")
            
        app.after(0, update_ui)
        
    except Exception as e:
        log_func(app, f"❌ Error fatal: {e}")
    finally:
        if browser and pw:
            try: close_browser(pw, browser)
            except: pass
        app.after(0, lambda: app.btn_iniciar_precios.configure(state="normal"))


# ═══════════════════════════════════════════════════════════════════
#  WORKER TEST: SUBIR PRECIOS (1 producto - modo dry-run)
# ═══════════════════════════════════════════════════════════════════

def _calcular_precio_dolar(precio_max, ganancia=0.10, tc=3.4):
    """(precio_max * 1.10) / 3.4 → precio en USD redondeado a 2 decimales"""
    return round((precio_max * (1 + ganancia)) / tc, 2)


def _buscar_match_local(producto_pc, precios_data):
    """
    Busca en precios_data el registro que coincida con el producto de Perú Compras.
    Estrategia 1: comparar C_Descripcion (PC) contra descripcin_fichaproducto (Local).
    Estrategia 2: buscar nro_parte local dentro de la C_Descripcion de Perú Compras.
    """
    desc_pc = (producto_pc.get("C_Descripcion") or "").upper().strip()

    for rec in precios_data:
        desc_local = (rec.get("descripcin_fichaproducto") or "").upper().strip()
        nro_local  = (rec.get("nro_parte_o_cdigo_nico_de_identificacin") or "").upper().strip()

        # Match exacto por descripción
        if desc_local and desc_pc and desc_local == desc_pc:
            return rec

        # Match si el número de parte local está dentro de la descripción de Perú Compras
        if nro_local and desc_pc and nro_local in desc_pc:
            return rec

    return None


def execute_test_precios(app, usuario, password, headless, log_func,
                         precios_data, acuerdo_val, catalogo_val, categoria_val):
    """
    TEST: navega a t_ProductoOfertadoAmp, aplica filtros, descarga TODOS los
    productos mediante peticiones POST directas usando el payload interceptado
    y las cookies de sesión del navegador (mucho más rápido), hace matching
    y muestra qué precio se insertaría. NO escribe nada todavía.
    """
    from automation.browser import init_browser, close_browser
    from automation.login import do_login
    from urllib.parse import parse_qs, urlencode
    import time

    pw = browser = None

    def re_enable():
        app.after(0, lambda: app.btn_test_precios.configure(state="normal"))
        app.after(0, lambda: app.btn_iniciar_precios.configure(state="normal"))

    try:
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass

        log_func(app, "✅ Navegador listo, iniciando login...")

        class DummyLog:
            def info(self, m): pass
            def ok(self, m): pass
            def warn(self, m): pass
            def error(self, m): pass

        stop = type('S', (), {'is_set': lambda self: False})()

        if not do_login(page, usuario, password, "", DummyLog(), stop, app.captcha_bridge):
            log_func(app, "❌ Falló el login.")
            re_enable()
            return

        log_func(app, "✅ Login OK. Navegando a t_ProductoOfertadoAmp...")
        page.goto(
            "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp",
            wait_until="networkidle", timeout=60_000
        )
        time.sleep(2)

        # ── 1. Seleccionar Acuerdo ──────────────────────────────────────────
        log_func(app, "⏳ Esperando a que el Acuerdo esté disponible...")
        wait_for_options_precios(page, '#ajaxAcuerdo', timeout=20)
        log_func(app, f"🔧 Seleccionando Acuerdo [{acuerdo_val}]...")
        page.evaluate(f"""
            () => {{
                const sel = document.querySelector('#ajaxAcuerdo');
                if (sel) {{
                    sel.value = '{acuerdo_val}';
                    sel.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }}
        """)

        # ── 2. Seleccionar Catálogo ────────────────────────────────────────
        log_func(app, "⏳ Esperando a que carguen los Catálogos...")
        wait_for_options_precios(page, '#ajaxCatalogo', timeout=25)
        log_func(app, f"🔧 Seleccionando Catálogo [{catalogo_val}]...")
        page.evaluate(f"""
            () => {{
                const sel = document.querySelector('#ajaxCatalogo');
                if (sel) {{
                    sel.value = '{catalogo_val}';
                    sel.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }}
        """)

        # ── 3. Seleccionar Categoría ───────────────────────────────────────
        log_func(app, "⏳ Esperando a que carguen las Categorías...")
        wait_for_options_precios(page, '#ajaxCategoria', timeout=25)
        log_func(app, f"🔧 Seleccionando Categoría [{categoria_val}]...")
        page.evaluate(f"""
            () => {{
                const sel = document.querySelector('#ajaxCategoria');
                if (sel) {{
                    sel.value = '{categoria_val}';
                    sel.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }}
        """)
        time.sleep(2)

        # ── 4. Click Iniciar Búsqueda y esperar tabla ─────────────────────
        log_func(app, "🔍 Haciendo click en Iniciar Búsqueda...")
        page.click("#btnBuscar")
        time.sleep(4)

        try:
            page.wait_for_selector("#btnNuevoProducto", timeout=10_000)
            log_func(app, "✅ Botón 'Agregar oferta' visible.")
        except Exception:
            log_func(app, "⚠ Botón 'Agregar oferta' no apareció, continuando...")

        # ── 5. Interceptar payload del primer request AJAX ─────────────────
        captured_payload = {}
        total_records = [0]

        def handle_request(req):
            if "_CatalogoProductoIndexJson" in req.url:
                captured_payload["body"] = req.post_data
                captured_payload["url"] = req.url

        def handle_response(response):
            if "_CatalogoProductoIndexJson" in response.url:
                try:
                    data = response.json()
                    total_records[0] = data.get("recordsTotal", 0)
                except Exception:
                    pass

        page.on("request", handle_request)
        page.on("response", handle_response)

        try:
            log_func(app, "📂 Navegando a CatalogoProductoIndex...")
            with page.expect_response(lambda r: "_CatalogoProductoIndexJson" in r.url, timeout=45000):
                page.goto(
                    "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp/CatalogoProductoIndex",
                    wait_until="domcontentloaded", timeout=60_000
                )
        except Exception as e:
            log_func(app, f"⚠ Error esperando respuesta de red: {e}")

        page.remove_listener("request", handle_request)
        page.remove_listener("response", handle_response)

        raw_body = captured_payload.get("body")
        url_endpoint = captured_payload.get("url")

        if not raw_body or not url_endpoint:
            log_func(app, "❌ Error: No se pudo capturar el payload de red. Abortando.")
            re_enable()
            return

        # ── 6. Descarga mediante peticiones directas API (Velocidad Máxima) ─
        log_func(app, "📡 Descargando todos los registros mediante llamadas API directas...")
        parsed_params = parse_qs(raw_body)
        total = total_records[0]
        log_func(app, f"📊 Total registros en Perú Compras: {total}")

        all_products = []
        PAGE_SIZE = 100  # Lote estable
        start = 0
        draw = 1

        while start < total:
            # Modificar parámetros de paginación
            parsed_params['start'] = [str(start)]
            parsed_params['length'] = [str(PAGE_SIZE)]
            parsed_params['draw'] = [str(draw)]
            
            # Re-codificar a urlencoded
            new_body = urlencode(parsed_params, doseq=True)

            # Petición POST directa usando el cliente de Playwright (comparte sesión/cookies)
            resp = page.request.post(
                url_endpoint,
                headers={
                    "Content-Type": "application/x-www-form-urlencoded",
                    "X-Requested-With": "XMLHttpRequest",
                    "Referer": "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp/CatalogoProductoIndex"
                },
                data=new_body,
                timeout=20000
            )

            if resp.status == 200:
                try:
                    data = resp.json()
                    records = data.get("data", [])
                    all_products.extend(records)
                    log_func(app, f"   Descargados: {len(all_products)}/{total}")
                except Exception as je:
                    log_func(app, f"   ❌ Error parseando JSON en lote {start}: {je}")
                    break
            else:
                log_func(app, f"   ❌ Error de red en lote {start}: Status {resp.status}")
                break

            start += PAGE_SIZE
            draw += 1
            time.sleep(0.3)  # Pausa mínima de seguridad

        log_func(app, f"\n✅ Descarga completa: {len(all_products)} productos")

        # ── 7. Matching con JSON local ─────────────────────────────────────
        log_func(app, f"\n🔎 Buscando matches contra {len(precios_data)} fichas locales...")

        matches_found = 0
        no_match = 0
        sample_matches = []

        for producto_pc in all_products:
            match = _buscar_match_local(producto_pc, precios_data)
            if match:
                matches_found += 1
                precio_max = match.get("precio_max", 0)
                precio_usd = _calcular_precio_dolar(precio_max)
                if len(sample_matches) < 3:
                    sample_matches.append({
                        "pc_desc": (producto_pc.get("C_Descripcion") or "")[:60],
                        "precio_max": precio_max,
                        "precio_usd": precio_usd,
                    })
            else:
                no_match += 1

        log_func(app, f"✅ Matches encontrados: {matches_found}/{len(all_products)}")
        log_func(app, f"⚠ Sin match:           {no_match}/{len(all_products)}")

        if sample_matches:
            log_func(app, "\n📋 Muestra de los primeros 3 matches:")
            for s in sample_matches:
                log_func(app, f"   [{s['pc_desc']}]")
                log_func(app, f"   precio_max={s['precio_max']} S/ → {s['precio_usd']} USD")
                log_func(app, "   ---")

        log_func(app, "\n✅ TEST completado. Revisa los logs.")

    except Exception as e:
        import traceback
        log_func(app, f"❌ Error fatal: {e}")
        log_func(app, traceback.format_exc())
    finally:
        if browser and pw:
            try: close_browser(pw, browser)
            except: pass
        re_enable()



# ═══════════════════════════════════════════════════════════════════
#  TRABAJADORES DE SUBIDA DE PRECIOS REAL
# ═══════════════════════════════════════════════════════════════════

def _buscar_match_pc(rec_local, all_products):
    nro_local = (rec_local.get("nro_parte_o_cdigo_nico_de_identificacin") or "").upper().strip()
    desc_local = (rec_local.get("descripcin_fichaproducto") or "").upper().strip()
    
    for p in all_products:
        desc_pc = (p.get("C_Descripcion") or "").upper().strip()
        if desc_local and desc_pc and desc_local == desc_pc:
            return p
        if nro_local and desc_pc and nro_local in desc_pc:
            return p
    return None

def _interpret_response_precios(text: str) -> str:
    t = text.lower()
    if any(w in t for w in ["excede", "supera", "máximo", "maximo", "mayor"]):
        return "EXCEDE"
    if any(w in t for w in ["inferior", "menor", "mínimo", "minimo"]):
        return "INFERIOR"
    if "error" in t:
        return "ERROR"
    return "OK"

def _enviar_oferta_precios(page, log_func, app):
    # Confirmar oferta llamando a los posibles endpoints
    for suffix in ["/EnviarOferta", "/Enviar_Oferta", "/ConfirmarOferta"]:
        url = f"https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp{suffix}"
        try:
            resp = page.request.post(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp/CatalogoProductoIndex"
            }, timeout=30000)
            if resp.status == 200:
                log_func(app, f"✅ API: Oferta confirmada vía {suffix} (Respuesta: {resp.text()[:150]})")
                return True
        except Exception as ex:
            log_func(app, f"⚠ Falló confirmación vía API ({suffix}): {ex}")
            continue
    log_func(app, "❌ No se pudo confirmar la oferta automáticamente por API.")
    return False

def execute_iniciar_precios(app, usuario, password, headless, log_func,
                            precios_data, acuerdo_val, catalogo_val, categoria_val):
    from automation.browser import init_browser, close_browser
    from automation.login import do_login
    from urllib.parse import parse_qs, urlencode
    import time
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from datetime import datetime

    pw = browser = None

    def re_enable():
        app.after(0, lambda: app.btn_test_precios.configure(state="normal"))
        app.after(0, lambda: app.btn_iniciar_precios.configure(state="normal"))

    try:
        pw, browser, page = init_browser(headless=headless)
        if not headless:
            try: page.set_viewport_size({"width": 1920, "height": 1080})
            except Exception: pass

        log_func(app, "✅ Navegador listo, iniciando login...")

        class DummyLog:
            def info(self, m): pass
            def ok(self, m): pass
            def warn(self, m): pass
            def error(self, m): pass

        stop = type('S', (), {'is_set': lambda self: False})()

        if not do_login(page, usuario, password, "", DummyLog(), stop, app.captcha_bridge):
            log_func(app, "❌ Falló el login.")
            re_enable()
            return

        log_func(app, "✅ Login OK. Navegando a t_ProductoOfertadoAmp...")
        page.goto(
            "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp",
            wait_until="domcontentloaded", timeout=60_000
        )
        time.sleep(2)

        # ── 1. Seleccionar Acuerdo ──────────────────────────────────────────
        log_func(app, "⏳ Esperando a que el Acuerdo esté disponible...")
        wait_for_options_precios(page, '#ajaxAcuerdo', timeout=20)
        log_func(app, f"🔧 Seleccionando Acuerdo [{acuerdo_val}]...")
        page.evaluate(f"""
            () => {{
                const sel = document.querySelector('#ajaxAcuerdo');
                if (sel) {{
                    sel.value = '{acuerdo_val}';
                    sel.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }}
        """)

        # ── 2. Seleccionar Catálogo ────────────────────────────────────────
        log_func(app, "⏳ Esperando a que carguen los Catálogos...")
        wait_for_options_precios(page, '#ajaxCatalogo', timeout=25)
        log_func(app, f"🔧 Seleccionando Catálogo [{catalogo_val}]...")
        page.evaluate(f"""
            () => {{
                const sel = document.querySelector('#ajaxCatalogo');
                if (sel) {{
                    sel.value = '{catalogo_val}';
                    sel.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }}
        """)

        # ── 3. Seleccionar Categoría ───────────────────────────────────────
        log_func(app, "⏳ Esperando a que carguen las Categorías...")
        wait_for_options_precios(page, '#ajaxCategoria', timeout=25)
        log_func(app, f"🔧 Seleccionando Categoría [{categoria_val}]...")
        page.evaluate(f"""
            () => {{
                const sel = document.querySelector('#ajaxCategoria');
                if (sel) {{
                    sel.value = '{categoria_val}';
                    sel.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }}
        """)
        time.sleep(2)

        # ── 4. Click Iniciar Búsqueda y esperar tabla ─────────────────────
        log_func(app, "🔍 Haciendo click en Iniciar Búsqueda...")
        page.click("#btnBuscar")
        time.sleep(4)

        try:
            page.wait_for_selector("#btnNuevoProducto", timeout=10_000)
            log_func(app, "✅ Botón 'Agregar oferta' visible.")
        except Exception:
            log_func(app, "⚠ Botón 'Agregar oferta' no apareció, continuando...")

        # ── 5. Interceptar payload del primer request AJAX ─────────────────
        captured_payload = {}
        total_records = [0]

        def handle_request(req):
            if "_CatalogoProductoIndexJson" in req.url:
                captured_payload["body"] = req.post_data
                captured_payload["url"] = req.url

        def handle_response(response):
            if "_CatalogoProductoIndexJson" in response.url:
                try:
                    data = response.json()
                    total_records[0] = data.get("recordsTotal", 0)
                except Exception:
                    pass

        page.on("request", handle_request)
        page.on("response", handle_response)

        try:
            log_func(app, "📂 Navegando a CatalogoProductoIndex...")
            with page.expect_response(lambda r: "_CatalogoProductoIndexJson" in r.url, timeout=45000):
                page.goto(
                    "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp/CatalogoProductoIndex",
                    wait_until="domcontentloaded", timeout=60_000
                )
        except Exception as e:
            log_func(app, f"⚠ Error esperando respuesta de red: {e}")

        page.remove_listener("request", handle_request)
        page.remove_listener("response", handle_response)

        raw_body = captured_payload.get("body")
        url_endpoint = captured_payload.get("url")

        if not raw_body or not url_endpoint:
            log_func(app, "❌ Error: No se pudo capturar el payload de red. Abortando.")
            re_enable()
            return

        # ── 6. Descarga completa del catálogo ──────────────────────────────
        log_func(app, "📡 Descargando catálogo completo de Perú Compras...")
        parsed_params = parse_qs(raw_body)
        total = total_records[0]
        log_func(app, f"📊 Total registros en Perú Compras: {total}")

        all_products = []
        PAGE_SIZE = 100
        start = 0
        draw = 1

        while start < total:
            parsed_params['start'] = [str(start)]
            parsed_params['length'] = [str(PAGE_SIZE)]
            parsed_params['draw'] = [str(draw)]
            new_body = urlencode(parsed_params, doseq=True)

            resp = page.request.post(
                url_endpoint,
                headers={
                    "Content-Type": "application/x-www-form-urlencoded",
                    "X-Requested-With": "XMLHttpRequest",
                    "Referer": "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp/CatalogoProductoIndex"
                },
                data=new_body,
                timeout=20000
            )

            if resp.status == 200:
                try:
                    data = resp.json()
                    records = data.get("data", [])
                    all_products.extend(records)
                except Exception as je:
                    log_func(app, f"   ❌ Error parseando JSON en lote {start}: {je}")
                    break
            else:
                log_func(app, f"   ❌ Error de red en lote {start}: Status {resp.status}")
                break

            start += PAGE_SIZE
            draw += 1
            time.sleep(0.3)

        log_func(app, f"✅ Catálogo descargado: {len(all_products)} productos")

        # ── 7. Bucle de Subida de Precios ──────────────────────────────────
        log_func(app, f"🚀 Iniciando inserción de precios en paralelo para {len(precios_data)} fichas locales...")
        
        report_rows = []
        oks = 0
        errors = 0
        no_matches = 0

        url_insert = "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp/Inserta_ProductoOfertadoTMP"

        for idx, rec in enumerate(precios_data, 1):
            desc_local = rec.get("descripcin_fichaproducto", "")
            nro_local = rec.get("nro_parte_o_cdigo_nico_de_identificacin", "")
            precio_max = rec.get("precio_max", 0)
            precio_usd = _calcular_precio_dolar(precio_max)

            # Buscar match en el catálogo descargado
            cat_match = _buscar_match_pc(rec, all_products)
            
            status = "PENDIENTE"
            server_response = ""
            catalogo_id = ""

            if cat_match:
                catalogo_id = str(cat_match.get("N_CatalogoProducto") or "")
                moneda = str(cat_match.get("C_MonedaOfertada") or "USD")

                # Realizar POST de inserción
                payload_insert = {
                    "N_CatalogoProducto": catalogo_id,
                    "C_MonedaOfertada": moneda,
                    "N_PrecioOfertado": str(precio_usd),
                }

                try:
                    resp_ins = page.request.post(
                        url_insert,
                        headers={
                            "Content-Type": "application/x-www-form-urlencoded",
                            "X-Requested-With": "XMLHttpRequest",
                            "Referer": "https://www.catalogos.perucompras.gob.pe/t_ProductoOfertadoAmp/CatalogoProductoIndex"
                        },
                        data=urlencode(payload_insert),
                        timeout=15000
                    )
                    
                    if resp_ins.status == 200:
                        server_response = resp_ins.text()
                        status = _interpret_response_precios(server_response)
                        if status == "OK":
                            oks += 1
                            log_func(app, f"   [{idx}/{len(precios_data)}] ✅ {nro_local} -> {precio_usd} USD (Insertado)")
                        else:
                            errors += 1
                            log_func(app, f"   [{idx}/{len(precios_data)}] ⚠ {nro_local} -> Rechazado: {server_response[:100]}")
                    else:
                        status = "ERROR"
                        server_response = f"HTTP Error Status {resp_ins.status}"
                        errors += 1
                        log_func(app, f"   [{idx}/{len(precios_data)}] ❌ {nro_local} -> Error de red: {resp_ins.status}")
                except Exception as ex:
                    status = "ERROR"
                    server_response = str(ex)
                    errors += 1
                    log_func(app, f"   [{idx}/{len(precios_data)}] ❌ {nro_local} -> Excepción: {ex}")
            else:
                status = "NO ENCONTRADO"
                server_response = "No se localizó descripción/número de parte en Perú Compras"
                no_matches += 1

            report_rows.append({
                "descripcion": desc_local,
                "nro_parte": nro_local,
                "precio_max_soles": precio_max,
                "precio_usd": precio_usd,
                "id_catalogo_producto": catalogo_id,
                "estado": status,
                "respuesta_servidor": server_response
            })
            
            # Pequeña pausa para no saturar al servidor
            time.sleep(0.1)

        # ── 8. Enviar/Confirmar Oferta Final ────────────────────────────────
        if oks > 0:
            log_func(app, f"💾 Confirmando y enviando {oks} ofertas ingresadas exitosamente...")
            try:
                log_func(app, "🔄 Recargando página de Perú Compras para refrescar la tabla...")
                page.reload(wait_until="domcontentloaded")
                time.sleep(3)
                
                # Buscamos el botón de enviar
                btn = page.locator("#btn_enviarOferta2, #btn_enviarOferta, button:text('Enviar oferta')").first
                if btn.is_visible():
                    log_func(app, "🖱 Haciendo click físico en el botón 'Enviar oferta' en la página...")
                    btn.click()
                    time.sleep(4)
                    
                    # Aceptar cuadro de diálogo SweetAlert o de confirmación si aparece
                    try:
                        confirm_btn = page.locator(".swal2-confirm, button:text('Aceptar'), button:text('Confirmar')").first
                        confirm_btn.wait_for(state="visible", timeout=5000)
                        confirm_btn.click()
                        log_func(app, "✅ Diálogo de confirmación aceptado en pantalla.")
                        time.sleep(3)
                    except Exception as ex:
                        log_func(app, f"ℹ No se detectó o no apareció popup de confirmación: {ex}")
                    log_func(app, "✅ Confirmación física completada.")
                else:
                    log_func(app, "⚠ Botón de enviar no visible en la página. Intentando fallback por API...")
                    _enviar_oferta_precios(page, log_func, app)
            except Exception as e:
                log_func(app, f"❌ Error confirmando en UI ({e}). Intentando fallback por API...")
                _enviar_oferta_precios(page, log_func, app)
        else:
            log_func(app, "⚠ No se insertó ningún precio exitosamente. Saltando confirmación final.")

        # ── 9. Generar Reporte Excel ───────────────────────────────────────
        log_func(app, "📊 Generando reporte Excel de resultados...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados Subida"
        
        # Estilos
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        fill_ok = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        fill_error = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        fill_warn = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        headers = ["Descripción Ficha", "Número de Parte", "Precio Máx (S/)", "Precio USD Calculado", "ID Catálogo Producto", "Estado Subida", "Respuesta Servidor"]
        ws.append(headers)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
            
        for r in report_rows:
            ws.append([
                r["descripcion"],
                r["nro_parte"],
                r["precio_max_soles"],
                r["precio_usd"],
                r["id_catalogo_producto"],
                r["estado"],
                r["respuesta_servidor"]
            ])
            
            curr_row = ws.max_row
            status_cell = ws.cell(row=curr_row, column=6)
            
            # Colorear según estado
            if r["estado"] == "OK":
                status_cell.fill = fill_ok
            elif r["estado"] in ["ERROR", "EXCEDE", "INFERIOR"]:
                status_cell.fill = fill_error
            elif r["estado"] == "NO ENCONTRADO":
                status_cell.fill = fill_warn
                
        # Autoajustar columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        report_name = f"reporte_subida_precios_{timestamp}.xlsx"
        report_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), report_name)
        wb.save(report_path)
        
        log_func(app, f"\n✅ PROCESO COMPLETADO:")
        log_func(app, f"   - Exitosos: {oks}")
        log_func(app, f"   - Fallidos: {errors}")
        log_func(app, f"   - Sin coincidencia: {no_matches}")
        log_func(app, f"\n📂 Reporte generado en:")
        log_func(app, f"   {report_path}")

    except Exception as e:
        import traceback
        log_func(app, f"❌ Error crítico en subida: {e}")
        log_func(app, traceback.format_exc())
    finally:
        if browser and pw:
            try: close_browser(pw, browser)
            except: pass
        re_enable()
