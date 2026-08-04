"""
audit_reporter.py — Módulo Auditor de Resultados e Informe de Auditoría (Excel / PDF)
Permite verificar rápidamente el estado de subidas y generar informes corporativos en Excel o PDF.
"""

import os
import time
from datetime import datetime

def audit_results(rows_data):
    """
    Analiza una lista de dicts o tuplas con la información del proceso.
    row: dict(parte, descripcion, precio, stock, estado, obs)
    Retorna un diccionario summary con métricas de auditoría.
    """
    total = len(rows_data)
    ok_cnt = 0
    warn_cnt = 0
    err_cnt = 0

    for r in rows_data:
        st = str(r.get("estado", "")).lower()
        if "éxito" in st or "exito" in st or "ok" in st or "✓" in st:
            ok_cnt += 1
        elif "no hallado" in st or "no encontrado" in st or "advertencia" in st or "⚠️" in st:
            warn_cnt += 1
        elif "error" in st or "fallo" in st or "✕" in st:
            err_cnt += 1
        else:
            # Si aún está pendiente o desconocido
            pass

    rate = (ok_cnt / total * 100.0) if total > 0 else 0.0

    return {
        "total": total,
        "ok": ok_cnt,
        "warn": warn_cnt,
        "err": err_cnt,
        "pending": max(0, total - (ok_cnt + warn_cnt + err_cnt)),
        "rate": round(rate, 1),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }


def export_excel_report(rows_data, summary, output_path, modulo_nombre="Publicación PDF"):
    """
    Genera un informe completo de auditoría en formato Excel (.xlsx).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Informe de Auditoría"
        ws.views.sheetView[0].showGridLines = True

        # Estilos corporativos
        fill_header = PatternFill(start_color="006CA8", end_color="006CA8", fill_type="solid")
        fill_sub = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        fill_ok = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")
        fill_warn = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")
        fill_err = PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid")

        font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_sub = Font(name="Segoe UI", size=10, bold=True, color="1A1A1A")
        font_normal = Font(name="Segoe UI", size=10, color="1A1A1A")
        font_bold = Font(name="Segoe UI", size=10, bold=True, color="1A1A1A")

        font_ok = Font(name="Segoe UI", size=10, bold=True, color="1B6B1B")
        font_warn = Font(name="Segoe UI", size=10, bold=True, color="854D0E")
        font_err = Font(name="Segoe UI", size=10, bold=True, color="8B1A1A")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        thin = Side(border_style="thin", color="C8C8C8")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 1. BANNER TÍTULO
        ws.merge_cells("A1:F2")
        cell_t = ws["A1"]
        cell_t.value = f"INFORME DE AUDITORÍA — PERÚ COMPRAS BOT v1.4\nMódulo: {modulo_nombre}"
        cell_t.font = font_title
        cell_t.fill = fill_header
        cell_t.alignment = align_center

        # 2. RESUMEN DE AUDITORÍA
        ws.merge_cells("A4:F4")
        cell_res = ws["A4"]
        cell_res.value = "RESUMEN DE AUDITORÍA Y VERIFICACIÓN"
        cell_res.font = font_sub
        cell_res.fill = fill_sub

        metrics = [
            ("Fecha y Hora:", summary.get("timestamp", "")),
            ("Total Productos Procesados:", summary.get("total", 0)),
            ("Subidos / Exitosos (✓):", summary.get("ok", 0)),
            ("No Hallados / Advertencias (⚠️):", summary.get("warn", 0)),
            ("Errores de Procesamiento (✕):", summary.get("err", 0)),
            ("Tasa de Éxito General:", f"{summary.get('rate', 0.0)}%"),
        ]

        row_idx = 5
        for label, val in metrics:
            ws.cell(row=row_idx, column=1, value=label).font = font_bold
            c_val = ws.cell(row=row_idx, column=2, value=val)
            c_val.font = font_normal
            if "Éxito" in label or "Exitosos" in label:
                c_val.font = font_ok
            elif "Errores" in label:
                c_val.font = font_err
            row_idx += 1

        row_idx += 1

        # 3. TABLA DE DETALLE POR PRODUCTO
        headers = ["N°", "N° de Parte", "Descripción / Marca", "Precio S/", "Stock", "Estado / Resultado Auditoría"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        row_idx += 1

        for i, r in enumerate(rows_data, 1):
            pn = str(r.get("parte", "") or r.get("nro_parte", "") or "")
            desc = str(r.get("descripcion", "") or r.get("marca", "") or "")
            precio = str(r.get("precio", "") or "")
            stock = str(r.get("stock", "") or "")
            st = str(r.get("estado", "") or "Pendiente")

            ws.cell(row=row_idx, column=1, value=i).alignment = align_center
            ws.cell(row=row_idx, column=2, value=pn).alignment = align_left
            ws.cell(row=row_idx, column=3, value=desc).alignment = align_left
            ws.cell(row=row_idx, column=4, value=precio).alignment = align_center
            ws.cell(row=row_idx, column=5, value=stock).alignment = align_center

            cell_st = ws.cell(row=row_idx, column=6, value=st)
            cell_st.alignment = align_left

            st_low = st.lower()
            if "éxito" in st_low or "exito" in st_low or "ok" in st_low or "✓" in st_low:
                cell_st.fill = fill_ok
                cell_st.font = font_ok
            elif "no hallado" in st_low or "advertencia" in st_low or "⚠️" in st_low:
                cell_st.fill = fill_warn
                cell_st.font = font_warn
            elif "error" in st_low or "fallo" in st_low or "✕" in st_low:
                cell_st.fill = fill_err
                cell_st.font = font_err
            else:
                cell_st.font = font_normal

            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).border = border_all

            row_idx += 1

        # Autoajustar ancho de columnas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.coordinate in ws.merged_cells:
                    continue
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        return True, output_path
    except Exception as e:
        return False, str(e)


def export_pdf_report(rows_data, summary, output_path, modulo_nombre="Publicación PDF"):
    """
    Genera un informe detallado de auditoría en formato PDF (.pdf) estructurado en HTML printable.
    """
    try:
        html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Informe de Auditoría — Perú Compras Bot</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 12px; color: #1A1A1A; margin: 20px; background: #FFF; }}
        .header {{ background: #006CA8; color: white; padding: 16px; text-align: center; border-radius: 4px; }}
        .header h1 {{ margin: 0; font-size: 18px; }}
        .header p {{ margin: 4px 0 0 0; font-size: 12px; color: #AACCDD; }}
        .resumen {{ background: #F8F9FA; border: 1px solid #C8C8C8; padding: 12px; margin: 16px 0; border-radius: 4px; }}
        .resumen h3 {{ margin-top: 0; color: #006CA8; border-bottom: 1px solid #C8C8C8; padding-bottom: 4px; }}
        .grid-sum {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; }}
        .stat-card {{ background: white; padding: 10px; border: 1px solid #E0E0E0; border-radius: 4px; text-align: center; }}
        .stat-card .val {{ font-size: 18px; font-weight: bold; margin-top: 4px; }}
        .val.ok {{ color: #1B6B1B; }}
        .val.warn {{ color: #854D0E; }}
        .val.err {{ color: #8B1A1A; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
        th {{ background: #006CA8; color: white; padding: 8px; text-align: left; font-size: 11px; }}
        td {{ padding: 8px; border-bottom: 1px solid #E0E0E0; font-size: 11px; }}
        tr:nth-child(even) {{ background: #F9F9F9; }}
        .tag-ok {{ color: #1B6B1B; font-weight: bold; background: #DFF0D8; padding: 2px 6px; border-radius: 3px; }}
        .tag-warn {{ color: #854D0E; font-weight: bold; background: #FCF8E3; padding: 2px 6px; border-radius: 3px; }}
        .tag-err {{ color: #8B1A1A; font-weight: bold; background: #F2DEDE; padding: 2px 6px; border-radius: 3px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>INFORME DE AUDITORÍA — PERÚ COMPRAS BOT v1.4</h1>
        <p>Módulo: {modulo_nombre} | Generado: {summary.get('timestamp', '')}</p>
    </div>

    <div class="resumen">
        <h3>Resumen Ejecutivo de Verificación</h3>
        <div class="grid-sum">
            <div class="stat-card">
                <div>Total Procesados</div>
                <div class="val">{summary.get('total', 0)}</div>
            </div>
            <div class="stat-card">
                <div>Fichas Subidas (✓)</div>
                <div class="val ok">{summary.get('ok', 0)}</div>
            </div>
            <div class="stat-card">
                <div>No Hallados / Errores (✕)</div>
                <div class="val err">{summary.get('err', 0) + summary.get('warn', 0)}</div>
            </div>
        </div>
        <p style="margin-top: 10px; font-weight: bold;">Tasa de Éxito Operativa: <span class="val ok">{summary.get('rate', 0.0)}%</span></p>
    </div>

    <h3>Detalle de Registros de Productos</h3>
    <table>
        <thead>
            <tr>
                <th>N°</th>
                <th>N° de Parte</th>
                <th>Descripción / Marca</th>
                <th>Precio S/</th>
                <th>Stock</th>
                <th>Resultado de Auditoría</th>
            </tr>
        </thead>
        <tbody>
"""
        for i, r in enumerate(rows_data, 1):
            pn = str(r.get("parte", "") or r.get("nro_parte", "") or "")
            desc = str(r.get("descripcion", "") or r.get("marca", "") or "")
            precio = str(r.get("precio", "") or "")
            stock = str(r.get("stock", "") or "")
            st = str(r.get("estado", "") or "Pendiente")

            st_low = st.lower()
            if "éxito" in st_low or "exito" in st_low or "ok" in st_low or "✓" in st_low:
                tag_cls = "tag-ok"
            elif "no hallado" in st_low or "advertencia" in st_low or "⚠️" in st_low:
                tag_cls = "tag-warn"
            elif "error" in st_low or "fallo" in st_low or "✕" in st_low:
                tag_cls = "tag-err"
            else:
                tag_cls = ""

            html_content += f"""
            <tr>
                <td>{i}</td>
                <td><strong>{pn}</strong></td>
                <td>{desc}</td>
                <td>{precio}</td>
                <td>{stock}</td>
                <td><span class="{tag_cls}">{st}</span></td>
            </tr>
"""

        html_content += """
        </tbody>
    </table>
</body>
</html>
"""
        # Escribir a archivo .html/.pdf para fácil exportación/lectura
        pdf_html_path = output_path if output_path.endswith(".html") else output_path + ".html"
        with open(pdf_html_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        return True, pdf_html_path
    except Exception as e:
        return False, str(e)
