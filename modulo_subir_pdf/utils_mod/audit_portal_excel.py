"""
audit_portal_excel.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Genera el Excel de auditoría del Auditor Portal Stock.

Estructura del Excel generado:
  • Hoja 1 "Resumen"   → cabecera institucional + cuadro semáforo de conteos
  • Hoja 2 "Detalle"  → tabla fila por fila: producto / stock_excel / stock_portal / resultado

Función pública:
  generar_excel_auditoria(filas, resumen, ruta_salida) -> (bool, str)
"""

from __future__ import annotations
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── Paleta institucional ──────────────────────────────────────────
_GRIS_HEADER = "2C3E50"
_BLANCO      = "FFFFFF"
_VERDE       = "1E8449"
_NARANJA     = "D68910"
_ROJO        = "922B21"

def _header_font(bold=True, size=11, color=_BLANCO):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _normal_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Hoja de Resumen ──────────────────────────────────────────────

def _build_summary_sheet(ws, resumen: dict):
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCD", [28, 18, 28, 18]):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 46

    # Titulo
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "AUDITORIA DE STOCK — PERU COMPRAS vs EXCEL LOCAL"
    c.font = Font(name="Calibri", bold=True, size=14, color=_BLANCO)
    c.fill = _fill(_GRIS_HEADER)
    c.alignment = _center()

    # Subtitulo
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"Generado el {resumen.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}"
    c.font = _normal_font(size=9, color="777777")
    c.fill = _fill("ECF0F1")
    c.alignment = _center()
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    # Semaforo
    semaforo = [
        ("Total Fichas Auditadas", resumen.get("total",   0), _GRIS_HEADER, "->"),
        ("Correctas (Coinciden)",  resumen.get("ok",      0), _VERDE,       "OK"),
        ("Diferencias de Stock",   resumen.get("dif",     0), _NARANJA,     "!="),
        ("No encontradas Portal",  resumen.get("missing", 0), _ROJO,        "X"),
    ]
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 42
    for col, (label, valor, color, icon) in enumerate(semaforo, start=1):
        lbl = ws.cell(row=4, column=col)
        lbl.value = label
        lbl.font = _header_font(size=9)
        lbl.fill = _fill(color)
        lbl.alignment = _center()
        lbl.border = _border()
        val = ws.cell(row=5, column=col)
        val.value = f"{icon}  {valor}"
        val.font = Font(name="Calibri", bold=True, size=20, color=color)
        val.fill = _fill("FDFEFE")
        val.alignment = _center()
        val.border = _border()

    ws.row_dimensions[6].height = 14

    # Extras
    extras = [
        ("Acuerdo Marco",        resumen.get("acuerdo",   "—")),
        ("Catalogo",             resumen.get("catalogo",  "—")),
        ("Categoria",            resumen.get("categoria", "—")),
        ("Tasa de coincidencia", f"{resumen.get('tasa', 0):.1f}%"),
        ("Excel cargado",        resumen.get("excel_file", "—")),
    ]
    ws.merge_cells("A7:D7")
    h = ws["A7"]
    h.value = "Parametros de la auditoria"
    h.font = _header_font(size=10)
    h.fill = _fill(_GRIS_HEADER)
    h.alignment = _left()
    h.border = _border()
    for i, (k, v) in enumerate(extras, start=8):
        ws.row_dimensions[i].height = 18
        ka = ws.cell(row=i, column=1)
        ka.value = k
        ka.font = _normal_font(bold=True, size=10)
        ka.fill = _fill("EBF5FB")
        ka.alignment = _left()
        ka.border = _border()
        ws.merge_cells(f"B{i}:D{i}")
        va = ws.cell(row=i, column=2)
        va.value = str(v)
        va.font = _normal_font(size=10)
        va.alignment = _left()
        va.border = _border()


# ── Hoja de Detalle ──────────────────────────────────────────────

_COLS_DETALLE = [
    ("N",               6),
    ("Parte (Excel)",  18),
    ("Desc. Excel",    40),
    ("Stock Excel",    14),
    ("Precio Excel S/",16),
    ("Ficha (Excel)",  14),
    ("Stock Portal",   14),
    ("Estado Portal",  14),
    ("Diferencia",     12),
    ("Resultado",      16),
]

def _build_detail_sheet(ws, filas: list):
    ws.title = "Detalle"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for i, (_, ancho) in enumerate(_COLS_DETALLE, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[1].height = 28
    for col, (nombre, _) in enumerate(_COLS_DETALLE, start=1):
        c = ws.cell(row=1, column=col)
        c.value = nombre
        c.font = _header_font(size=10)
        c.fill = _fill(_GRIS_HEADER)
        c.alignment = _center()
        c.border = _border()

    for idx, fila in enumerate(filas, start=2):
        ws.row_dimensions[idx].height = 18
        resultado = fila.get("resultado", "")
        if resultado == "OK":
            row_fill = _fill("EAF7EA")
            res_fill = _fill(_VERDE)
            res_font = Font(name="Calibri", bold=True, size=10, color=_BLANCO)
        elif resultado == "DIFERENCIA":
            row_fill = _fill("FEF9E7")
            res_fill = _fill(_NARANJA)
            res_font = Font(name="Calibri", bold=True, size=10, color=_BLANCO)
        elif resultado == "NO ENCONTRADO":
            row_fill = _fill("FDEDEC")
            res_fill = _fill(_ROJO)
            res_font = Font(name="Calibri", bold=True, size=10, color=_BLANCO)
        else:
            row_fill = _fill(_BLANCO)
            res_fill = _fill("BDC3C7")
            res_font = _normal_font(size=10)

        diferencia = fila.get("diferencia", "—")
        if isinstance(diferencia, (int, float)) and diferencia != 0:
            dif_str = f"{'+' if diferencia > 0 else ''}{diferencia}"
        else:
            dif_str = str(diferencia) if diferencia is not None else "—"

        vals = [
            idx - 1,
            fila.get("parte", ""),
            fila.get("descripcion", ""),
            fila.get("stock_excel", ""),
            fila.get("precio_excel", ""),
            fila.get("ficha", ""),
            fila.get("stock_portal", "—"),
            fila.get("estado_portal", "—"),
            dif_str,
            resultado,
        ]
        center_cols = {1, 4, 5, 6, 7, 8, 9, 10}
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=idx, column=col)
            c.value = val
            c.border = _border()
            c.alignment = _center() if col in center_cols else _left()
            if col == 10:
                c.fill = res_fill
                c.font = res_font
            else:
                c.fill = row_fill
                c.font = _normal_font(size=10)


# ── Función pública ──────────────────────────────────────────────

def generar_excel_auditoria(filas: list, resumen: dict, ruta_salida: str):
    """
    Genera el Excel de auditoria.

    filas: list[dict] — cada dict con:
        parte, descripcion, stock_excel, precio_excel, ficha,
        stock_portal, estado_portal, diferencia, resultado
        resultado in {"OK", "DIFERENCIA", "NO ENCONTRADO"}

    resumen: dict — total, ok, dif, missing, tasa, acuerdo, catalogo,
        categoria, excel_file, timestamp

    ruta_salida: str — path completo del .xlsx

    Retorna (True, ruta_salida) | (False, mensaje_error)
    """
    if not _HAS_OPENPYXL:
        return False, "openpyxl no esta instalado."
    try:
        wb = Workbook()
        wb.remove(wb.active)
        ws_resumen = wb.create_sheet("Resumen")
        _build_summary_sheet(ws_resumen, resumen)
        ws_detalle = wb.create_sheet("Detalle")
        _build_detail_sheet(ws_detalle, filas)
        wb.active = ws_resumen
        os.makedirs(os.path.dirname(ruta_salida) if os.path.dirname(ruta_salida) else ".", exist_ok=True)
        wb.save(ruta_salida)
        return True, ruta_salida
    except Exception as e:
        import traceback
        return False, f"{e}\n{traceback.format_exc()}"
