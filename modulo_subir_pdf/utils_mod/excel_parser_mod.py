"""
excel_parser_mod.py — Lector del Excel de entrada para modificación de productos.

Columnas esperadas (auto-detectadas por nombre):
  - N° de Parte / Código / Part Number  → clave de búsqueda
  - PDF / Archivo / Ficha               → ruta absoluta o relativa al PDF a subir
  - Certificaciones / Certs             → texto de certificaciones (opcional)
  - Resto de columnas (PROCESADOR, RAM, ALMACENAMIENTO, etc.) → características
    cuyos valores se comparan contra PeruCompras
  - Columnas "CERTIFICACION 1".."CERTIFICACION 5" → certificaciones esperadas

Retorna una lista de dicts con las claves normalizadas.
"""
import os
import re
from typing import Optional

try:
    import openpyxl
except ImportError as e:
    raise ImportError("openpyxl es requerido: pip install openpyxl") from e


# Alias de nombres de columna reconocidos (insensible a mayúsculas/acentos)
_PARTE_ALIASES = [
    "n° de parte", "n° parte", "num parte", "numero parte",
    "código", "codigo", "cod", "part number", "partnumber",
    "código único", "codigo unico", "nro_parte", "nro parte",
]
_FICHA_ALIASES = [
    "ficha n°", "ficha nº", "ficha n", "ficha", "id ficha",
    "id producto", "id_producto", "numero ficha", "nro ficha",
]
_PDF_ALIASES = [
    "pdf", "archivo", "ficha", "ruta pdf", "ruta_pdf",
    "ficha técnica", "ficha tecnica", "documento",
]
_CERT_ALIASES = [
    "certificaciones", "certificacion", "certs", "cert",
    "certif", "certificado",
]

# Características de ficha técnica (en PeruCompras el nombre de columna
# coincide con la "Característica" del HTML). Alias cubre variantes de
# mayúsculas/acentos. Si tu Excel tiene más columnas no listadas acá,
# se incluyen automáticamente como características siempre que no sean
# parte/pdf/cert y tengan valor no vacío.
_KNOWN_CHARS = {
    "PROCESADOR", "RAM", "ALMACENAMIENTO", "LAN", "WLAN", "USB", "VGA",
    "HDMI", "SIST. OPER", "SIST. OPERATIVO", "UNIDAD OPTICA",
    "TECLADO", "MOUSE", "SUITE OFIMATICA", "G. F", "UNIDAD",
    "MARCA", "MODELO", "EMPAQUE DE FABRICA", "EMPAQUE DE FÁBRICA",
    "SISTEMA DE MANEJO DE RAEE",
}

# Columnas no-carácter que deben ignorarse (metadatos, no se comparan)
# IMAGEN (PDF) y PRECIO SUGERIDO NO se filtran — se leen como campos separados
_NON_CHAR_COLS = {
    "N°", "Nº", "NRO", "N", "FICHA N°", "FICHA Nº", "FICHA N",
    "FICHA TECNICA (PDF)",
}


def _normalize(text: str) -> str:
    """Normaliza texto para comparación: minúsculas, sin acentos."""
    text = str(text).lower().strip()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n",
        "°": "", "º": "",  # símbolos ordinales
    }
    for a, b in replacements.items():
        text = text.replace(a, b)
    return text


def _match_col(header: str, aliases: list[str]) -> bool:
    norm = _normalize(header)
    # Si la normalización quedó en 1-2 chars, exigir match exacto (evita que "n" matchee "n° de parte")
    if len(norm) <= 2:
        return any(norm == _normalize(a) for a in aliases)
    return any(
        _normalize(a) in norm
        for a in aliases
        if len(_normalize(a)) >= 3
    )


def get_sheets(path: str) -> list[str]:
    """Retorna la lista de hojas del Excel."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def detect_columns(path: str, sheet: str) -> dict:
    """
    Detecta automáticamente las columnas de parte, PDF, certificaciones,
    características y certificaciones esperadas.
    Retorna un dict con:
      - 'parte_col'   : nombre exacto de la columna N° de Parte
      - 'ficha_col'   : nombre exacto de la columna Ficha N° (o None)
      - 'pdf_col'     : nombre exacto de la columna PDF (o None)
      - 'cert_col'    : nombre exacto de la columna Certificaciones (o None)
      - 'char_cols'   : lista de nombres de columnas de características
      - 'cert_expected_cols' : lista de nombres de columnas CERTIFICACION N
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    parte_col = ficha_col = pdf_col = cert_col = None
    char_cols = []
    cert_expected_cols = []

    cert_re = re.compile(r"^certificacion\s*\d+$", re.IGNORECASE)

    # Solo leer la PRIMERA fila (header) — no iterar filas de datos
    first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if first_row:
        for cell in first_row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if not val:
                continue
            norm = _normalize(val)
            # Prioridad 1: CERTIFICACION N → cert_expected_cols (nunca cert_col legacy)
            if cert_re.match(norm):
                cert_expected_cols.append(val)
                continue
            # Prioridad 2: parte / pdf / cert legacy / ficha
            if parte_col is None and _match_col(val, _PARTE_ALIASES):
                parte_col = val
                continue
            if ficha_col is None and _match_col(val, _FICHA_ALIASES):
                # Evitar confundir "FICHA TECNICA (PDF)" con ficha
                if "tecnica" not in norm:
                    ficha_col = val
                continue
            if pdf_col is None and _match_col(val, _PDF_ALIASES):
                if _normalize(val).startswith("ficha tecnica") or _normalize(val) == "pdf":
                    pdf_col = val
                continue
            if cert_col is None and _match_col(val, _CERT_ALIASES):
                cert_col = val
                continue
            # Prioridad 3: columna desconocida → característica
            if _normalize(val) not in {_normalize(x) for x in _NON_CHAR_COLS}:
                char_cols.append(val)

    wb.close()
    return {
        "parte_col": parte_col,
        "ficha_col": ficha_col,
        "pdf_col": pdf_col,
        "cert_col": cert_col,
        "char_cols": char_cols,
        "cert_expected_cols": cert_expected_cols,
    }


def parse_excel(
    path: str,
    sheet: str,
    parte_col: str,
    pdf_col: Optional[str] = None,
    cert_col: Optional[str] = None,
) -> list[dict]:
    """
    Lee el Excel y retorna una lista de dicts con:
      {
        "parte":   str,         # N° de Parte
        "pdf":     str | None,  # Ruta al PDF (o None)
        "certs":   str | None,  # Texto de certificaciones (o None)
        "caracteristicas": dict,  # {nombre_caracteristica: valor_esperado, ...}
        "certs_esperadas": list,  # ['ISO 9001', 'CE O UE', ...]
        "_row_idx": int,        # Índice de fila (0-based, desde header+1)
      }
    Filas con parte vacía son ignoradas.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw:
        return []

    # Encontrar header row — usar la primera fila con contenido
    header_row_idx = 0
    headers = []
    for i, row in enumerate(rows_raw[:10]):
        vals = [str(v) if v is not None else "" for v in row]
        if any(v.strip() for v in vals):
            header_row_idx = i
            headers = vals
            break
    else:
        return []

    # Mapear columna de parte: match exacto contra parte_col o primer columna
    col_idx: dict[str, int] = {}
    parte_found = False
    if parte_col:
        for j, h in enumerate(headers):
            if h.strip() == parte_col.strip():
                col_idx["parte"] = j
                parte_found = True
                break
    if not parte_found:
        # Fallback: buscar por alias
        for j, h in enumerate(headers):
            if h and _match_col(h.strip(), _PARTE_ALIASES):
                col_idx["parte"] = j
                parte_found = True
                break
    if not parte_found:
        # Fallback extremo: columna 0
        col_idx["parte"] = 0

    # Mapear columnas opcionales (cert legacy: NO matchear CERTIFICACION N)
    cert_re = re.compile(r"^certificacion\s*\d+$", re.IGNORECASE)
    for col_name, aliases in [("pdf", _PDF_ALIASES), ("certs", _CERT_ALIASES)]:
        for j, h in enumerate(headers):
            if h and _match_col(h.strip(), aliases):
                # Si es CERTIFICACION N, no es el cert legacy
                if cert_re.match(_normalize(h.strip())):
                    continue
                col_idx[col_name] = j
                break

    # Detectar columnas de características, certs esperadas, imagen y precio
    char_col_indices = []
    cert_expected_indices = []
    imagen_idx = None
    precio_idx = None
    for j, h in enumerate(headers):
        if not h:
            continue
        norm = _normalize(h)
        if j in (col_idx.get("parte"), col_idx.get("ficha"), col_idx.get("pdf"), col_idx.get("certs")):
            continue
        if cert_re.match(norm):
            cert_expected_indices.append(j)
        elif "imagen" in norm and "pdf" in norm:
            imagen_idx = j
        elif "precio" in norm and "sugerido" in norm:
            precio_idx = j
        elif norm in ("imagen", "foto"):
            imagen_idx = j
        elif norm == "precio":
            precio_idx = j
        elif norm not in {_normalize(x) for x in _NON_CHAR_COLS}:
            char_col_indices.append((j, h.strip()))

    # Detectar columna de ficha también en parse_excel
    ficha_idx = None
    for j, h in enumerate(headers):
        if h and _match_col(h.strip(), _FICHA_ALIASES):
            if "tecnica" not in _normalize(h.strip()):
                ficha_idx = j
                break

    # Parsear filas de datos
    results = []
    for i, row in enumerate(rows_raw[header_row_idx + 1:]):
        if all(v is None for v in row):
            continue

        parte_val = ""
        ficha_val = None
        pdf_val = None
        certs_val = None
        imagen_val = None
        precio_val = None
        chars = {}
        certs_esp = []

        idx = col_idx.get("parte", 0)
        if idx < len(row):
            v = row[idx]
            parte_val = str(v).strip() if v is not None else ""

        if not parte_val:
            continue

        # Leer ficha
        if ficha_idx is not None and ficha_idx < len(row):
            v = row[ficha_idx]
            if v is not None and str(v).strip():
                ficha_val = str(v).strip()

        if col_idx.get("pdf") is not None and col_idx["pdf"] < len(row):
            v = row[col_idx["pdf"]]
            if v is not None and str(v).strip():
                pdf_val = str(v).strip()

        if col_idx.get("certs") is not None and col_idx["certs"] < len(row):
            v = row[col_idx["certs"]]
            if v is not None and str(v).strip():
                certs_val = str(v).strip()

        # Leer imagen (columna IMAGEN (PDF))
        if imagen_idx is not None and imagen_idx < len(row):
            v = row[imagen_idx]
            if v is not None and str(v).strip():
                imagen_val = str(v).strip()

        # Leer precio (columna PRECIO SUGERIDO)
        if precio_idx is not None and precio_idx < len(row):
            v = row[precio_idx]
            if v is not None and str(v).strip():
                precio_val = str(v).strip()

        for j, name in char_col_indices:
            if j < len(row):
                v = row[j]
                if v is not None and str(v).strip():
                    chars[name] = str(v).strip()

        for j in cert_expected_indices:
            if j < len(row):
                v = row[j]
                if v is not None and str(v).strip():
                    certs_esp.append(str(v).strip())

        results.append({
            "parte": parte_val,
            "ficha": ficha_val or "",
            "imagen": imagen_val or "",
            "precio": precio_val or "",
            "pdf": pdf_val,
            "certs": certs_val,
            "caracteristicas": chars,
            "certs_esperadas": certs_esp,
            "_row_idx": i,
        })

    return results
