# -*- coding: utf-8 -*-
"""
utils_mod/excel_report_designer.py — Generador y Diseñador Profesional de Reportes Excel.

Crea hojas de cálculo de nivel ejecutivo con:
1. Dashboard de Métricas y KPIs (incluyendo tiempo total transcurrido).
2. Tabla de Detalle estructurada con anchos auto-ajustados y formatos de moneda/número.
3. Formato condicional sobrio y limpio (sin colores chillones ni estilos desordenados).
4. Paneles congelados y autofiltros activos.
"""

from __future__ import annotations
import os
import time
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def format_elapsed_time(seconds: float) -> str:
    """Formatea segundos en una cadena legible: '02m 45s' o '01h 14m 20s'."""
    if seconds < 0:
        return "0s"
    s = int(round(seconds))
    hrs = s // 3600
    mins = (s % 3600) // 60
    secs = s % 60
    if hrs > 0:
        return f"{hrs:02d}h {mins:02d}m {secs:02d}s"
    elif mins > 0:
        return f"{mins:02d}m {secs:02d}s"
    else:
        return f"{secs}s"


# ── Paleta Institucional Ejecutiva ────────────────────────────────
_AZUL_PRINCIPAL = "003366"   # Azul marino institucional
_AZUL_HEADER    = "0F4C81"   # Azul pizarra para encabezados
_AZUL_CLARO     = "EBF5FB"   # Fondo sutil para tarjetas
_GRIS_FONDO     = "F8FAFC"   # Fondo alternado de filas
_GRIS_BORDE     = "CBD5E1"   # Borde gris suave
_BLANCO         = "FFFFFF"
_TEXTO_OSCURO   = "1E293B"

# Estados
_VERDE_FILL  = "D1F2EB"
_VERDE_TXT   = "0E6251"
_ROJO_FILL   = "FADBD8"
_ROJO_TXT    = "78281F"
_AMBAR_FILL  = "FCF3CF"
_AMBAR_TXT   = "7D6608"
_AZUL_FILL   = "E8F8F5"
_AZUL_TXT    = "1B4F72"


def _font(size=10, bold=False, color=_TEXTO_OSCURO):
    return Font(name="Segoe UI", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color=_GRIS_BORDE)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center", wrap_text=True)


def build_executive_excel_report(
    rows_data: list[dict],
    summary: dict,
    output_path: str,
    modulo_nombre: str = "Automatización Perú Compras",
    headers_config: list[tuple[str, str, str, int]] | None = None
) -> tuple[bool, str]:
    """
    Genera un informe Excel corporativo de 2 hojas: 'Resumen Ejecutivo' y 'Detalle de Resultados'.
    
    headers_config: Lista de tuplas (key_dict, Titulo_Columna, alineacion ['left','center','right'], ancho_minimo)
    """
    if not _HAS_OPENPYXL:
        return False, "Error: 'openpyxl' no está instalado en el entorno."

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # HOJA 1: RESUMEN EJECUTIVO / DASHBOARD
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        ws_resumen = wb.create_sheet("Resumen Ejecutivo")
        ws_resumen.sheet_view.showGridLines = True

        for col, width in zip("ABCDE", [26, 20, 24, 20, 20]):
            ws_resumen.column_dimensions[col].width = width

        # Banner Principal
        ws_resumen.merge_cells("A1:E2")
        c1 = ws_resumen["A1"]
        c1.value = f"INFORME DE CONTROL Y AUDITORÍA — PERÚ COMPRAS BOT v1.4\n{modulo_nombre.upper()}"
        c1.font = Font(name="Segoe UI", size=13, bold=True, color=_BLANCO)
        c1.fill = _fill(_AZUL_PRINCIPAL)
        c1.alignment = _center()
        ws_resumen.row_dimensions[1].height = 24
        ws_resumen.row_dimensions[2].height = 24

        # Sub-barra de Fecha y Tiempo
        ws_resumen.merge_cells("A3:E3")
        c3 = ws_resumen["A3"]
        ts = summary.get("timestamp", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        tiempo_txt = summary.get("elapsed_time", "—")
        c3.value = f"Generado: {ts}   |   Tiempo total de ejecución: {tiempo_txt}"
        c3.font = _font(size=9, bold=True, color=_AZUL_HEADER)
        c3.fill = _fill("E2E8F0")
        c3.alignment = _center()
        ws_resumen.row_dimensions[3].height = 20

        # Espaciador
        ws_resumen.row_dimensions[4].height = 10

        # Tarjetas KPI de Métricas Clave (5 columnas)
        total = summary.get("total", len(rows_data))
        ok_cnt = summary.get("ok", 0)
        warn_cnt = summary.get("warn", summary.get("dif", 0))
        err_cnt = summary.get("err", summary.get("missing", 0))
        rate = summary.get("rate", (ok_cnt / total * 100.0) if total > 0 else 0.0)

        kpis = [
            ("TOTAL REGISTROS", total, _AZUL_PRINCIPAL, "E2E8F0"),
            ("CORRECTOS (✓)", ok_cnt, _VERDE_TXT, _VERDE_FILL),
            ("DIFERENCIAS / ADVERTENCIAS", warn_cnt, _AMBAR_TXT, _AMBAR_FILL),
            ("ERRORES / NO HALLADOS", err_cnt, _ROJO_TXT, _ROJO_FILL),
            ("TASA DE ÉXITO", f"{rate:.1f}%", _AZUL_HEADER, _AZUL_CLARO),
        ]

        ws_resumen.row_dimensions[5].height = 22
        ws_resumen.row_dimensions[6].height = 36

        for col_i, (kpi_title, kpi_val, text_col, bg_col) in enumerate(kpis, start=1):
            # Título KPI
            t_cell = ws_resumen.cell(row=5, column=col_i, value=kpi_title)
            t_cell.font = Font(name="Segoe UI", size=8, bold=True, color=text_col)
            t_cell.fill = _fill(bg_col)
            t_cell.alignment = _center()
            t_cell.border = _border()

            # Valor KPI
            v_cell = ws_resumen.cell(row=6, column=col_i, value=kpi_val)
            v_cell.font = Font(name="Segoe UI", size=18, bold=True, color=text_col)
            v_cell.fill = _fill(_BLANCO)
            v_cell.alignment = _center()
            v_cell.border = _border()

        # Espaciador
        ws_resumen.row_dimensions[7].height = 14

        # Tabla de Parámetros de Operación
        ws_resumen.merge_cells("A8:E8")
        h_par = ws_resumen["A8"]
        h_par.value = "PARÁMETROS Y METADATOS DE LA OPERACIÓN"
        h_par.font = Font(name="Segoe UI", size=10, bold=True, color=_BLANCO)
        h_par.fill = _fill(_AZUL_HEADER)
        h_par.alignment = _left()
        ws_resumen.row_dimensions[8].height = 24

        parametros = [
            ("Módulo Ejecutado:", modulo_nombre),
            ("Acuerdo Marco:", summary.get("acuerdo", "EXT-CE-2022-5")),
            ("Catálogo Electrónico:", summary.get("catalogo", "—")),
            ("Categoría Seleccionada:", summary.get("categoria", "—")),
            ("Archivo Fuente Excel:", summary.get("excel_file", summary.get("archivo", "—"))),
            ("Usuario del Portal:", summary.get("usuario", "—")),
            ("Duración del Proceso:", summary.get("elapsed_time", "—")),
        ]

        curr_r = 9
        for label, val in parametros:
            ws_resumen.row_dimensions[curr_r].height = 20
            lbl_c = ws_resumen.cell(row=curr_r, column=1, value=label)
            lbl_c.font = _font(size=9, bold=True, color=_TEXTO_OSCURO)
            lbl_c.fill = _fill("F1F5F9")
            lbl_c.border = _border()
            lbl_c.alignment = _left()

            ws_resumen.merge_cells(f"B{curr_r}:E{curr_r}")
            val_c = ws_resumen.cell(row=curr_r, column=2, value=str(val))
            val_c.font = _font(size=9, color=_TEXTO_OSCURO)
            val_c.border = _border()
            val_c.alignment = _left()
            curr_r += 1

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # HOJA 2: DETALLE DE RESULTADOS
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        ws_detalle = wb.create_sheet("Detalle de Resultados")
        ws_detalle.sheet_view.showGridLines = True
        ws_detalle.freeze_panes = "A2"

        # Configuración por defecto de columnas si no se especifica
        if not headers_config:
            headers_config = [
                ("idx",         "N°",                    "center", 6),
                ("parte",       "N° de Parte",           "center", 18),
                ("ficha",       "Ficha Técnica",         "center", 16),
                ("descripcion", "Descripción del Producto", "left", 45),
                ("stock_excel", "Stock Excel",           "center", 14),
                ("stock_portal","Stock Portal",          "center", 14),
                ("precio",      "Precio S/",             "right",  16),
                ("estado",      "Estado / Resultado",    "center", 20),
                ("obs",         "Observaciones / Detalle","left",  35),
            ]

        # Fila 1: Encabezados de tabla
        ws_detalle.row_dimensions[1].height = 28
        for col_i, (k, col_title, align_mode, min_w) in enumerate(headers_config, start=1):
            col_letter = get_column_letter(col_i)
            c = ws_detalle.cell(row=1, column=col_i, value=col_title)
            c.font = Font(name="Segoe UI", size=10, bold=True, color=_BLANCO)
            c.fill = _fill(_AZUL_HEADER)
            c.alignment = _center()
            c.border = _border()
            ws_detalle.column_dimensions[col_letter].width = min_w

        # Filas de datos
        for row_i, r in enumerate(rows_data, start=2):
            ws_detalle.row_dimensions[row_i].height = 20
            # Color alterno sutil para filas
            row_bg = _BLANCO if (row_i % 2 == 0) else _GRIS_FONDO

            st_raw = str(r.get("estado", r.get("resultado", ""))).strip()
            st_low = st_raw.lower()

            # Determinar estilo de píldora para la columna Estado
            if any(w in st_low for w in ("ok", "éxito", "exito", "correcto", "coincide", "✓")):
                st_fill = _fill(_VERDE_FILL)
                st_font = Font(name="Segoe UI", size=9, bold=True, color=_VERDE_TXT)
            elif any(w in st_low for w in ("diferencia", "advertencia", "modificado", "⚠️", "!=")):
                st_fill = _fill(_AMBAR_FILL)
                st_font = Font(name="Segoe UI", size=9, bold=True, color=_AMBAR_TXT)
            elif any(w in st_low for w in ("error", "fallo", "no encontrado", "no hallado", "✕", "x")):
                st_fill = _fill(_ROJO_FILL)
                st_font = Font(name="Segoe UI", size=9, bold=True, color=_ROJO_TXT)
            else:
                st_fill = _fill(row_bg)
                st_font = _font(size=9, color=_TEXTO_OSCURO)

            for col_i, (k, col_title, align_mode, _) in enumerate(headers_config, start=1):
                cell = ws_detalle.cell(row=row_i, column=col_i)

                # Extraer valor según la clave
                if k == "idx":
                    val = row_i - 1
                elif k == "parte":
                    val = r.get("parte", r.get("nro_parte", ""))
                elif k == "ficha":
                    val = r.get("ficha", r.get("ficha_tecnica", r.get("cod_ficha", "")))
                elif k == "descripcion":
                    val = r.get("descripcion", r.get("desc", r.get("marca", "")))
                elif k == "stock_excel":
                    val = r.get("stock_excel", r.get("stock", ""))
                elif k == "stock_portal":
                    val = r.get("stock_portal", "—")
                elif k == "precio":
                    val = r.get("precio", r.get("precio_excel", ""))
                elif k in ("estado", "resultado"):
                    val = st_raw
                else:
                    val = r.get(k, "")

                # Formateo numérico si es precio
                if k == "precio" and isinstance(val, (int, float)):
                    cell.number_format = '"S/ " #,##0.00'
                elif isinstance(val, str) and val.replace(".", "", 1).isdigit() and k == "precio":
                    try:
                        val = float(val)
                        cell.number_format = '"S/ " #,##0.00'
                    except Exception:
                        pass

                cell.value = val
                cell.border = _border()

                # Alineación
                if align_mode == "center":
                    cell.alignment = _center()
                elif align_mode == "right":
                    cell.alignment = _right()
                else:
                    cell.alignment = _left()

                # Aplicar color
                if k in ("estado", "resultado"):
                    cell.fill = st_fill
                    cell.font = st_font
                else:
                    cell.fill = _fill(row_bg)
                    cell.font = _font(size=9, color=_TEXTO_OSCURO)

        # Activar Auto-filtro en la tabla de detalle
        ws_detalle.auto_filter.ref = f"A1:{get_column_letter(len(headers_config))}{len(rows_data)+1}"

        # Ajuste inteligente de ancho de columnas
        for col in ws_detalle.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))
            # Respetar el ancho mínimo especificado en headers_config
            curr_w = ws_detalle.column_dimensions[col_letter].width or 12
            ws_detalle.column_dimensions[col_letter].width = max(curr_w, min(max_len + 3, 50))

        # Guardar archivo
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        wb.save(output_path)
        wb.close()
        return True, output_path

    except Exception as e:
        import traceback
        return False, f"{e}\n{traceback.format_exc()}"
