import os
import datetime
import openpyxl
from openpyxl.styles import PatternFill

YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def find_data_start(ws) -> int:
    """Encuentra la primera fila de datos (header + 1)."""
    for row_idx in range(1, min(ws.max_row + 1, 11)):
        for cell in ws[row_idx]:
            if cell.value is not None and str(cell.value).strip():
                return row_idx + 1
    return 2


def write_colored_results(source_path: str, sheet_name: str, results: list[dict]) -> str:
    """
    Colorea el Excel según los resultados.
    Verde (ok) = 100% match, Amarillo (differ) = falla, Rojo (not_found) = no existe.
    """
    wb = openpyxl.load_workbook(source_path)
    ws = wb[sheet_name]

    data_start = find_data_start(ws)

    for r in results:
        row_idx = r.get("index", 0)
        status = r.get("status", "")
        excel_row = data_start + row_idx

        if status == "ok":
            fill = GREEN
        elif status == "differ":
            fill = YELLOW
        elif status == "certs_already_exist":
            fill = BLUE
        elif status in ("not_found", "pdf_failed", "save_failed", "session_lost"):
            fill = RED
        else:
            fill = RED

        for col_idx in range(1, ws.max_column + 1):
            try:
                ws.cell(row=excel_row, column=col_idx).fill = fill
            except Exception:
                pass

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(source_path)
    out_path = f"{base}_procesado_{ts}{ext}"
    wb.save(out_path)
    wb.close()
    return out_path
