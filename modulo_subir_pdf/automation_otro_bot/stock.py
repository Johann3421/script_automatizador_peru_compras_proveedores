"""
stock.py — Flujo de Actualización de Stock (réplica del otro bot en Playwright)
================================================================================
Replica el modo "stock" del bot PeruCompras (REPLICAR_OTRO_BOT.txt) usando
Playwright sync en vez de Selenium.

Funcionalidades:
- Login con credenciales + OCR de CAPTCHA
- Navegación con truco de retroceso (back() + get(BASE) + get(MEJORA))
- Selección de filtros (Acuerdo > Catálogo > Categoría)
- Búsqueda por número de parte
- Actualización de stock via modal
- Reintentos con backoff (MAX_REINTENTOS=3)
- Reporte Excel con 3 hojas (Resumen, Detalle, Solo Fallidos)
- Clasificación de errores
- Aprendizaje adaptativo (registra tipos de fallo recurrentes)

Diferencias con el bot Selenium original:
- Usa Playwright sync en vez de Selenium WebDriver
- Usa fetch() desde JS para llamadas API
- OCR con pytesseract (igual al original)
- Selectores CSS adaptados (los mismos que el portal)
"""
import re
import os
import sys
import json
import time
import threading
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════

BASE_URL = "https://www.catalogos.perucompras.gob.pe"
LOGIN_URL = f"{BASE_URL}/AccesoGeneral"
MEJORA_URL = f"{BASE_URL}/MejoraBasica"
MEJORA_COBERTURA_URL = f"{BASE_URL}/MejoraCobertura"
MEJORA_PLAZO_URL = f"{BASE_URL}/MejoraPlazo/IndexMejora"

# Defaults del portal
ACUERDO_TEXTO = "EXT-CE-2022-5 COMPUTADORAS DE ESCRITORIO, COMPUTADORAS PORTÁTILES Y ESCÁNERES"
CATALOGO_TEXTO = "COMPUTADORAS DE ESCRITORIO"
CATEGORIA_TEXTO = "MONITOR"

# Tiempos (segundos)
WAIT_NORMAL = 30
WAIT_LARGO = 60
WAIT_CORTO = 10
PAUSA_ENTRE_PRODUCTOS = 2
MAX_REINTENTOS = 3
MAX_INTENTOS_LOGIN = 5

# Selectores del portal
_INPUT_USUARIO = ["#ID_Usuario", "#txtUsuario", "#N_Usuario",
                  "input[name*='Usuario' i]:not([type='hidden'])"]
_INPUT_PASSWORD = ["#Contrasena", "#txtPassword", "#N_Password", "input[type='password']"]
_INPUT_CAPTCHA = ["#CodigoCaptcha", "#txtCaptcha", "input[name*='aptcha' i]:not([type='hidden'])"]
_IMG_CAPTCHA = ["#imgCaptcha", "img[src*='aptcha' i]"]
_BTN_LOGIN = ["button[type='submit']", "#btnLogin", "#btnIngresar",
              "button:has-text('Ingresar')", "input[type='submit']"]

# Estado global
RESULTADOS = []  # [{"Parte", "Stock", "Estado", "Tipo de Fallo", "Descripción", "Duración"}]
STOP_EVENT = None
PAUSA_EVENT = None
ANALIZADOR = None


# ══════════════════════════════════════════════════════════════════════════════
# ADAPTADOR DE LOGGER PARA REUTILIZAR do_login DEL FLUJO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

class _LogAdapter:
    """Adapta una función log_func() simple a la interfaz LogWriter de automation/login.py."""
    def __init__(self, log_func):
        self._log = log_func

    def info(self, msg: str):
        self._log(msg)

    def ok(self, msg: str):
        self._log(msg)

    def error(self, msg: str):
        self._log(msg)

    def warn(self, msg: str):
        self._log(msg)

    def progress(self, current: int, total: int):
        pass

    def done(self, ok_count: int, error_count: int):
        pass


# ══════════════════════════════════════════════════════════════════════════════
# RE-LOGIN AUTOMÁTICO (igual que bulk_subir_pdf.py)
# ══════════════════════════════════════════════════════════════════════════════

def _is_logged_in(page) -> bool:
    """Verifica si seguimos logueados: URL + ausencia de formulario login visible."""
    try:
        url = page.url.lower()
        # 1. Si estamos explícitamente en login/accesogeneral → NO logueado
        if "accesogeneral" in url or "login" in url:
            return False
        # 2. Si el formulario de login está visible AHORA (no momentáneamente) → NO logueado
        # Usar wait_for con state="visible" y timeout corto para confirmar que REALMENTE está ahí
        try:
            user_input = page.locator("#ID_Usuario").first
            # wait_for con timeout=2000 confirma que permanece visible, no es un flash
            user_input.wait_for(state="visible", timeout=2000)
            return False  # formulario visible confirmado = sesión expirada
        except Exception:
            # No apareció o desapareció → no hay formulario → logueado
            pass
        return True
    except Exception:
        return True


def _esta_en_mejorabasica(page) -> bool:
    """Verifica que estemos en la página de MejoraBasica (donde se editan productos)."""
    try:
        url = page.url.lower()
        return "mejorabasica" in url
    except Exception:
        return False


def _tiene_campos_login(page) -> bool:
    """Verifica si la página actual tiene los campos de login visibles."""
    try:
        user_input = page.locator("#ID_Usuario").first
        return user_input.count() > 0 and user_input.is_visible(timeout=2000)
    except Exception:
        return False


def _relogin(page, usuario: str, password: str, log_func, stop_event=None, captcha_bridge=None) -> bool:
    """Vuelve a la pagina de login, se loguea de cero y navega a MejoraBasica.

    Usa automation.login.do_login (la misma funcion del login inicial) porque
    cierra mejor los modales del portal PeruCompras.
    """
    if stop_event is None:
        import threading
        stop_event = threading.Event()
    log_func("🔁 Sesión expirada. Volviendo a login para re-loguear de cero...")
    try:
        if stop_event and stop_event.is_set():
            return False


        # Importar do_login del flujo principal (mismo que login inicial)
        try:
            from automation.login import do_login
        except Exception:
            _root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            if _root not in sys.path:
                sys.path.insert(0, _root)
            from automation.login import do_login

        # 1. Navegar a login usando domcontentloaded (más rápido y evita timeouts por modales)
        log_func("  Navegando a pagina de login...")
        page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60_000)
        if stop_event and stop_event.is_set():
            return False
        time.sleep(2)

        # 2. Login fresco como la primera vez, con do_login (mejor manejo de modales)
        log_func("  Intentando do_login...")
        log_adapter = _LogAdapter(log_func)
        ok = do_login(page, usuario, password, "", log_adapter, stop_event, captcha_bridge, max_retries=5)
        if stop_event and stop_event.is_set():
            return False
        if not ok:
            log_func("  ❌ Login falló tras sesión expirada")
            return False

        # 3. Navegar a MejoraBasica
        log_func("  ✅ Re-login exitoso. Navegando a MejoraBasica...")
        page.goto(MEJORA_URL, wait_until="domcontentloaded", timeout=60_000)
        if stop_event and stop_event.is_set():
            return False
        page.wait_for_load_state("networkidle", timeout=30_000)
        time.sleep(2)
        return True
    except Exception as e:
        log_func(f"  ❌ Re-login falló: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════════════

import sys, os, time, json, re
from datetime import datetime

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
if sys.stderr and hasattr(sys.stderr, "reconfigure"):
    try:
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

def log(msg):
    txt = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    try:
        print(txt)
    except Exception:
        try:
            enc = getattr(sys.stdout, 'encoding', None) or 'utf-8'
            print(txt.encode(enc, errors='replace').decode(enc, errors='replace'))
        except Exception:
            pass



# ══════════════════════════════════════════════════════════════════════════════
# LIMPIEZA DE MODALES
# ══════════════════════════════════════════════════════════════════════════════

_CLEAR_MODALS_JS = r"""
(function() {
    // 1. Click en botones de cierre típicos
    const closeButtons = document.querySelectorAll(
        '[data-dismiss="modal"], .modal-header button.close, .modal .close, ' +
        'button[aria-label="Close"], button[title="Close"], button[title="Cerrar"], ' +
        '._wModal_close, ._wModal_btn_ok, ._wModal_btn_cancel, ' +
        '.swal2-confirm, .swal2-cancel, .swal2-close'
    );
    closeButtons.forEach(btn => { try { btn.click(); } catch(e) {} });

    // 2. Eliminar modales por selector
    const sels = [
        '._wModal', '._wModal_delete', '._wModal_bg',
        '.sweet-alert', '.swal2-container', '.swal2-popup', '.swal2-backdrop',
        '.modal.fade.in', '.modal.fade.show', '.modal.show', '.modal-backdrop',
        '[role="dialog"]', '[aria-modal="true"]', 'dialog[open]'
    ];
    sels.forEach(s => {
        try { document.querySelectorAll(s).forEach(el => el.remove()); } catch(e) {}
    });

    // 3. Resetear body
    document.body.style.overflow = '';
    document.body.style.paddingRight = '';
    document.body.classList.remove('modal-open');

    return true;
})();
"""


def clear_modals(page):
    """Limpia todos los modales colgados."""
    try:
        page.evaluate(_CLEAR_MODALS_JS)
        return True
    except Exception as e:
        log(f"⚠ clear_modals error: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# LOGIN CON OCR
# ══════════════════════════════════════════════════════════════════════════════

def _trigger_materialize(page, input_id):
    """Dispara eventos input/change/blur en un input para Materialize CSS."""
    try:
        page.evaluate(f"""
            var el = document.getElementById('{input_id}');
            if (el) {{
                el.dispatchEvent(new Event('input', {{bubbles: true}}));
                el.dispatchEvent(new Event('change', {{bubbles: true}}));
                el.dispatchEvent(new Event('blur', {{bubbles: true}}));
            }}
        """)
    except Exception as e:
        log(f"⚠ _trigger_materialize error: {e}")


def _solve_captcha(page) -> str | None:
    """OCR del CAPTCHA con 4 thresholds."""
    try:
        import pytesseract
        from PIL import Image
        import io
    except ImportError:
        return None
    try:
        img_el = page.locator("#imgCaptcha").first
        img_el.wait_for(timeout=5000)
        img_bytes = img_el.screenshot()
        img = Image.open(io.BytesIO(img_bytes))
        img = img.resize((img.width * 4, img.height * 4), Image.LANCZOS).convert("L")
        for threshold in [140, 120, 100, 80]:
            binary = img.point(lambda p: 255 if p > threshold else 0)
            text = pytesseract.image_to_string(
                binary,
                config="--psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
            ).strip().upper()
            text = re.sub(r"[^A-Z0-9]", "", text)
            if 4 <= len(text) <= 8:
                return text
    except Exception as e:
        log(f"⚠ OCR error: {e}")
    return None


def _type_field(page, selector_list, value, materialize_id=None):
    """Escribe value en el primer selector que funcione."""
    for sel in selector_list:
        try:
            el = page.locator(sel).first
            if el.count() == 0:
                continue
            el.click()
            el.fill("")  # limpiar
            el.fill(str(value))
            if materialize_id:
                _trigger_materialize(page, materialize_id)
            return True
        except Exception:
            continue
    return False


def login_with_ocr(page, usuario, password, captcha_bridge=None, max_intentos=MAX_INTENTOS_LOGIN, skip_goto=False) -> bool:
    """Login automático con OCR del CAPTCHA.

    Si el OCR no está disponible, retorna False (caller debe caer a login manual).
    skip_goto: si True, asume que ya estamos en LOGIN_URL y no navega de nuevo.
    """
    log(f"🔐 Login automático con OCR ({usuario})")
    if not skip_goto:
        page.goto(LOGIN_URL, wait_until="networkidle", timeout=60_000)
        page.wait_for_load_state("domcontentloaded")
        # Limpiar modales que aparezcan tras cargar la página de login
        clear_modals(page)
    time.sleep(3)

    for intento in range(1, max_intentos + 1):
        if STOP_EVENT and STOP_EVENT.is_set():
            return False
        log(f"🔐 Intento {intento}/{max_intentos}")

        # Limpiar modales antes de CADA paso
        clear_modals(page)

        # Verificar si ya estamos logueados
        url = page.url.lower()
        if "accesogeneral" not in url and "login" not in url:
            log(f"✅ Ya estamos en {page.url}, saltando login")
            return True

        # Tipear credenciales
        if not _type_field(page, _INPUT_USUARIO, usuario, "ID_Usuario"):
            log(f"⚠ Intento {intento}: no se pudo tipear usuario")
            page.goto(LOGIN_URL)
            time.sleep(2)
            continue
        time.sleep(0.3)

        if not _type_field(page, _INPUT_PASSWORD, password, "Contrasena"):
            log(f"⚠ Intento {intento}: no se pudo tipear contraseña")
            page.goto(LOGIN_URL)
            time.sleep(2)
            continue
        time.sleep(0.5)

        # Resolver CAPTCHA
        if captcha_bridge:
            try:
                img_el = page.locator("#imgCaptcha").first
                img_bytes = img_el.screenshot()
                texto = captcha_bridge.request(img_bytes)
                if texto:
                    texto = re.sub(r"[^A-Z0-9]", "", texto.upper())
            except Exception:
                texto = None
        else:
            texto = _solve_captcha(page)

        if not texto:
            log(f"⚠ Intento {intento}: OCR no resolvió CAPTCHA")
            page.goto(LOGIN_URL)
            time.sleep(2)
            continue

        # Tipear CAPTCHA
        if not _type_field(page, _INPUT_CAPTCHA, texto, "CodigoCaptcha"):
            log(f"⚠ Intento {intento}: no se pudo tipear CAPTCHA")
            page.goto(LOGIN_URL)
            time.sleep(2)
            continue
        time.sleep(0.3)

        # Limpiar modales antes del submit
        clear_modals(page)

        # Click submit
        clicked = False
        for sel in _BTN_LOGIN:
            try:
                btn = page.locator(sel).first
                if btn.count() > 0:
                    btn.click(timeout=5_000)
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            # Fallback JS
            try:
                page.evaluate("""
                    var b = document.querySelector('button[type=submit], #btnLogin, #btnIngresar');
                    if (b) b.click();
                """)
            except Exception:
                pass

        time.sleep(2)
        clear_modals(page)
        time.sleep(2)

        # Verificar éxito
        url = page.url.lower()
        if "accesogeneral" not in url and "login" not in url:
            log(f"✅ Login exitoso. URL: {page.url}")
            return True

        # ValidarAcceso -> back() y reintentar
        if "validaracceso" in url:
            log(f"⚠ ValidarAcceso detectada, back()")
            try:
                page.go_back()
                time.sleep(2)
            except Exception:
                pass
            continue

        log(f"⚠ Intento {intento}: seguimos en {url}, reintentando...")
        page.goto(LOGIN_URL)
        time.sleep(2)

    log(f"❌ Login falló tras {max_intentos} intentos")
    return False


# ══════════════════════════════════════════════════════════════════════════════
# VALIDACIÓN DE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _normalizar_parte(valor) -> str:
    if valor is None:
        return ""
    try:
        import math
        if isinstance(valor, float) and math.isnan(valor):
            return ""
    except Exception:
        pass
    return str(valor).strip()


def _normalizar_stock(valor):
    """Convierte a int, retorna None si es inválido."""
    if valor is None:
        return None
    try:
        import math
        if isinstance(valor, float) and math.isnan(valor):
            return None
    except Exception:
        pass
    s = str(valor).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def analizar_excel_stock(excel_path) -> dict:
    """Analiza el Excel de stock. Retorna {valido: bool, df: [...], errores: [...]}"""
    if not os.path.isfile(excel_path):
        return {"valido": False, "df": [], "errores": [f"No existe: {excel_path}"]}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return {"valido": False, "df": [], "errores": [f"Error abriendo Excel: {e}"]}

    # Leer encabezados
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"valido": False, "df": [], "errores": ["Excel vacío"]}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    log(f"📋 Encabezados: {headers}")

    # Detectar columnas
    col_parte = None
    col_stock = None
    col_ficha = None
    for i, h in enumerate(headers):
        h_low = h.lower()
        if col_parte is None and any(k in h_low for k in ["parte", "n_parte", "n° de parte"]):
            col_parte = i
        if col_stock is None and "stock" in h_low:
            col_stock = i
        if col_ficha is None and any(k in h_low for k in ["ficha", "id_catalogo"]):
            col_ficha = i

    errores = []
    if col_parte is None:
        errores.append("Falta columna 'Parte'")
    if col_stock is None:
        errores.append("Falta columna 'Stock'")

    if errores:
        return {"valido": False, "df": [], "errores": errores}

    # Parsear filas
    data = []
    for i, row in enumerate(rows[1:], start=2):
        parte = _normalizar_parte(row[col_parte]) if col_parte < len(row) else ""
        stock = _normalizar_stock(row[col_stock]) if col_stock < len(row) else None
        ficha = ""
        if col_ficha is not None and col_ficha < len(row):
            f = row[col_ficha]
            if f is not None:
                ficha = str(f).strip()

        if not parte and stock is None:
            continue  # fila vacía
        if not parte:
            errores.append(f"Fila {i}: falta 'Parte'")
            continue
        if stock is None:
            errores.append(f"Fila {i}: stock inválido '{row[col_stock]}'")
            continue

        data.append({"Parte": parte, "Stock": stock, "Ficha": ficha})

    wb.close()
    return {"valido": len(data) > 0, "df": data, "errores": errores, "total": len(data)}


# ══════════════════════════════════════════════════════════════════════════════
# NAVEGACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def paso2_navegacion_stock(page):
    """Truco de retroceso + ir a MejoraBasica."""
    try:
        page.go_back()
        time.sleep(2)
    except Exception:
        pass
    try:
        page.goto(BASE_URL, wait_until="networkidle", timeout=60_000)
        time.sleep(2)
    except Exception:
        pass
    try:
        page.goto(MEJORA_URL, wait_until="networkidle", timeout=60_000)
        time.sleep(3)
    except Exception as e:
        log(f"⚠ Error navegando a MejoraBasica: {e}")


def seleccionar_por_texto_flexible(page, select_id, texto_objetivo):
    """Selecciona option que matchea exacto, contiene, o está contenido."""
    # Normalizar acentos
    def normalizar(s):
        if not s:
            return ""
        s = str(s).lower().strip()
        repl = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}
        for a, b in repl.items():
            s = s.replace(a, b)
        return s

    target = normalizar(texto_objetivo)
    options = page.evaluate(f"""
        () => {{
            const sel = document.querySelector('{select_id}');
            if (!sel) return [];
            return Array.from(sel.options).map(o => ({{
                value: o.value, text: o.text, norm: o.text.toLowerCase()
            }}));
        }}
    """)
    if not options:
        return False

    # 1. Match exacto (normalizado)
    for opt in options:
        if normalizar(opt["text"]) == target:
            page.select_option(select_id, value=opt["value"])
            return True

    # 2. Contiene al target
    for opt in options:
        if target in normalizar(opt["text"]):
            page.select_option(select_id, value=opt["value"])
            return True

    # 3. El target contiene al option
    for opt in options:
        if normalizar(opt["text"]) in target:
            page.select_option(select_id, value=opt["value"])
            return True

    # 4. Coincidencia por código o palabra principal (p.ej. "EXT-CE-2022-5")
    words = [w for w in target.split() if len(w) >= 3]
    if words:
        first_code = words[0]
        for opt in options:
            if first_code in normalizar(opt["text"]):
                page.select_option(select_id, value=opt["value"])
                return True

    # 5. Coincidencia directa por ID de value (p.ej. "252" o "11740")
    for opt in options:
        if str(opt["value"]).strip() == str(texto_objetivo).strip():
            page.select_option(select_id, value=opt["value"])
            return True

    return False




def _wait_for_select_options(page, select_id, timeout_ms=30_000):
    """Espera a que un <select> tenga opciones con value no vacio."""
    try:
        page.wait_for_function(f"""
            () => {{
                const sel = document.querySelector('{select_id}');
                if (!sel) return false;
                const opts = Array.from(sel.options).filter(o => o.value && o.value.trim() !== '');
                return opts.length > 0;
            }}
        """, timeout=timeout_ms)
        return True
    except Exception:
        return False


def paso3_filtros_stock(page, acuerdo=ACUERDO_TEXTO, catalogo=CATALOGO_TEXTO, categoria=CATEGORIA_TEXTO):
    """Selecciona Acuerdo > Catálogo > Categoría y espera que cargue la tabla."""
    try:
        # Esperar que el select de acuerdo esté disponible
        page.wait_for_selector("#ajaxAcuerdo, select[name*='cuerdo']", timeout=WAIT_NORMAL * 1000)
        time.sleep(2)
        if not seleccionar_por_texto_flexible(page, "#ajaxAcuerdo", acuerdo):
            log(f"⚠ No se encontró Acuerdo: {acuerdo}")
            return False
        # Esperar a que el catalogo cargue opciones
        time.sleep(3)
        _wait_for_select_options(page, "#ajaxCatalogo")

        if not seleccionar_por_texto_flexible(page, "#ajaxCatalogo", catalogo):
            log(f"⚠ No se encontró Catálogo: {catalogo}")
            return False
        # Esperar a que la categoria cargue opciones
        time.sleep(3)
        _wait_for_select_options(page, "#ajaxCategoria")

        if not seleccionar_por_texto_flexible(page, "#ajaxCategoria", categoria):
            log(f"⚠ No se encontró Categoría: {categoria}")
            return False

        # Esperar a que la tabla y el buscador dinamico carguen (el portal es lento)
        log("⏳ Esperando carga de tabla de productos...")
        time.sleep(3)
        try:
            page.wait_for_selector(
                ".loading, .spinner, .fa-spinner, .progress, .ajax-loading",
                state="detached",
                timeout=60_000,
            )
        except Exception:
            pass
        try:
            page.locator("input[type='search'][aria-controls='TablaProductos']").first.wait_for(
                state="visible", timeout=30_000
            )
        except Exception:
            pass
        time.sleep(2)
        return True
    except Exception as e:
        log(f"⚠ Error en paso3_filtros: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# ACTUALIZACIÓN DE STOCK
# ══════════════════════════════════════════════════════════════════════════════

def clasificar_error(mensaje: str) -> str:
    msg = str(mensaje).lower()
    if "browser has been closed" in msg or "browser was closed" in msg or "connection closed" in msg:
        return "Navegador cerrado"
    if "no se encontraron resultados" in msg:
        return "Producto no encontrado en la tabla"
    if "n_stock" in msg or "campo de stock" in msg:
        return "Modal de stock no se abrió"
    if "timeout" in msg or "timed out" in msg:
        return "Tiempo de espera agotado"
    if "no such element" in msg:
        return "Elemento de página no encontrado"
    if "stale" in msg:
        return "Página cambió durante la operación"
    return "Error inesperado"


def _browser_cerrado(mensaje: str) -> bool:
    """Detecta si el error se debe a que el usuario cerró el navegador."""
    msg = str(mensaje).lower()
    return any(p in msg for p in [
        "browser has been closed",
        "browser was closed",
        "connection closed",
        "target page, context or browser has been closed",
    ])


def _find_exact_matching_row(page, parte: str):
    """
    Busca en la tabla de productos la fila que contenga EXACTAMENTE el número de parte.
    Evita seleccionar por error 'PARTE-1' cuando se busca 'PARTE'.
    """
    target = str(parte).strip().upper()
    target_clean = " ".join(target.split())
    if not target_clean:
        return None

    rows = page.locator("#TablaProductos tbody tr, table tbody tr").all()
    if not rows:
        return None

    # Paso 1: Coincidencia EXACTA celda por celda (td.text == target)
    for r in rows:
        try:
            cells = r.locator("td").all()
            for c in cells:
                c_text = " ".join(c.inner_text(timeout=300).strip().upper().split())
                if c_text == target_clean:
                    return r
        except Exception:
            continue

    # Paso 2: Coincidencia por delimitador de palabra estricto (evita prefijos/sufijos como -1, _v2, etc.)
    pattern = re.compile(r'(?<![A-Z0-9#\-_/])' + re.escape(target_clean) + r'(?![A-Z0-9#\-_/])')
    for r in rows:
        try:
            r_text = " ".join(r.inner_text(timeout=300).strip().upper().split())
            if pattern.search(r_text):
                return r
        except Exception:
            continue

    # Paso 3: Fallback a tr:has-text de Playwright
    try:
        exact_tr = page.locator(f"tr:has-text('{parte}')")
        if exact_tr.count() > 0:
            return exact_tr.first
    except Exception:
        pass

    return None


def actualizar_producto(page, parte: str, stock: int, ficha: str = "", stop_event=None) -> tuple[bool, str]:
    """Actualiza el stock de un producto. Retorna (éxito, mensaje_error)."""
    try:
        if stop_event and stop_event.is_set():
            return False, "Detenido por usuario"
        # Si hay ficha, navegar directo
        if ficha and ficha.strip():
            edit_url = f"{BASE_URL}/t_CatalogoProductoMarca/CatalogoProductoEdit?ID_CatalogoProducto={ficha}&C_EstadoNav=OBSERVADO&C_Moneda=USD"
            page.goto(edit_url, wait_until="networkidle", timeout=60_000)
            if stop_event and stop_event.is_set():
                return False, "Detenido por usuario"
            time.sleep(2)
        else:
            # 0. Asegurar que no hay modal cubriendo la tabla
            try:
                page.evaluate("""() => {
                    document.querySelectorAll('.modal').forEach(mm => {
                        mm.classList.remove('in', 'show');
                        mm.style.display = 'none';
                    });
                    document.querySelectorAll('.modal-backdrop').forEach(b => b.remove());
                    document.body.classList.remove('modal-open');
                    document.body.style.overflow = '';
                }""")
            except Exception:
                pass

            # 0. Limpiar cualquier overlay residual que intercepte clicks
            try:
                page.evaluate("""() => {
                    document.querySelectorAll('#_wModal_bg, ._wModal_bg, ._wModal_holder').forEach(el => el.remove());
                    document.querySelectorAll('.modal.in, .modal.show').forEach(m => {
                        m.classList.remove('in', 'show');
                        m.style.display = 'none';
                    });
                    document.body.classList.remove('modal-open');
                }""")
            except Exception:
                pass

            # ── PASO 1: Vaciar #C_Descripcion ANTES de que la tabla cargue
            # CRITICAL: si el campo superior tiene texto, el portal filtra la tabla
            # a 1 solo resultado. DataTables entonces solo ve ese 1 item y la búsqueda
            # de abajo no encuentra nada. Limpiar via JS directo (sin .fill ni click).
            search_input = page.locator("#C_Descripcion, input[name='C_Descripcion']").first
            buscar_btn   = page.locator("#btnBuscar, button[name='btnBuscar'], .btn-primary:has-text('Buscar'), button:has-text('Iniciar Búsqueda')").first
            try:
                page.evaluate("""() => {
                    const el = document.querySelector('#C_Descripcion, input[name="C_Descripcion"]');
                    if (el && el.value.trim() !== '') {
                        el.value = '';
                        el.dispatchEvent(new Event('input',  {bubbles: true}));
                        el.dispatchEvent(new Event('change', {bubbles: true}));
                    }
                }""")
            except Exception:
                pass

            # ── PASO 2: Esperar a que DataTables cargue la tabla COMPLETA (>1 fila)
            # Si la tabla tiene solo 1 fila, es señal de que #C_Descripcion filtró.
            # Esperamos hasta 15s para que el buscador de DataTables sea visible
            # y la tabla tenga más de 1 fila.
            dynamic_search = page.locator(
                "input[type='search'][aria-controls='TablaProductos'], #TablaProductos_filter input"
            ).first
            has_dynamic = False
            for _ in range(30):
                if stop_event and stop_event.is_set():
                    return False, "Detenido por usuario"
                try:
                    if dynamic_search.count() > 0 and dynamic_search.is_visible():
                        n_rows = page.locator("#TablaProductos tbody tr").count()
                        if n_rows > 1:
                            has_dynamic = True
                            break
                        # Solo 1 fila visible: intentar limpiar el campo superior de nuevo
                        if n_rows == 1:
                            try:
                                page.evaluate("""() => {
                                    const el = document.querySelector('#C_Descripcion, input[name="C_Descripcion"]');
                                    if (el) { el.value = ''; el.dispatchEvent(new Event('input', {bubbles:true})); }
                                }""")
                            except Exception:
                                pass
                except Exception:
                    pass
                time.sleep(0.5)

            if stop_event and stop_event.is_set():
                return False, "Detenido por usuario"

            # ── PASO 3: Buscar el número de parte en el buscador correcto
            if has_dynamic:
                # Usar DataTables search (cuadro de ABAJO) — no toca #C_Descripcion
                try:
                    dynamic_search.click(force=True)
                    dynamic_search.fill("")
                    dynamic_search.fill(str(parte))
                    dynamic_search.press("Enter")
                    dynamic_search.dispatch_event("input")
                    time.sleep(1.0)
                except Exception:
                    has_dynamic = False

            if not has_dynamic:
                # Fallback: usar #C_Descripcion + btnBuscar (solo si DataTables no existe)
                try:
                    if search_input.count() > 0:
                        search_input.fill("")
                        search_input.fill(str(parte))
                        time.sleep(0.5)
                    if buscar_btn.count() > 0:
                        buscar_btn.click(force=True, timeout=5_000)
                    elif search_input.count() > 0:
                        search_input.press("Enter")
                except Exception as e:
                    return False, f"No se pudo escribir en #C_Descripcion o buscar: {e}"
                time.sleep(2)

            # 4. Verificar que la fila encontrada contiene EXACTAMENTE la parte buscada
            row = _find_exact_matching_row(page, parte)
            if not row or row.count() == 0:
                return False, f"No se encontraron resultados para {parte}"

            # Verificar que la fila realmente contiene la parte buscada.
            # ponytail: inner_text() a veces trunca/corta el texto visible; buscamos
            # en cada celda y en el HTML completo de la fila (texto + atributos).
            parte_upper = str(parte).upper()
            try:
                found_parte = False
                # 1) Buscar en cada celda visible
                cells = row.locator("td").all()
                for cell in cells:
                    try:
                        cell_text = cell.inner_text(timeout=1_000).strip()
                        if parte_upper in cell_text.upper():
                            found_parte = True
                            break
                    except Exception:
                        continue

                # 2) Si no esta visible, buscar en el HTML crudo de la fila
                if not found_parte:
                    row_html = row.inner_html(timeout=2_000).upper()
                    if parte_upper in row_html:
                        found_parte = True

                if not found_parte:
                    row_text = ""
                    try:
                        row_text = row.inner_text(timeout=2_000)
                    except Exception:
                        pass
                    return False, f"Fila encontrada no contiene '{parte}'. Texto: {row_text[:300]}"
            except Exception:
                pass

            # 5. Click en "Existencias" de ESA fila específica
            try:
                exist_link = row.locator("a:has-text('Existencias'), button:has-text('Existencias')").first
                if exist_link.count() > 0:
                    exist_link.click(force=True, timeout=5_000)
                else:
                    # Fallback: cualquier link/botón de edición (NO usar .first ciego)
                    row.locator("a, button").first.click(force=True, timeout=5_000)
            except Exception as e:
                return False, f"No se pudo clickear Existencias: {e}"
            time.sleep(2)
            if stop_event and stop_event.is_set():
                return False, "Detenido por usuario"

        # 4b. Esperar a que el modal de edición de existencias esté visible
        try:
            page.wait_for_selector(
                ".modal:visible, [role='dialog']:visible, #MensajeModal:visible",
                timeout=10_000,
            )
        except Exception:
            pass
        time.sleep(1)
        if stop_event and stop_event.is_set():
            return False, "Detenido por usuario"

        # 5. Buscar input de stock EDITABLE (excluir readonly y N_StockAnt)
        # ponytail: input[name='N_StockAnt'] es readonly con el stock viejo;
        # el campo editable suele ser #N_Stock, input[name='N_Stock'] sin sufijo Ant.
        # Si no se encuentra, buscar cualquier input editable con id/name que contenga "stock" o "existencia".
        stock_input = page.locator(
            "#N_Stock:not([readonly]), input[name='N_Stock']:not([readonly])"
        ).first
        if stock_input.count() == 0:
            # fallback: cualquier input editable (no readonly) con id/name que matchee stock/exist
            stock_input = page.locator(
                "input:not([readonly])[id*='tock' i]:not([id$='Ant']):not([name$='Ant']),"
                "input:not([readonly])[name*='tock' i]:not([id$='Ant']):not([name$='Ant']),"
                "input:not([readonly])[id*='xist' i],"
                "input:not([readonly])[name*='xist' i]"
            ).first
        if stock_input.count() == 0:
            # Diagnóstico: loguear qué inputs hay en el modal
            try:
                all_inputs = page.locator("input").all()
                diag = []
                for inp in all_inputs[:15]:
                    attrs = inp.evaluate(
                        "el => ({id: el.id, name: el.name, readonly: el.readOnly, "
                        "value: el.value, type: el.type, visible: el.offsetParent !== null})"
                    )
                    diag.append(str(attrs))
                log_func(f"   🔍 Inputs visibles en modal: {diag}")
            except Exception:
                pass
            return False, "Campo de stock editable no encontrado"

        # 6. Limpiar y escribir nuevo stock (sanitizado)
        try:
            clean_stock = str(int(float(str(stock).replace(',', '').strip())))
        except Exception:
            clean_stock = str(stock).strip()
        stock_input.fill("")
        stock_input.fill(clean_stock)
        time.sleep(0.5)

        if stop_event and stop_event.is_set():
            return False, "Detenido por usuario"

        # 7. Click Guardar (botón submit del form formStock)
        # ponytail: usar #btn_guardar (id real en ExistenciasModal). No usar force=True
        # porque es submit de formulario.
        save_btn = page.locator("#btn_guardar").first
        if save_btn.count() == 0:
            return False, "Botón Guardar no encontrado"
        try:
            save_btn.click(timeout=10_000)
        except Exception:
            # Si está tapado, forzamos click (aunque puede no hacer submit real)
            save_btn.click(force=True, timeout=3_000)
        if stop_event and stop_event.is_set():
            return False, "Detenido por usuario"

        # 7b. Modal de CONFIRMACIÓN: esperar el _wModal (el modal real del portal)
        # ponytail: NO es #divPopUpConfirmacion; el modal real es ._wModal_holder con ._wModal_btn_ok "Sí"
        try:
            page.wait_for_selector(
                "._wModal_holder:visible, ._wModal_delete:visible",
                timeout=15_000,
            )
            # Clickear "Sí" del modal de confirmación
            page.locator("._wModal_btn_ok, ._wModal_btn_blue").first.click(
                force=True, timeout=3_000
            )
        except Exception:
            # Fallback: buscar por texto Sí dentro del modal
            try:
                page.locator("._wModal:has-text('Está seguro') ._wModal_btn_ok, ._wModal:has-text('seguro') ._wModal_btn_blue").first.click(
                    force=True, timeout=3_000
                )
            except Exception:
                pass

        time.sleep(2)
        if stop_event and stop_event.is_set():
            return False, "Detenido por usuario"

        # 8. Modal de ÉXITO: esperar #MensajeModal y cerrarlo con #btnSalir
        try:
            page.wait_for_selector(
                "#MensajeModal .modal-body:has-text('Actualización'), #MensajeModal.in, #MensajeModal[style*='display: block']",
                timeout=15_000,
            )
            page.locator("#MensajeModal #btnSalir, #MensajeModal button[name='btnCerrar']").first.click(
                force=True, timeout=3_000
            )
        except Exception:
            # Fallback final: forzar cierre de #MensajeModal
            try:
                page.evaluate("""() => {
                    const m = document.getElementById('MensajeModal');
                    if (m) {
                        m.classList.remove('in', 'show');
                        m.style.display = 'none';
                        m.setAttribute('aria-hidden', 'true');
                    }
                }""")
            except Exception:
                pass
        if stop_event and stop_event.is_set():
            return False, "Detenido por usuario"

        time.sleep(1)

        return True, ""
    except Exception as e:
        return False, str(e)


def _get_field(row, keys, default=""):
    if isinstance(row, dict):
        for k in keys:
            if k in row and row[k] is not None:
                return row[k]
        row_lower = {str(k).lower().strip(): v for k, v in row.items()}
        for k in keys:
            kl = str(k).lower().strip()
            if kl in row_lower and row_lower[kl] is not None:
                return row_lower[kl]
    return default


def paso4_actualizar_stock(page, df: list, pausa: float = PAUSA_ENTRE_PRODUCTOS,
                           log_func=None, usuario: str = "", password: str = "",
                           captcha_bridge=None, acuerdo: str = ACUERDO_TEXTO,
                           catalogo: str = CATALOGO_TEXTO, categoria: str = CATEGORIA_TEXTO) -> int:

    """Itera el DataFrame y actualiza cada producto. Retorna cantidad de éxitos."""
    if log_func is None:
        log_func = log

    exitos = 0
    total = len(df)

    for i, row in enumerate(df, 1):
        if STOP_EVENT and STOP_EVENT.is_set():
            log_func(f"⏹ Detención solicitada")
            break

        # ── Check session / página correcta antes de cada producto ──
        # Solo re-loginear si: no estamos logueados O no estamos en MejoraBasica
        # NO re-loginear por fallos de producto (producto no encontrado != sesión expirada)
        if not _is_logged_in(page) or not _esta_en_mejorabasica(page):
            if usuario and password:
                log_func("🔁 No estamos en MejoraBasica o sesión expirada. Re-logueando...")
                if not _relogin(page, usuario, password, log_func, STOP_EVENT, captcha_bridge):
                    log_func("❌ Re-login fallido, deteniendo")
                    break
                # Re-aplicar filtros tras re-login (usar los mismos que al inicio)
                if not paso3_filtros_stock(page, acuerdo, catalogo, categoria):
                    log_func("❌ No se pudieron reaplicar filtros tras re-login")
                    break
            else:
                log_func("❌ Sesión perdida y no hay credenciales para re-loguear")
                break

        # Pausa si está activada
        if PAUSA_EVENT and not PAUSA_EVENT.is_set():
            log_func(f"⏸ En pausa ({i}/{total})")
            while PAUSA_EVENT and not PAUSA_EVENT.is_set():
                if STOP_EVENT and STOP_EVENT.is_set():
                    break
                time.sleep(0.5)

        if STOP_EVENT and STOP_EVENT.is_set():
            log_func("⏹ Detención solicitada durante pausa, saliendo...")
            break

        parte = str(_get_field(row, ["Parte", "parte", "PARTE", "N° Parte", "num_parte", "codigo"], default="")).strip()
        stock = _get_field(row, ["Stock", "stock", "STOCK", "Cantidad", "cantidad"], default=0)
        ficha = str(_get_field(row, ["Ficha", "ficha", "FICHA", "id", "N° Ficha"], default="")).strip()



        log_func(f"📦 [{i}/{total}] {parte} (stock={stock}, ficha={ficha})")
        t0 = time.time()

        exito = False
        error_msg = ""
        for reintento in range(1, MAX_REINTENTOS + 1):
            if STOP_EVENT and STOP_EVENT.is_set():
                break
            exito, error_msg = actualizar_producto(page, parte, stock, ficha, STOP_EVENT)
            if exito:
                break
            # Si el usuario cerró el navegador, no tiene sentido reintentar
            if _browser_cerrado(error_msg):
                log_func("   ⏹ Navegador cerrado, cancelando reintentos...")
                break
            log_func(f"   ⚠ Reintento {reintento}/{MAX_REINTENTOS}: {error_msg}")
            # Sleep interrumpible por STOP_EVENT
            for _ in range(4):
                if STOP_EVENT and STOP_EVENT.is_set():
                    break
                time.sleep(0.5)

        # Si se pidió detención o el browser cerró, intentar recuperación o salir
        if STOP_EVENT and STOP_EVENT.is_set():
            log_func("⏹ Detención solicitada, saliendo...")
            break
        if not exito and (_browser_cerrado(error_msg) or "session" in error_msg.lower() or "expir" in error_msg.lower()):
            if usuario and password:
                log_func("🔁 Sesión o conexión interrumpida. Re-logueando automáticamente...")
                if _relogin(page, usuario, password, log_func, STOP_EVENT, captcha_bridge):
                    log_func("  Re-aplicando filtros...")
                    paso3_filtros_stock(page, acuerdo, catalogo, categoria)
                    continue
            log_func("⏹ No se pudo recuperar la sesión, deteniendo flujo...")
            break


        duracion = time.time() - t0
        tipo_fallo = "" if exito else clasificar_error(error_msg)
        estado = "OK" if exito else "FALLO"

        RESULTADOS.append({
            "Parte": parte, "Stock": stock, "Ficha": ficha,
            "Estado": estado, "Tipo de Fallo": tipo_fallo,
            "Descripción": error_msg, "Duración (seg)": round(duracion, 1),
        })

        if exito:
            exitos += 1
            log_func(f"   ✅ OK en {duracion:.1f}s")
        else:
            log_func(f"   ❌ FALLO: {tipo_fallo}")

        time.sleep(pausa)

    return exitos


# ══════════════════════════════════════════════════════════════════════════════
# REPORTES EXCEL
# ══════════════════════════════════════════════════════════════════════════════

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def generar_reporte_excel(output_path: str, acuerdo="", catalogo="", categoria="") -> str:
    """Genera el reporte Excel con 3 hojas."""
    if not RESULTADOS:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1["A1"] = "Reporte de Actualización de Stock"
        ws1["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws1["A3"] = f"Acuerdo: {acuerdo}"
        ws1["A4"] = f"Catálogo: {catalogo}"
        ws1["A5"] = f"Categoría: {categoria}"
        ws1["A7"] = "Estado del Proceso"
        ws1["B7"] = "No se registraron productos o el proceso finalizó antes de procesar la lista."
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path


    wb = openpyxl.Workbook()
    total = len(RESULTADOS)
    exitos = sum(1 for r in RESULTADOS if r["Estado"] == "OK")
    fallidos = total - exitos
    ratio = (exitos / total * 100) if total > 0 else 0

    # ── Hoja 1: Resumen ──
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1["A1"] = "Reporte de Actualización de Stock"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A3"] = f"Acuerdo: {acuerdo}"
    ws1["A4"] = f"Catálogo: {catalogo}"
    ws1["A5"] = f"Categoría: {categoria}"

    ws1["A7"] = "Total productos"
    ws1["B7"] = total
    ws1["A8"] = "Exitosos"
    ws1["B8"] = exitos
    ws1["A9"] = "Fallidos"
    ws1["B9"] = fallidos
    ws1["A10"] = "Ratio éxito"
    ws1["B10"] = f"{ratio:.1f}%"

    # Gráfico de pie
    pie = PieChart()
    pie.title = "Éxito vs Fallo"
    # Escribir datos en celdas para que el chart pueda referenciarlos
    ws1["D7"] = "Éxitos"
    ws1["E7"] = exitos
    ws1["D8"] = "Fallos"
    ws1["E8"] = fallidos
    pie_data = Reference(ws1, min_col=5, min_row=7, max_col=5, max_row=8)
    pie_cats = Reference(ws1, min_col=4, min_row=7, max_col=4, max_row=8)
    pie.add_data(pie_data)
    pie.set_categories(pie_cats)
    ws1.add_chart(pie, "G7")

    # Conteo por tipo de fallo
    tipos = {}
    for r in RESULTADOS:
        if r["Estado"] == "FALLO":
            t = r["Tipo de Fallo"] or "Sin clasificar"
            tipos[t] = tipos.get(t, 0) + 1
    if tipos:
        ws1["A13"] = "Tipo de Fallo"
        ws1["B13"] = "Cantidad"
        ws1["A13"].fill = HEADER_FILL
        ws1["B13"].fill = HEADER_FILL
        ws1["A13"].font = HEADER_FONT
        ws1["B13"].font = HEADER_FONT
        for i, (t, n) in enumerate(sorted(tipos.items(), key=lambda x: -x[1])):
            ws1.cell(row=14 + i, column=1, value=t)
            ws1.cell(row=14 + i, column=2, value=n)

        # Gráfico de barras
        bar = BarChart()
        bar.title = "Fallos por tipo"
        bar.type = "col"
        data_ref = Reference(ws1, min_col=2, min_row=14, max_col=2, max_row=13 + len(tipos))
        cats_ref = Reference(ws1, min_col=1, min_row=14, max_col=1, max_row=13 + len(tipos))
        bar.add_data(data_ref)
        bar.set_categories(cats_ref)
        ws1.add_chart(bar, "D13")

    # ── Hoja 2: Detalle por Producto ──
    ws2 = wb.create_sheet("Detalle por Producto")
    headers = ["#", "Parte", "Stock", "Ficha", "Estado", "Tipo de Fallo",
               "Descripción", "Duración (seg)"]
    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(RESULTADOS, start=2):
        ws2.cell(row=i, column=1, value=i - 1)
        ws2.cell(row=i, column=2, value=r["Parte"])
        ws2.cell(row=i, column=3, value=r["Stock"])
        ws2.cell(row=i, column=4, value=r.get("Ficha", ""))
        ws2.cell(row=i, column=5, value=r["Estado"])
        ws2.cell(row=i, column=6, value=r["Tipo de Fallo"])
        ws2.cell(row=i, column=7, value=r["Descripción"])
        ws2.cell(row=i, column=8, value=r["Duración (seg)"])
        fill = GREEN if r["Estado"] == "OK" else RED
        for c in range(1, 9):
            ws2.cell(row=i, column=c).fill = fill

    # Auto-ajustar columnas
    for c in range(1, 9):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # ── Hoja 3: Solo Fallidos ──
    ws3 = wb.create_sheet("Solo Fallidos")
    fallidos_list = [r for r in RESULTADOS if r["Estado"] == "FALLO"]
    for c, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, r in enumerate(fallidos_list, start=2):
        ws3.cell(row=i, column=1, value=i - 1)
        ws3.cell(row=i, column=2, value=r["Parte"])
        ws3.cell(row=i, column=3, value=r["Stock"])
        ws3.cell(row=i, column=4, value=r.get("Ficha", ""))
        ws3.cell(row=i, column=5, value=r["Estado"])
        ws3.cell(row=i, column=6, value=r["Tipo de Fallo"])
        ws3.cell(row=i, column=7, value=r["Descripción"])
        ws3.cell(row=i, column=8, value=r["Duración (seg)"])
        for c in range(1, 9):
            ws3.cell(row=i, column=c).fill = RED
    for c in range(1, 9):
        ws3.column_dimensions[get_column_letter(c)].width = 18

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# FLUJO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def ejecutar_stock(page, excel_path, usuario, password, acuerdo=ACUERDO_TEXTO,
                  catalogo=CATALOGO_TEXTO, categoria=CATEGORIA_TEXTO,
                  pausa=PAUSA_ENTRE_PRODUCTOS, captcha_bridge=None,
                  log_func=None) -> str:
    """Ejecuta el flujo completo de stock. Retorna path del reporte."""
    global RESULTADOS, STOP_EVENT
    RESULTADOS = []

    if log_func is None:
        log_func = log

    log_func("=" * 60)
    log_func("MODO STOCK - Actualización de existencias")
    log_func("=" * 60)

    # 1. Validar Excel
    log_func(f"📋 Validando Excel: {excel_path}")
    validacion = analizar_excel_stock(excel_path)
    if not validacion["valido"]:
        log_func(f"❌ Excel inválido: {validacion['errores']}")
        return ""
    df = validacion["df"]
    log_func(f"✅ {len(df)} productos cargados")

    # 2. Login
    if not login_with_ocr(page, usuario, password, captcha_bridge):
        log_func("❌ Login falló")
        return ""

    # 3. Navegación
    log_func("📍 Navegando a MejoraBasica...")
    paso2_navegacion_stock(page)

    # 4. Filtros
    log_func(f"📋 Aplicando filtros: {acuerdo} > {catalogo} > {categoria}")
    if not paso3_filtros_stock(page, acuerdo, catalogo, categoria):
        log_func("❌ No se pudieron aplicar los filtros")
        return ""

    # 5. Actualizar stock
    log_func(f"📦 Actualizando {len(df)} productos...")
    exitos = paso4_actualizar_stock(page, df, pausa, log_func, usuario, password, captcha_bridge,
                                   acuerdo, catalogo, categoria)
    log_func(f"✅ {exitos}/{len(df)} actualizados")

    # 6. Generar reporte
    output_dir = os.path.dirname(excel_path) or "."
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(output_dir, f"reporte_stock_{ts}.xlsx")
    generar_reporte_excel(report_path, acuerdo, catalogo, categoria)
    log_func(f"📊 Reporte: {report_path}")

    return report_path
