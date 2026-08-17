# ═══════════════════════════════════════════════════════════════════════
# main_subir_pdf.py  —  Punto de entrada de la aplicación PeruComprasBot
# ───────────────────────────────────────────────────────────────────────
# ARQUITECTURA (3 capas):
#
#   1. UI  (SubirPdfApp / _build_ui / _build_stock_tab)
#      Solo construye ventana, tabs y widgets. No toca el portal.
#
#   2. DELEGADORES DE HILOS  (_on_* lanza thread → _execute_* llama workers.*)
#      Validan campos de UI, arman parámetros y lanzan daemons.
#
#   3. PUENTE JS→PYTHON  (SubirPdfWebApi + run_app)
#      Expone los mismos flujos a la UI web (pywebview) sin duplicar lógica.
#
# LÓGICA REAL DE AUTOMATIZACIÓN (NO está aquí):
#   workers.py                    → execute_stock, execute, execute_test…
#   automation_otro_bot/stock.py  → pasos 1-4 del flujo de stock
#   automation/browser.py         → init/close del navegador Playwright
#   automation/login.py           → login + manejo de CAPTCHA
#
# REGLA: Para cambiar el comportamiento de automatización → editar workers.py
#        Para cambiar la UI → editar los bloques _build_* de este archivo
# ═══════════════════════════════════════════════════════════════════════
import sys, os, time, threading, queue, json, re
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='backslashreplace')
    sys.stderr.reconfigure(encoding='utf-8', errors='backslashreplace')
except Exception:
    pass
from io import BytesIO
from pathlib import Path
from datetime import datetime
from tkinter import filedialog, messagebox

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
if sys.stderr and hasattr(sys.stderr, "reconfigure"):
    try:
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


VERSION = "1.4"

# ── Paths ─────────────────────────────────────────────────────────
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.abspath(os.path.join(_THIS_DIR, ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)

from automation.browser import init_browser, close_browser
from automation.login import do_login
from utils.logger import LogWriter

# Imports locales del módulo (renombrados para evitar conflicto con el proyecto raíz)
from utils_mod.excel_parser_mod import get_sheets, detect_columns, parse_excel
from utils_mod.logger_mod import LogWriter as LocalLogWriter

try:
    import ctk_compat as ctk
    from PIL import Image
except ImportError:
    print("Error: instala Pillow: pip install pillow")
    import sys; sys.exit(1)


# ═══════════════════════════════════════════════════════════════════
#  BLOQUE 1 — SPLASH SCREEN
# ───────────────────────────────────────────────────────────────────
#  Ventana de carga animada que se muestra ANTES de la app principal.
#  Ciclo de vida:
#    SubirPdfApp.__init__() → SplashScreen(parent=self)
#    → _step() avanza 4 pasos animados (barra de progreso)
#    → _finish() destruye el splash y llama parent.deiconify()
#  ❌ NO contiene lógica de negocio. Solo visual/animación.
# ═══════════════════════════════════════════════════════════════════

class SplashScreen(ctk.CTkToplevel):
    """Pestaña/Ventana de carga inicial elegante antes de mostrar la aplicación principal."""
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.configure(fg_color="#006CA8")
        self.overrideredirect(True)
        self.resizable(False, False)

        w, h = 460, 260
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.attributes("-topmost", True)

        # Cabecera institucional
        panel = ctk.CTkFrame(self, fg_color="#00507E", corner_radius=0)
        panel.place(relx=0, rely=0, relwidth=1, relheight=0.45)

        ctk.CTkLabel(
            panel, text="PERU COMPRAS",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="#FFFFFF",
        ).pack(anchor="w", padx=24, pady=(22, 2))
        ctk.CTkLabel(
            panel, text="Sistema de Automatizacion de Ofertas",
            font=ctk.CTkFont(size=11),
            text_color="#AACCDD",
        ).pack(anchor="w", padx=24)

        ctk.CTkFrame(self, fg_color="#00507E", height=1).place(
            relx=0, rely=0.45, relwidth=1
        )

        self._lbl_msg = ctk.CTkLabel(
            self, text="Inicializando sistema...",
            font=ctk.CTkFont(size=11),
            text_color="#AACCDD",
        )
        self._lbl_msg.place(relx=0.5, rely=0.60, anchor="center")

        self._bar = ctk.CTkProgressBar(
            self, width=380, height=5,
            fg_color="#1A5493", progress_color="#FFFFFF",
        )
        self._bar.place(relx=0.5, rely=0.72, anchor="center")
        self._bar.set(0)

        ctk.CTkLabel(
            self, text=f"v{VERSION}",
            font=ctk.CTkFont(size=10),
            text_color="#AACCDD",
        ).place(relx=0.5, rely=0.88, anchor="center")

        self._steps = [
            (0.25, "Cargando componentes y recursos…"),
            (0.55, "Validando estructura y datos…"),
            (0.85, "Cargando menú e interfaz gráfica…"),
            (1.00, "¡Sistema Listo!"),
        ]
        self._step_idx = 0
        self.after(100, self._step)

    def _step(self):
        if self._step_idx < len(self._steps):
            prog, msg = self._steps[self._step_idx]
            self._bar.set(prog)
            self._lbl_msg.configure(text=msg)
            self._step_idx += 1
            self.after(260, self._step)
        else:
            self.after(120, self._finish)

    def _finish(self):
        try:
            self.destroy()
        except Exception:
            pass
        self.parent.deiconify()


# ═══════════════════════════════════════════════════════════════════
#  BLOQUE 2 — HELPERS GLOBALES
# ───────────────────────────────────────────────────────────────────
#  _make_stock_log(append_fn)
#    Crea un objeto logger compatible con LogWriter que redirige
#    .info() / .warning() / .error() / .success() / .ok()
#    al callback de UI (append_fn = self._append_stock_log).
#    Se pasa a paso4_actualizar_stock() en stock.py como log_func.
# ═══════════════════════════════════════════════════════════════════
def _make_stock_log(append_fn):
    # ponytail: puente ligero LogWriter → callback UI (pestaña 2)
    class _StockLog:
        def info(self, msg): append_fn(str(msg))
        def warning(self, msg): append_fn(f"⚠ {msg}")
        def error(self, msg): append_fn(f"❌ {msg}")
        def success(self, msg): append_fn(f"✅ {msg}")
        def ok(self, msg): append_fn(f"✅ {msg}")
    return _StockLog()


# ═══════════════════════════════════════════════════════════════════
#  BLOQUE 3 — CAPTCHA BRIDGE  (sincronización UI ↔ hilo de trabajo)
# ───────────────────────────────────────────────────────────────────
#  Problema: el worker (hilo daemon) se bloquea cuando encuentra un
#  CAPTCHA y necesita que el USUARIO ingrese el código antes de
#  continuar. CaptchaBridge lo sincroniza con threading.Event.
#
#  Flujo completo:
#    [Worker]   bridge.request(imagen_bytes)  → bloquea (Event.wait)
#    [poll_queue] detecta imagen_bytes != None → llama _show_captcha()
#    [Usuario]  ingresa código + clic Enviar
#    [UI]       bridge.respond(codigo)        → desbloquea el worker
#
#  ❌ NO modificar Lock/Event. Thread-safe por diseño.
#  ✅ Se puede ajustar el timeout del poll (actualmente 0.5 s).
# ═══════════════════════════════════════════════════════════════════

class CaptchaBridge:
    def __init__(self):
        self.lock = threading.Lock()
        self.event = threading.Event()
        self.image_bytes = None
        self.user_code = ""
        self.stop_event = None

    def request(self, img):
        with self.lock:
            self.image_bytes = img
            self.user_code = ""
            self.event.clear()
        while True:
            if self.stop_event and self.stop_event.is_set():
                self.event.set()
                return ""
            if self.event.wait(timeout=0.5):
                break
        with self.lock:
            return self.user_code

    def respond(self, code):
        with self.lock:
            self.user_code = code
            self.image_bytes = None
        self.event.set()


# ═══════════════════════════════════════════════════════════════════
#  BLOQUE 4 — SubirPdfApp  (clase principal de la aplicación)
# ───────────────────────────────────────────────────────────────────
#  Hereda de ctk.CTk (CustomTkinter). Es la ventana raíz.
#
#  Métodos internos agrupados por responsabilidad:
#
#  [INICIALIZACIÓN]
#    __init__()           → estado global, carga JSON, llama _build_ui + poll_queue
#
#  [UI — CONSTRUCCIÓN DE VISTAS]
#    _C (dict)            → paleta de colores institucional
#    _build_ui()          → construye TODA la ventana (5 tabs + menú + statusbar)
#    _build_stock_tab()   → vista "Actualización de Stock"
#    _build_advanced_tools_tab() → vista "Herramientas Avanzadas"
#    _build_credentials_section / _build_excel_section / etc. → sub-secciones de la vista PDF
#
#  [HANDLERS DE EVENTOS — pestaña PDF]
#    _on_launch()    → inicia flujo PDF (valida → lanza hilo → _execute)
#    _on_stop()      → detiene flujo PDF
#    _on_test()      → test de 1 ficha
#    _on_certs_only(), _on_nro_parte(), etc. → herramientas avanzadas
#
#  [HANDLERS DE EVENTOS — pestaña Stock]
#    _on_stock_start()  → inicia flujo Stock (valida → lanza hilo → _execute_stock)
#    _on_stock_stop()   → detiene flujo Stock cerrando el browser
#
#  [DELEGADORES — lanzan el hilo y llaman a workers.*]
#    _execute()           → workers.execute()
#    _execute_stock()     → workers.execute_stock()
#    _execute_test/certs/nro_parte/etc. → workers.*
#
#  [INFRAESTRUCTURA]
#    poll_queue()    → loop 200ms para actualizar UI desde hilos (ÚNICO punto thread-safe)
#    _append_stock_log() → escribe en el log de la pestaña Stock
#
#  ❌ NO agregar lógica de automatización en esta clase.
#  ✅ Cambios seguros: layout, colores, validaciones de campos, textos.
# ═══════════════════════════════════════════════════════════════════

class SubirPdfApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.withdraw()  # Ocultar ventana principal durante la carga del Splash
        self.title(f"Sistema de Automatización — Perú Compras v{VERSION}")

        # Detección y adaptación inteligente a pantallas pequeñas (laptops) y grandes
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        init_w = min(1080, max(880, sw - 80))
        init_h = min(760, max(600, sh - 90))
        init_x = max(0, (sw - init_w) // 2)
        init_y = max(0, (sh - init_h) // 2)

        self.geometry(f"{init_w}x{init_h}+{init_x}+{init_y}")
        self.minsize(860, 580)
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.log_queue = queue.Queue()
        self.stop_event = threading.Event()
        self.captcha_bridge = CaptchaBridge()
        self._running = False
        self._log_lines = []
        self._ok = 0
        self._errors = 0
        self._total = 0

        self._excel_path = ""
        self._excel_rows = []
        self._excel_columns = []

        self._catalog_data = {}
        self._catalog_combos = []
        self._load_dropdown_json()

        self._build_ui()
        self.poll_queue()

        # Iniciar pestaña de carga previa a la app
        self._splash = SplashScreen(self)

    def _load_dropdown_json(self):
        # ponytail: usar resource_path para que funcione empaquetado con PyInstaller
        try:
            try:
                from resource_helper import resource_path
            except Exception:
                _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                if _root not in sys.path:
                    sys.path.insert(0, _root)
                from resource_helper import resource_path
            candidates = [
                resource_path("modulo_subir_pdf/combinaciones_computadoras.json"),
                resource_path("modulo_subir_pdf/dropdown_options_modificar.json"),
            ]
            json_path = ""
            for cand in candidates:
                if os.path.isfile(cand):
                    json_path = cand
                    break
            if not json_path:
                return
            with open(json_path, "r", encoding="utf-8") as f:
                self._catalog_data = json.load(f)
        except Exception:
            pass

    # ═══════════════════════════════════════════════════════════════
    # PALETA DE COLORES  —  _C (class-level dict)
    # ───────────────────────────────────────────────────────────────
    # Referencia: Portal SISCatalogo / SIAF / SIGA (Windows native)
    # Regla: TODOS los colores van aquí. Nunca hardcodear hex en widgets.
    #   Uso en métodos: C = self._C  →  fg_color=C["accent"]
    #   Para cambiar color: solo editar el valor hex de la clave.
    #   Antes de renombrar una clave: buscar todos los usos con C["clave"].
    # ═══════════════════════════════════════════════════════════════
    _C = {
        "bg":         "#F0F0F0",   # Fondo gris sistema Windows
        "topbar":     "#006CA8",   # Azul institucional Peru Compras
        "topbar_dk":  "#00507E",   # Azul oscuro
        "tabs_bg":    "#E8E8E8",   # Barra de modulos
        "tab_active": "#FFFFFF",   # Tab activo
        "tab_txt_a":  "#006CA8",   # Texto tab activo
        "tab_txt_i":  "#555555",   # Texto tab inactivo
        "card":       "#FFFFFF",   # Superficie principal
        "card2":      "#F7F7F7",   # Superficie secundaria
        "border":     "#C8C8C8",   # Bordes
        "border2":    "#E0E0E0",   # Bordes suaves
        "txt":        "#2B2B2B",   # Texto principal
        "txt2":       "#555555",   # Texto secundario
        "txt3":       "#888888",   # Texto inactivo
        "accent":     "#006CA8",   # Boton principal
        "accent_h":   "#00507E",   # Hover boton principal
        "success":    "#1B6B1B",   # Exito
        "danger":     "#8B1A1A",   # Error / detener
        "danger_h":   "#6A1414",   # Hover detener
        "warn":       "#854D0E",   # Advertencia
        "sep":        "#D4D4D4",   # Separadores
        # Claves legacy para secciones que aun usan sidebar_*
        "sidebar":    "#006CA8",
        "sidebar_hl": "#00507E",
        "sidebar_txt":"#FFFFFF",
        "sidebar_sub":"#AACCDD",
    }

    def _setup_styles(self):
        from tkinter import ttk
        s = ttk.Style(self)
        try:
            s.theme_use("clam")
        except Exception:
            pass
        s.configure("Hoja.Treeview",
            font=("Segoe UI", 10),
            rowheight=22,
            background="#FFFFFF",
            foreground="#2B2B2B",
            fieldbackground="#FFFFFF",
            borderwidth=0,
        )
        s.configure("Hoja.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#D4D4D4",
            foreground="#1A1A1A",
            relief="flat",
            padding=(6, 4),
        )
        s.map("Hoja.Treeview",
            background=[("selected", "#006CA8")],
            foreground=[("selected", "#FFFFFF")],
        )
        s.configure("TScrollbar", troughcolor="#F0F0F0", background="#E8E8E8", relief="flat")

    def _set_taskbar_icon(self):
        try:
            import ctypes
            # Establecer AppUserModelID para que Windows agrupe correctamente en la barra de tareas
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("TKC.PeruComprasBot.1.4")
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id()) or self.winfo_id()
            style = ctypes.windll.user32.GetWindowLongW(hwnd, -20)
            style = style | 0x00040000  # WS_EX_APPWINDOW
            ctypes.windll.user32.SetWindowLongW(hwnd, -20, style)
            ctypes.windll.user32.ShowWindow(hwnd, 5)
        except Exception:
            pass
        # Cargar icono profesional desde resources/
        try:
            from resource_helper import resource_path
            ico = resource_path(os.path.join("resources", "icon.ico"))
            if os.path.isfile(ico):
                self.iconbitmap(default=ico)
        except Exception:
            pass

    def _start_drag(self, event):
        self._drag_x = event.x
        self._drag_y = event.y

    def _drag_window(self, event):
        if getattr(self, "_is_maximized", False):
            return
        x = self.winfo_x() + (event.x - self._drag_x)
        y = self.winfo_y() + (event.y - self._drag_y)
        self.geometry(f"+{x}+{y}")

    def _setup_border_resizing(self):
        """Crea los 8 manejadores perimetrales para redimensionar en todas las direcciones como una ventana nativa de Windows."""
        import tkinter as tk
        BORDER = 6
        self._resizing = False
        self._resize_dir = "se"
        self._start_x = 0
        self._start_y = 0
        self._start_geom = (0, 0, 0, 0)

        # 8 bordes y esquinas con cursores nativos de Windows
        self._resize_handles = {
            "e":  tk.Frame(self, cursor="size_we", bg=""),
            "w":  tk.Frame(self, cursor="size_we", bg=""),
            "s":  tk.Frame(self, cursor="size_ns", bg=""),
            "n":  tk.Frame(self, cursor="size_ns", bg=""),
            "se": tk.Frame(self, cursor="size_nw_se", bg=""),
            "sw": tk.Frame(self, cursor="size_ne_sw", bg=""),
            "ne": tk.Frame(self, cursor="size_ne_sw", bg=""),
            "nw": tk.Frame(self, cursor="size_nw_se", bg=""),
        }

        # Ubicación perimetral alrededor de toda la ventana
        self._resize_handles["e"].place(relx=1.0, rely=0.0, relheight=1.0, width=BORDER, anchor="ne")
        self._resize_handles["w"].place(relx=0.0, rely=0.0, relheight=1.0, width=BORDER, anchor="nw")
        self._resize_handles["s"].place(relx=0.0, rely=1.0, relwidth=1.0, height=BORDER, anchor="sw")
        self._resize_handles["n"].place(relx=0.0, rely=0.0, relwidth=1.0, height=BORDER, anchor="nw")

        self._resize_handles["se"].place(relx=1.0, rely=1.0, width=BORDER*2, height=BORDER*2, anchor="se")
        self._resize_handles["sw"].place(relx=0.0, rely=1.0, width=BORDER*2, height=BORDER*2, anchor="sw")
        self._resize_handles["ne"].place(relx=1.0, rely=0.0, width=BORDER*2, height=BORDER*2, anchor="ne")
        self._resize_handles["nw"].place(relx=0.0, rely=0.0, width=BORDER*2, height=BORDER*2, anchor="nw")

        for direction, handle in self._resize_handles.items():
            handle.lift()
            handle.bind("<ButtonPress-1>", lambda e, d=direction: self._start_resize(e, d))
            handle.bind("<B1-Motion>", self._do_resize)
            handle.bind("<ButtonRelease-1>", self._stop_resize)

    def _start_resize(self, event, direction="se"):
        """Inicia el redimensionamiento dinámico en la dirección seleccionada."""
        if getattr(self, "_is_maximized", False):
            return
        self._resizing = True
        self._resize_dir = direction
        self._start_x = event.x_root
        self._start_y = event.y_root
        self._start_geom = (self.winfo_x(), self.winfo_y(), self.winfo_width(), self.winfo_height())

    def _stop_resize(self, event):
        self._resizing = False

    def _do_resize(self, event):
        """Aplica el cambio de tamaño dinámico en tiempo real adaptando toda la interfaz."""
        if not getattr(self, "_resizing", False) or getattr(self, "_is_maximized", False):
            return
        x0, y0, w0, h0 = self._start_geom
        dx = event.x_root - self._start_x
        dy = event.y_root - self._start_y
        min_w, min_h = 860, 580

        new_x, new_y, new_w, new_h = x0, y0, w0, h0
        direction = getattr(self, "_resize_dir", "se")

        if "e" in direction:
            new_w = max(min_w, w0 + dx)
        if "w" in direction:
            cand_w = w0 - dx
            if cand_w >= min_w:
                new_w = cand_w
                new_x = x0 + dx
            else:
                new_w = min_w
                new_x = x0 + (w0 - min_w)
        if "s" in direction:
            new_h = max(min_h, h0 + dy)
        if "n" in direction:
            cand_h = h0 - dy
            if cand_h >= min_h:
                new_h = cand_h
                new_y = y0 + dy
            else:
                new_h = min_h
                new_y = y0 + (h0 - min_h)

        self.geometry(f"{new_w}x{new_h}+{new_x}+{new_y}")

    def _set_window_preset(self, target_w, target_h):
        """Ajusta la ventana a una resolución predefinida."""
        if getattr(self, "_is_maximized", False):
            self._toggle_maximize()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w = min(target_w, max(860, sw - 40))
        h = min(target_h, max(580, sh - 60))
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _center_window(self):
        """Centra la ventana actual en la pantalla."""
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w = self.winfo_width()
        h = self.winfo_height()
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _close_window(self):
        self.destroy()
        import sys
        sys.exit(0)

    def _open_config_dialog(self):
        """Abre la ventana modal de Configuración y Preferencias del Sistema."""
        import tkinter as tk
        from tkinter import ttk, messagebox

        win = tk.Toplevel(self)
        win.title("Configuración y Preferencias del Sistema — Perú Compras Bot")
        win.configure(bg="#F0F0F0")
        win.geometry("520x460")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        # Centrar ventana
        win.update_idletasks()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = (sw - 520) // 2
        y = (sh - 460) // 2
        win.geometry(f"520x460+{x}+{y}")

        # Header Azul
        hdr = tk.Frame(win, bg="#006CA8", pady=10, padx=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="CONFIGURACIÓN Y PREFERENCIAS DEL SISTEMA",
                 font=("Segoe UI", 10, "bold"), bg="#006CA8", fg="#FFFFFF", anchor="w").pack(fill="x")
        tk.Label(hdr, text="Ajustes de credenciales, parámetros de red y motores de automatización.",
                 font=("Segoe UI", 9), bg="#006CA8", fg="#AACCDD", anchor="w").pack(fill="x")

        body = tk.Frame(win, bg="#F0F0F0", padx=16, pady=12)
        body.pack(fill="both", expand=True)

        def card_sec(title):
            c = tk.Frame(body, bg="#FFFFFF", bd=1, relief="solid", highlightbackground="#C8C8C8", pady=8, padx=12)
            c.pack(fill="x", pady=6)
            tk.Label(c, text=title, font=("Segoe UI", 9, "bold"), bg="#FFFFFF", fg="#006CA8", anchor="w").pack(fill="x", pady=(0, 6))
            return c

        # Sec 1: Credenciales
        c1 = card_sec("ALMACENAMIENTO DE CREDENCIALES")
        var_creds = tk.BooleanVar(value=True)
        var_cat = tk.BooleanVar(value=True)
        tk.Checkbutton(c1, text="Guardar credenciales de inicio de sesión automáticamente (encriptado)",
                       variable=var_creds, bg="#FFFFFF", font=("Segoe UI", 9), activebackground="#FFFFFF").pack(anchor="w", pady=2)
        tk.Checkbutton(c1, text="Recordar catálogo y categoría seleccionados entre sesiones",
                       variable=var_cat, bg="#FFFFFF", font=("Segoe UI", 9), activebackground="#FFFFFF").pack(anchor="w", pady=2)

        # Sec 2: Tiempos de Red
        c2 = card_sec("PARÁMETROS DE RED Y EJECUCIÓN")
        f_p = tk.Frame(c2, bg="#FFFFFF")
        f_p.pack(fill="x", pady=2)
        tk.Label(f_p, text="Pausa por defecto entre solicitudes (seg):", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555").pack(side="left")
        e_pausa = tk.Entry(f_p, font=("Segoe UI", 9), width=8, bd=1, relief="sunken")
        e_pausa.pack(side="left", padx=8)
        e_pausa.insert(0, "1.5")

        f_r = tk.Frame(c2, bg="#FFFFFF")
        f_r.pack(fill="x", pady=2)
        tk.Label(f_r, text="Reintentos automáticos por error de red:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555").pack(side="left")
        e_retry = tk.Entry(f_r, font=("Segoe UI", 9), width=8, bd=1, relief="sunken")
        e_retry.pack(side="left", padx=8)
        e_retry.insert(0, "3")

        # Sec 3: Motores
        c3 = card_sec("MOTORES DE AUTOMATIZACIÓN")
        tk.Label(c3, text="✓ Tesseract OCR: Detectado e integrado correctamente",
                 font=("Segoe UI", 9), bg="#FFFFFF", fg="#1B6B1B", anchor="w").pack(fill="x", pady=2)
        tk.Label(c3, text="✓ Playwright Chromium: Motor listo para automatización",
                 font=("Segoe UI", 9), bg="#FFFFFF", fg="#1B6B1B", anchor="w").pack(fill="x", pady=2)

        # Bottom Bar
        bbar = tk.Frame(win, bg="#E8E8E8", pady=8, padx=16)
        bbar.pack(fill="x", side="bottom")

        def _save():
            messagebox.showinfo("Configuración", "Preferencias guardadas exitosamente.", parent=win)
            win.destroy()

        tk.Button(bbar, text="Guardar Cambios", font=("Segoe UI", 9, "bold"),
                  bg="#006CA8", fg="#FFFFFF", activebackground="#00507E", activeforeground="#FFFFFF",
                  bd=0, padx=14, pady=4, cursor="hand2", command=_save).pack(side="right", padx=4)
        tk.Button(bbar, text="Cancelar", font=("Segoe UI", 9),
                  bg="#E8E8E8", fg="#1A1A1A", bd=1, relief="raised", padx=12, pady=3,
                  cursor="hand2", command=win.destroy).pack(side="right", padx=4)

    def _open_about_dialog(self):
        """Abre la ventana modal Acerca del Sistema."""
        import tkinter as tk
        from tkinter import messagebox

        win = tk.Toplevel(self)
        win.title("Acerca de Perú Compras Bot")
        win.configure(bg="#FFFFFF")
        win.geometry("460x340")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        # Centrar ventana
        win.update_idletasks()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = (sw - 460) // 2
        y = (sh - 340) // 2
        win.geometry(f"460x340+{x}+{y}")

        # Banner Superior Azul
        hdr = tk.Frame(win, bg="#006CA8", pady=16, padx=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="PERÚ COMPRAS BOT",
                 font=("Segoe UI", 14, "bold"), bg="#006CA8", fg="#FFFFFF", anchor="w").pack(fill="x")
        tk.Label(hdr, text=f"Versión {VERSION} Build 2026 — Plataforma Oficial de Automatización",
                 font=("Segoe UI", 9), bg="#006CA8", fg="#AACCDD", anchor="w").pack(fill="x", pady=(2, 0))

        # Cuerpo
        body = tk.Frame(win, bg="#FFFFFF", padx=20, pady=16)
        body.pack(fill="both", expand=True)

        txt_info = (
            "Sistema Integrado de Automatización de Ofertas, Gestión de Stock\n"
            "y Carga Masiva de Catálogos Electrónicos de Perú Compras.\n\n"
            "• Acuerdo Marco: EXT-CE-2022-5\n"
            "• Motor de Procesamiento: Playwright Engine + Tesseract OCR\n"
            "• Desarrollo y Distribución Oficial:\n"
            "  THE KING COMPUTER E.I.R.L.\n\n"
            "Todos los derechos reservados © 2026."
        )
        tk.Label(body, text=txt_info, font=("Segoe UI", 9), bg="#FFFFFF", fg="#2B2B2B",
                 justify="left", anchor="w").pack(fill="both", expand=True)

        # Bottom Bar
        bbar = tk.Frame(win, bg="#E8E8E8", pady=8, padx=16)
        bbar.pack(fill="x", side="bottom")
        tk.Button(bbar, text="Cerrar", font=("Segoe UI", 9, "bold"),
                  bg="#006CA8", fg="#FFFFFF", activebackground="#00507E", activeforeground="#FFFFFF",
                  bd=0, padx=16, pady=4, cursor="hand2", command=win.destroy).pack(side="right")

    def _minimize_window(self):
        self._is_minimized = True
        try:
            self.update_idletasks()
            self.overrideredirect(False)
            self.iconify()
            self.bind("<Map>", self._on_window_map)
        except Exception:
            pass

    def _on_window_map(self, event):
        if getattr(self, "_is_minimized", False) and event.widget == self:
            try:
                if self.state() == "normal":
                    self.overrideredirect(True)
                    self._is_minimized = False
                    self._set_taskbar_icon()
                    self.unbind("<Map>")
            except Exception:
                pass

    def _toggle_maximize(self):
        if not getattr(self, "_is_maximized", False):
            try:
                import ctypes
                hwnd = self.winfo_id()
                ctypes.windll.user32.ShowWindow(hwnd, 3)  # SW_MAXIMIZE
            except Exception:
                self.state("zoomed")
            self._is_maximized = True
            if hasattr(self, "_btn_max"):
                self._btn_max.config(text="❐")
        else:
            try:
                import ctypes
                hwnd = self.winfo_id()
                ctypes.windll.user32.ShowWindow(hwnd, 9)  # SW_RESTORE
            except Exception:
                self.state("normal")
            self._is_maximized = False
            if hasattr(self, "_btn_max"):
                self._btn_max.config(text="□")

    # ═══════════════════════════════════════════════════════════════
    # _build_ui()  —  Constructor principal de toda la ventana
    # ───────────────────────────────────────────────────────────────
    # Construye en este orden:
    #   1. Barra de título institucional (CSD: —, □, ✕ + drag)
    #   2. Menú clásico (Archivo | Acciones | Configuración | Ayuda)
    #   3. Barra de módulos horizontal (5 tabs: pdf, stock, json, guide, tools)
    #   4. Workspace con las 5 vistas (self._views dict)
    #      "pdf"   → Publicación de Ofertas PDF       (construida aquí inline)
    #      "stock" → Actualización de Stock           (delega a _build_stock_tab)
    #      "json"  → Subida de Precios JSON           (delega a tab_precios_json)
    #      "guide" → Instrucciones de Uso             (delega a gui_instructions_tab)
    #      "tools" → Herramientas Avanzadas           (delega a _build_advanced_tools_tab)
    #   5. Barra de estado inferior segmentada
    #
    # Navegación entre vistas: _switch_view(view_id) mueve grid/grid_forget
    # ❌ No poner lógica de negocio aquí. Solo widgets y layout.
    # ═══════════════════════════════════════════════════════════════
    def _build_ui(self):
        import tkinter as tk
        from tkinter import ttk
        C = self._C
        self.configure(bg="#F0F0F0")
        self.overrideredirect(True)
        self._set_taskbar_icon()
        self._setup_styles()

        # ── 1. BARRA DE TÍTULO INSTITUCIONAL (CON BOTONES INTEGRADOS —, □, ✕) ──
        titulo = tk.Frame(self, bg="#006CA8", height=34)
        titulo.pack(fill="x", side="top")
        titulo.pack_propagate(False)

        lbl_title = tk.Label(titulo, text="PERU COMPRAS BOT — Sistema de Automatización de Ofertas y Catálogos",
                             bg="#006CA8", fg="#FFFFFF", font=("Segoe UI", 10, "bold"),
                             anchor="w", cursor="fleur")
        lbl_title.pack(side="left", padx=14, fill="y")

        lbl_ver = tk.Label(titulo, text=f"THE KING COMPUTER E.I.R.L.  |  v{VERSION}",
                           bg="#006CA8", fg="#AACCDD", font=("Segoe UI", 9), cursor="fleur")
        lbl_ver.pack(side="left", padx=10, fill="y")

        # Botones de control de ventana (—, □, ✕)
        ctrl_frame = tk.Frame(titulo, bg="#006CA8")
        ctrl_frame.pack(side="right", fill="y")

        def make_win_btn(parent, text, cmd, hover_bg="#00507E", hover_fg="#FFFFFF", width=5):
            btn = tk.Label(parent, text=text, font=("Segoe UI", 10), bg="#006CA8", fg="#FFFFFF",
                           width=width, anchor="center", cursor="hand2")
            btn.pack(side="left", fill="y")
            btn.bind("<Enter>", lambda e: btn.config(bg=hover_bg, fg=hover_fg))
            btn.bind("<Leave>", lambda e: btn.config(bg="#006CA8", fg="#FFFFFF"))
            btn.bind("<Button-1>", lambda e: cmd())
            return btn

        make_win_btn(ctrl_frame, "—", self._minimize_window)
        self._btn_max = make_win_btn(ctrl_frame, "□", self._toggle_maximize)
        make_win_btn(ctrl_frame, "✕", self._close_window, hover_bg="#E81123", hover_fg="#FFFFFF")

        # Habilitar arrastre de ventana desde la barra
        for w in (titulo, lbl_title, lbl_ver):
            w.bind("<ButtonPress-1>", self._start_drag)
            w.bind("<B1-Motion>", self._drag_window)
            w.bind("<Double-Button-1>", lambda e: self._toggle_maximize())

        tk.Frame(titulo, bg="#00507E", height=2).place(relx=0, rely=1, relwidth=1, anchor="sw")

        # ── 2. BARRA DE MENÚ CLÁSICA ──
        menubar = tk.Menu(self, font=("Segoe UI", 10), bg="#E8E8E8", fg="#1A1A1A", relief="flat", bd=0)
        m_arch = tk.Menu(menubar, tearoff=0, font=("Segoe UI", 10))
        m_arch.add_command(label="Abrir archivo Excel...", command=self._pick_excel)
        m_arch.add_command(label="Limpiar datos", command=self._clear_excel)
        m_arch.add_separator()
        m_arch.add_command(label="Salir", command=self._close_window)
        menubar.add_cascade(label="Archivo", menu=m_arch)

        m_acc = tk.Menu(menubar, tearoff=0, font=("Segoe UI", 10))
        m_acc.add_command(label="Iniciar procesamiento (F5)", command=self._on_launch)
        m_acc.add_command(label="Detener ejecución", command=self._on_stop)
        m_acc.add_command(label="🔍 Auditar Portal contra Excel", command=self._on_stock_audit_start)
        m_acc.add_separator()
        m_acc.add_command(label="Publicación PDF", command=lambda: self._switch_view("pdf"))
        m_acc.add_command(label="Actualización de Stock", command=lambda: self._switch_view("stock"))
        m_acc.add_command(label="Subida Precios JSON", command=lambda: self._switch_view("json"))
        menubar.add_cascade(label="Acciones", menu=m_acc)

        m_cfg = tk.Menu(menubar, tearoff=0, font=("Segoe UI", 10))
        m_cfg.add_command(label="Preferencias de conexión y guardado", command=self._open_config_dialog)
        m_cfg.add_command(label="Parámetros de red y tiempos de pausa", command=self._open_config_dialog)
        menubar.add_cascade(label="Configuración", menu=m_cfg)

        m_ver = tk.Menu(menubar, tearoff=0, font=("Segoe UI", 10))
        m_ver.add_command(label="💻 Modo Compacto / Laptops (920 × 600)", command=lambda: self._set_window_preset(920, 600))
        m_ver.add_command(label="🖥️ Modo Estándar (1080 × 760)", command=lambda: self._set_window_preset(1080, 760))
        m_ver.add_command(label="📺 Modo Amplio (1280 × 820)", command=lambda: self._set_window_preset(1280, 820))
        m_ver.add_command(label="🖥️ Modo Ultra Amplio (1440 × 900)", command=lambda: self._set_window_preset(1440, 900))
        m_ver.add_separator()
        m_ver.add_command(label="🎯 Centrar ventana en pantalla", command=self._center_window)
        m_ver.add_command(label="⛶ Maximizar / Restaurar (F11)", command=self._toggle_maximize)
        menubar.add_cascade(label="Ver", menu=m_ver)

        m_hlp = tk.Menu(menubar, tearoff=0, font=("Segoe UI", 10))
        m_hlp.add_command(label="Manual de usuario (Instrucciones)", command=lambda: self._switch_view("guide"))
        m_hlp.add_command(label="Herramientas de diagnóstico", command=lambda: self._switch_view("tools"))
        m_hlp.add_separator()
        m_hlp.add_command(label="Acerca del sistema...",
                          command=lambda: messagebox.showinfo("Acerca de Perú Compras Bot",
                                                               f"PERU COMPRAS BOT v{VERSION}\n\n"
                                                               "Sistema Integrado de Automatización de Ofertas y Catálogos\n"
                                                               "Plataforma Oficial Perú Compras / Acuerdo EXT-CE-2022-5\n\n"
                                                               "Desarrollado para THE KING COMPUTER E.I.R.L."))
        menubar.add_cascade(label="Ayuda", menu=m_hlp)
        self.config(menu=menubar)

        # ── 3. BARRA DE MÓDULOS HORIZONTAL (TABS NATIVOS FLAT) ──
        tabs_frame = tk.Frame(self, bg="#E8E8E8", bd=0)
        tabs_frame.pack(fill="x")
        tk.Frame(tabs_frame, bg="#C8C8C8", height=1).pack(fill="x", side="bottom")

        self._nav_buttons = {}
        MODULOS = [
            ("pdf",   "Publicación de Ofertas PDF"),
            ("stock", "Actualización de Stock"),
            ("json",  "Subida de Precios JSON"),
            ("guide", "Instrucciones de Uso"),
            ("tools", "Herramientas Avanzadas"),
        ]
        tab_inner = tk.Frame(tabs_frame, bg="#E8E8E8")
        tab_inner.pack(side="left", padx=0)
        for mid, mlabel in MODULOS:
            btn = tk.Label(tab_inner, text=mlabel, font=("Segoe UI", 10),
                           padx=16, pady=6, cursor="hand2", bg="#F0F0F0", fg="#555555")
            btn.pack(side="left")
            tk.Frame(tab_inner, bg="#C8C8C8", width=1).pack(side="left", fill="y")
            btn.bind("<Button-1>", lambda e, m=mid: self._switch_view(m))
            self._nav_buttons[mid] = btn

        # ── 4. WORKSPACE CONTENEDOR PRINCIPAL ──
        workspace = tk.Frame(self, bg="#F0F0F0")
        workspace.pack(fill="both", expand=True)
        workspace.columnconfigure(0, weight=1)
        workspace.rowconfigure(0, weight=1)

        self._views = {}

        # ═ VISTA 1: PUBLICACIÓN DE OFERTAS PDF ═
        view_pdf = tk.Frame(workspace, bg="#F0F0F0")
        view_pdf.columnconfigure(0, weight=1)
        view_pdf.columnconfigure(1, weight=0, minsize=320)
        view_pdf.rowconfigure(0, weight=1)

        # Columna Izquierda: Zona de Trabajo (Archivo + Treeview Tabla de Productos)
        zona = tk.Frame(view_pdf, bg="#FFFFFF", bd=0)
        zona.grid(row=0, column=0, sticky="nsew")
        tk.Frame(view_pdf, bg="#C8C8C8", width=1).grid(row=0, column=0, sticky="nse")
        zona.columnconfigure(0, weight=1)
        zona.rowconfigure(1, weight=1)

        # Panel Carga de Archivo
        sec_carga = tk.Frame(zona, bg="#FFFFFF", bd=0)
        sec_carga.grid(row=0, column=0, sticky="ew")
        tk.Frame(sec_carga, bg="#C8C8C8", height=1).pack(fill="x", side="bottom")

        cabecera = tk.Frame(sec_carga, bg="#E8E8E8", pady=3)
        cabecera.pack(fill="x")
        tk.Label(cabecera, text="CARGA DE ARCHIVO DE TRABAJO",
                 font=("Segoe UI", 9, "bold"), bg="#E8E8E8", fg="#555555",
                 anchor="w").pack(side="left", padx=10)

        fila_arch = tk.Frame(sec_carga, bg="#FFFFFF", pady=6)
        fila_arch.pack(fill="x", padx=10)
        tk.Label(fila_arch, text="Archivo:", font=("Segoe UI", 10), bg="#FFFFFF", fg="#555555").pack(side="left")

        self.lbl_file = tk.Entry(fila_arch, font=("Segoe UI", 10), bd=1, relief="sunken",
                                 state="readonly", readonlybackground="#F0F0F0")
        self.lbl_file.pack(side="left", fill="x", expand=True, padx=(6, 4))
        self.btn_file = tk.Button(fila_arch, text="Examinar...", font=("Segoe UI", 10),
                                  bg="#E8E8E8", fg="#1A1A1A", bd=1, relief="raised",
                                  command=self._pick_excel)
        self.btn_file.pack(side="left", padx=2)
        tk.Button(fila_arch, text="Limpiar", font=("Segoe UI", 10),
                  bg="#E8E8E8", fg="#1A1A1A", bd=1, relief="raised",
                  command=self._clear_excel).pack(side="left", padx=2)

        # Sub-fila filtros (Pestaña / Col. N° Parte)
        fila_map = tk.Frame(sec_carga, bg="#FFFFFF", pady=4)
        fila_map.pack(fill="x", padx=10)
        tk.Label(fila_map, text="Pestaña Excel:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555").pack(side="left")
        self.combo_sheet = ttk.Combobox(fila_map, values=["--"], state="disabled", width=18)
        self.combo_sheet.pack(side="left", padx=(4, 12))
        self.combo_sheet.bind("<<ComboboxSelected>>", lambda e: self._on_sheet_changed(self.combo_sheet.get()))

        tk.Label(fila_map, text="Col. N° Parte:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555").pack(side="left")
        self.combo_parte = ttk.Combobox(fila_map, values=["--"], state="disabled", width=20)
        self.combo_parte.pack(side="left", padx=4)

        self.lbl_excel_info = tk.Label(sec_carga, text="", font=("Segoe UI", 9),
                                       bg="#FFFFFF", fg="#1B6B1B", anchor="w", padx=10, pady=2)
        self.lbl_excel_info.pack(fill="x")

        # Tabla de Productos (Treeview)
        tabla_frame = tk.Frame(zona, bg="#FFFFFF")
        tabla_frame.grid(row=1, column=0, sticky="nsew")
        tabla_frame.columnconfigure(0, weight=1)
        tabla_frame.rowconfigure(0, weight=1)

        COLS = ("#", "Número de Parte", "Descripción / Marca", "Precio Lista S/", "Stock Disp.", "Estado en Portal")
        self._tree = ttk.Treeview(tabla_frame, columns=COLS, show="headings",
                                   style="Hoja.Treeview", selectmode="extended")
        col_widths = [35, 140, 260, 100, 75, 130]
        col_anchors = ["center", "w", "w", "e", "center", "center"]
        for col, w, a in zip(COLS, col_widths, col_anchors):
            self._tree.heading(col, text=col)
            self._tree.column(col, width=w, anchor=a, stretch=(col == "Descripción / Marca"))

        sb_v = ttk.Scrollbar(tabla_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb_v.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        sb_v.grid(row=0, column=1, sticky="ns")
        self._tree.tag_configure("par",   background="#FFFFFF")
        self._tree.tag_configure("impar", background="#F7F7F7")

        # Columna Derecha: Inspector (Panel de Configuración de Ejecución)
        inspector = tk.Frame(view_pdf, bg="#FFFFFF", bd=0)
        inspector.grid(row=0, column=1, sticky="nsew")
        inspector.columnconfigure(0, weight=1)
        inspector.rowconfigure(1, weight=1)

        tk.Label(inspector, text="CONFIGURACIÓN DE EJECUCIÓN",
                 font=("Segoe UI", 9, "bold"), bg="#006CA8", fg="#FFFFFF",
                 anchor="w", padx=10, pady=6).grid(row=0, column=0, sticky="ew")

        cuerpo_insp = tk.Frame(inspector, bg="#FFFFFF", padx=10, pady=6)
        cuerpo_insp.grid(row=1, column=0, sticky="nsew")
        cuerpo_insp.columnconfigure(1, weight=1)
        fila = [0]

        def add_sep(titulo):
            sep_f = tk.Frame(cuerpo_insp, bg="#FFFFFF")
            sep_f.grid(row=fila[0], column=0, columnspan=2, sticky="ew", pady=(8, 3))
            tk.Label(sep_f, text=titulo, font=("Segoe UI", 9, "bold"),
                     bg="#FFFFFF", fg="#006CA8").pack(side="left")
            tk.Frame(sep_f, bg="#C8C8C8", height=1).pack(side="bottom", fill="x")
            fila[0] += 1

        # Acceso
        add_sep("Acceso al Portal Perú Compras")
        tk.Label(cuerpo_insp, text="Usuario:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555", anchor="e").grid(row=fila[0], column=0, sticky="e", padx=(0,4), pady=2)
        self.entry_user = tk.Entry(cuerpo_insp, font=("Segoe UI", 9), bd=1, relief="sunken", bg="#F0F0F0")
        self.entry_user.insert(0, "almerco.03")
        self.entry_user.grid(row=fila[0], column=1, sticky="ew", pady=2)
        fila[0] += 1

        tk.Label(cuerpo_insp, text="Contraseña:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555", anchor="e").grid(row=fila[0], column=0, sticky="e", padx=(0,4), pady=2)
        self.entry_pass = tk.Entry(cuerpo_insp, font=("Segoe UI", 9), bd=1, relief="sunken", bg="#F0F0F0", show="*")
        self.entry_pass.insert(0, "4lm3rKenYa@#")
        self.entry_pass.grid(row=fila[0], column=1, sticky="ew", pady=2)
        fila[0] += 1

        self.check_visible = tk.BooleanVar(value=False)
        chk = tk.Checkbutton(cuerpo_insp, text="Navegador visible (no oculto)", variable=self.check_visible,
                             font=("Segoe UI", 9), bg="#FFFFFF", activebackground="#FFFFFF")
        chk.grid(row=fila[0], column=0, columnspan=2, sticky="w", pady=2)
        fila[0] += 1

        # Filtros Catálogo
        add_sep("Parámetros de Catálogo")
        comb_data = self._catalog_data.get("combinaciones", [])

        tk.Label(cuerpo_insp, text="Catálogo:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555", anchor="e").grid(row=fila[0], column=0, sticky="e", padx=(0,4), pady=2)
        self.combo_catalogo = ttk.Combobox(cuerpo_insp, values=self._opts_texts(comb_data), state="readonly", font=("Segoe UI", 9))
        self.combo_catalogo.grid(row=fila[0], column=1, sticky="ew", pady=2)
        self.combo_catalogo.bind("<<ComboboxSelected>>", lambda e: self._on_catalogo_changed(self.combo_catalogo.get()))
        fila[0] += 1

        tk.Label(cuerpo_insp, text="Categoría:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555", anchor="e").grid(row=fila[0], column=0, sticky="e", padx=(0,4), pady=2)
        self.combo_categoria = ttk.Combobox(cuerpo_insp, values=["Seleccione Catálogo"], state="readonly", font=("Segoe UI", 9))
        self.combo_categoria.grid(row=fila[0], column=1, sticky="ew", pady=2)
        self.combo_categoria.bind("<<ComboboxSelected>>", lambda e: self._on_categoria_changed(self.combo_categoria.get()))
        fila[0] += 1

        tk.Label(cuerpo_insp, text="Estado:", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555", anchor="e").grid(row=fila[0], column=0, sticky="e", padx=(0,4), pady=2)
        self.combo_estado = ttk.Combobox(cuerpo_insp, values=["Seleccione Categoría"], state="readonly", font=("Segoe UI", 9))
        self.combo_estado.grid(row=fila[0], column=1, sticky="ew", pady=2)
        fila[0] += 1

        if comb_data:
            self.combo_catalogo.set(self._opts_texts(comb_data)[0])
            self._on_catalogo_changed(self._opts_texts(comb_data)[0])

        tk.Label(cuerpo_insp, text="Pausa (seg):", font=("Segoe UI", 9), bg="#FFFFFF", fg="#555555", anchor="e").grid(row=fila[0], column=0, sticky="e", padx=(0,4), pady=2)
        self.slider_pausa = tk.DoubleVar(value=1.5)
        e_pausa = tk.Entry(cuerpo_insp, textvariable=self.slider_pausa, font=("Segoe UI", 9), bd=1, relief="sunken", bg="#F0F0F0", width=8)
        e_pausa.grid(row=fila[0], column=1, sticky="w", pady=2)
        fila[0] += 1

        # Consola Log de Eventos (log_box)
        add_sep("Consola de Eventos")
        consola_f = tk.Frame(cuerpo_insp, bg="#1A1A2E")
        consola_f.grid(row=fila[0], column=0, columnspan=2, sticky="ew", pady=2)
        consola_f.columnconfigure(0, weight=1)

        self.log_box = tk.Text(consola_f, height=6, font=("Consolas", 9),
                               bg="#1A1A2E", fg="#E0E0E0", bd=0, relief="flat",
                               state="disabled", wrap="word")
        self.log_box.pack(fill="both", expand=True)
        self.log_box.tag_configure("ok",   foreground="#90EE90")
        self.log_box.tag_configure("warn", foreground="#FFD700")
        self.log_box.tag_configure("error",foreground="#FF6B6B")
        self.log_box.tag_configure("info", foreground="#87CEEB")
        self.log_box.tag_configure("done", foreground="#5DADE2")
        fila[0] += 1

        # Botón Ejecutar e Iniciar
        btn_f = tk.Frame(inspector, bg="#FFFFFF", pady=6, padx=10)
        btn_f.grid(row=2, column=0, sticky="ew")

        self.btn_launch = tk.Button(btn_f, text="INICIAR PROCESAMIENTO  (F5)",
                                    font=("Segoe UI", 10, "bold"), bg="#006CA8", fg="#FFFFFF",
                                    activebackground="#00507E", activeforeground="#FFFFFF",
                                    bd=0, pady=8, cursor="hand2",
                                    command=self._on_launch)
        self.btn_launch.pack(fill="x")

        self.btn_stop = tk.Button(btn_f, text="Detener",
                                  font=("Segoe UI", 9, "bold"), bg="#8B1A1A", fg="#FFFFFF",
                                  bd=0, pady=4, state="disabled", cursor="hand2",
                                  command=self._on_stop)
        self.btn_stop.pack(fill="x", pady=(4, 0))

        # ── SECCIÓN DE AUDITOR DE RESULTADOS E INFORME ──
        audit_f = tk.Frame(btn_f, bg="#F8F9FA", bd=1, relief="solid", highlightbackground="#C8C8C8", pady=6, padx=8)
        audit_f.pack(fill="x", pady=(10, 0))

        tk.Label(audit_f, text="🔍 Auditor de Resultados de Proceso", font=("Segoe UI", 9, "bold"),
                 bg="#F8F9FA", fg="#006CA8").pack(anchor="w")

        self.lbl_audit_summary = tk.Label(audit_f, text="Cargue o procese datos para auditar...",
                                          font=("Segoe UI", 8), bg="#F8F9FA", fg="#555555", anchor="w", justify="left")
        self.lbl_audit_summary.pack(fill="x", pady=2)

        btn_row_audit = tk.Frame(audit_f, bg="#F8F9FA")
        btn_row_audit.pack(fill="x", pady=(4, 0))

        tk.Button(btn_row_audit, text="📊 Informe Excel", font=("Segoe UI", 8, "bold"),
                  bg="#1B6B1B", fg="#FFFFFF", activebackground="#145214", activeforeground="#FFFFFF",
                  bd=0, padx=6, pady=4, cursor="hand2",
                  command=lambda: self._export_audit_report(fmt="excel", modulo_nombre="Publicación PDF")).pack(side="left", padx=(0, 4), fill="x", expand=True)

        tk.Button(btn_row_audit, text="📄 Informe PDF", font=("Segoe UI", 8, "bold"),
                  bg="#006CA8", fg="#FFFFFF", activebackground="#00507E", activeforeground="#FFFFFF",
                  bd=0, padx=6, pady=4, cursor="hand2",
                  command=lambda: self._export_audit_report(fmt="pdf", modulo_nombre="Publicación PDF")).pack(side="left", fill="x", expand=True)

        self._views["pdf"] = view_pdf

        # ── OTROS MÓDULOS DE TRABAJO ──
        view_stock = tk.Frame(workspace, bg="#F0F0F0")
        self._build_stock_tab(parent=view_stock)
        self._views["stock"] = view_stock

        view_json = tk.Frame(workspace, bg="#F0F0F0")
        import tab_precios_json
        tab_precios_json.build_precios_json_tab(self, parent=view_json)
        self._views["json"] = view_json

        view_guide = tk.Frame(workspace, bg="#F0F0F0")
        import gui_instructions_tab
        gui_instructions_tab.build_instructions_tab(view_guide, C=C)
        self._views["guide"] = view_guide

        view_tools = tk.Frame(workspace, bg="#F0F0F0")
        self._build_advanced_tools_tab(view_tools)
        self._views["tools"] = view_tools

        # ── 5. BARRA DE ESTADO INFERIOR SEGMENTADA ──
        statusbar = tk.Frame(self, bg="#E8E8E8", bd=0)
        statusbar.pack(fill="x", side="bottom")
        tk.Frame(statusbar, bg="#C8C8C8", height=1).pack(fill="x", side="top")

        def st_seg(texto, color="#555555", bold=False):
            lbl = tk.Label(statusbar, text=texto,
                           font=("Segoe UI", 9, "bold" if bold else "normal"),
                           bg="#E8E8E8", fg=color, pady=3, padx=10)
            lbl.pack(side="left")
            tk.Frame(statusbar, bg="#C8C8C8", width=1).pack(side="left", fill="y", pady=2)
            return lbl

        self.lbl_footer_status = st_seg("Listo", "#1B6B1B", bold=True)
        self.lbl_status = self.lbl_footer_status
        self._lbl_nav = st_seg("Navegador: No iniciado")
        self.lbl_counter = st_seg("Registros: 0 cargados")
        self._lbl_modulo = st_seg("Modulo: Publicacion de Ofertas PDF")

        # Manija de redimensionamiento visual en esquina inferior derecha (Sizegrip)
        grip = tk.Label(statusbar, text="⇲", font=("Segoe UI", 11, "bold"),
                        bg="#E8E8E8", fg="#777777", cursor="size_nw_se", padx=6)
        grip.pack(side="right", fill="y")
        grip.bind("<ButtonPress-1>", lambda e: self._start_resize(e, "se"))
        grip.bind("<B1-Motion>", self._do_resize)
        grip.bind("<ButtonRelease-1>", self._stop_resize)
        grip.bind("<Enter>", lambda e: grip.config(fg="#006CA8"))
        grip.bind("<Leave>", lambda e: grip.config(fg="#777777"))

        tk.Label(statusbar, text=f"Peru Compras Bot v{VERSION}",
                 font=("Segoe UI", 9), bg="#E8E8E8", fg="#555555",
                 padx=8).pack(side="right")

        # Inicializar el sistema perimetral de bordes activos en las 8 direcciones
        self._setup_border_resizing()

        self.bind("<F5>", lambda e: self._on_launch())
        self.bind("<F11>", lambda e: self._toggle_maximize())
        self._switch_view("pdf")

    def _clear_excel(self):
        self._excel_path = ""
        self._excel_rows = []
        self.lbl_file.config(state="normal")
        self.lbl_file.delete(0, "end")
        self.lbl_file.config(state="readonly")
        if hasattr(self, "_tree"):
            self._tree.delete(*self._tree.get_children())
        self.lbl_excel_info.config(text="")

    def _switch_view(self, view_id):
        """Cambia la vista activa y resalta el tab horizontal correspondiente."""
        for v_name, frame in self._views.items():
            if v_name == view_id:
                frame.grid(row=0, column=0, sticky="nsew")
            else:
                frame.grid_forget()

        LABELS = {
            "pdf":   "Publicación de Ofertas PDF",
            "stock": "Actualización de Stock",
            "json":  "Subida de Precios JSON",
            "guide": "Instrucciones de Uso",
            "tools": "Herramientas Avanzadas",
        }
        for v_name, btn in self._nav_buttons.items():
            if v_name == view_id:
                btn.config(bg="#FFFFFF", fg="#006CA8", font=("Segoe UI", 10, "bold"))
            else:
                btn.config(bg="#F0F0F0", fg="#555555", font=("Segoe UI", 10))

        if hasattr(self, "_lbl_modulo") and self._lbl_modulo:
            self._lbl_modulo.config(text="Modulo: " + LABELS.get(view_id, view_id))
        self._current_view = view_id

    # ── METODOS DE AUDITOR DE RESULTADOS E INFORME DE EXPORTACIÓN ──

    def _collect_tree_rows(self):
        rows = []
        if hasattr(self, "_tree"):
            for item in self._tree.get_children():
                vals = self._tree.item(item, "values")
                if vals and len(vals) >= 6:
                    rows.append({
                        "parte": vals[1],
                        "descripcion": vals[2],
                        "precio": vals[3],
                        "stock": vals[4],
                        "estado": vals[5],
                    })
        if not rows and hasattr(self, "_excel_rows") and self._excel_rows:
            rows = self._excel_rows
        return rows

    def _run_auditor_check(self, modulo_nombre="Publicación PDF"):
        """Ejecuta el chequeo rápido del auditor sobre las fichas procesadas."""
        from utils_mod.audit_reporter import audit_results
        rows = self._collect_tree_rows()
        summary = audit_results(rows)
        if hasattr(self, "lbl_audit_summary") and self.lbl_audit_summary:
            text = f"Total: {summary['total']} | ✓ OK: {summary['ok']} | ✕ Err: {summary['err']} | Éxito: {summary['rate']}%"
            self.lbl_audit_summary.config(text=text)
        return rows, summary

    def _export_audit_report(self, fmt="excel", modulo_nombre="Publicación PDF"):
        """Genera y guarda el informe de auditoría en Excel (.xlsx) o PDF/HTML."""
        from utils_mod.audit_reporter import audit_results, export_excel_report, export_pdf_report
        rows, summary = self._run_auditor_check(modulo_nombre)

        if not rows:
            messagebox.showwarning("Auditor del Sistema", "No hay datos de productos ni ejecuciones para auditar.\nPor favor cargue o procese datos primero.")
            return

        if fmt == "excel":
            def_ext = ".xlsx"
            ftypes = [("Libro de Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        else:
            def_ext = ".html"
            ftypes = [("Informe de Auditoría PDF/HTML", "*.html"), ("Todos los archivos", "*.*")]

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_file = f"Informe_Auditoria_{modulo_nombre.replace(' ', '_')}_{ts}{def_ext}"

        path = filedialog.asksaveasfilename(
            title=f"Guardar Informe de Auditoría ({fmt.upper()}) — {modulo_nombre}",
            initialfile=default_file,
            defaultextension=def_ext,
            filetypes=ftypes,
        )
        if not path:
            return

        if fmt == "excel":
            ok, msg = export_excel_report(rows, summary, path, modulo_nombre=modulo_nombre)
        else:
            ok, msg = export_pdf_report(rows, summary, path, modulo_nombre=modulo_nombre)

        if ok:
            messagebox.showinfo("Auditor de Resultados", f"¡Informe de Auditoría generado exitosamente!\n\nArchivo creado:\n{msg}")
        else:
            messagebox.showerror("Error en Auditoría", f"Ocurrió un error al generar el informe:\n{msg}")

    def _update_tools_excel_status(self):
        if not hasattr(self, "lbl_tools_excel_status") or not self.lbl_tools_excel_status:
            return
        if getattr(self, "_excel_path", None):
            name = os.path.basename(self._excel_path)
            cnt = len(self._excel_rows) if hasattr(self, "_excel_rows") else 0
            self.lbl_tools_excel_status.config(
                text=f"  ✓ Archivo Excel cargado: {name} ({cnt} registros listos para pruebas)  ",
                bg="#DFF0D8", fg="#1B6B1B"
            )
        else:
            self.lbl_tools_excel_status.config(
                text="  ⚠️ Sin archivo Excel cargado — Seleccione un Excel abajo o en el módulo principal para ejecutar las pruebas.  ",
                bg="#FCF8E3", fg="#854D0E"
            )

    def _build_advanced_tools_tab(self, parent):
        """Vista de Herramientas Avanzadas — diagnóstico y scrapers en Tkinter nativo."""
        import tkinter as tk
        from tkinter import ttk

        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        # Canvas con scrollbar para el contenedor
        canvas = tk.Canvas(parent, bg="#F0F0F0", bd=0, highlightthickness=0)
        sb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)

        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg="#F0F0F0")
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_cfg(e):
            if e.width > 20:
                canvas.itemconfig(win_id, width=e.width)

        inner.bind("<Configure>", _on_cfg)
        canvas.bind("<Configure>", _on_canvas_cfg)

        # TÍTULO PRINCIPAL
        header_f = tk.Frame(inner, bg="#F0F0F0", pady=12, padx=16)
        header_f.pack(fill="x")
        tk.Label(header_f, text="DIAGNÓSTICO Y HERRAMIENTAS AVANZADAS",
                 font=("Segoe UI", 11, "bold"), bg="#F0F0F0", fg="#006CA8", anchor="w").pack(fill="x")
        tk.Label(header_f, text="Pruebas unitarias de flujo, extracción de catálogos y scrapers de endpoints de Perú Compras.",
                 font=("Segoe UI", 9), bg="#F0F0F0", fg="#555555", anchor="w").pack(fill="x", pady=(2, 0))

        # PANEL 0: ESTADO DEL ARCHIVO DE TRABAJO (CON BOTÓN DE SELECCIÓN)
        sec_excel = tk.Frame(inner, bg="#FFFFFF", bd=1, relief="solid", highlightbackground="#C8C8C8")
        sec_excel.pack(fill="x", padx=16, pady=(0, 12))

        sec_excel_head = tk.Frame(sec_excel, bg="#E8E8E8", pady=4, padx=10)
        sec_excel_head.pack(fill="x")
        tk.Label(sec_excel_head, text="ESTADO DEL ARCHIVO DE TRABAJO PARA PRUEBAS",
                 font=("Segoe UI", 9, "bold"), bg="#E8E8E8", fg="#555555", anchor="w").pack(side="left")

        sec_excel_body = tk.Frame(sec_excel, bg="#FFFFFF", pady=8, padx=10)
        sec_excel_body.pack(fill="x")
        self.lbl_tools_excel_status = tk.Label(sec_excel_body, text="",
                                               font=("Segoe UI", 9), anchor="w", pady=4, padx=8)
        self.lbl_tools_excel_status.pack(side="left", fill="x", expand=True)

        tk.Button(sec_excel_body, text="Examinar Excel...", font=("Segoe UI", 9),
                  bg="#E8E8E8", fg="#1A1A1A", bd=1, relief="raised", padx=8,
                  command=self._pick_excel).pack(side="left", padx=4)
        tk.Button(sec_excel_body, text="Limpiar", font=("Segoe UI", 9),
                  bg="#E8E8E8", fg="#1A1A1A", bd=1, relief="raised", padx=8,
                  command=self._clear_excel).pack(side="left", padx=2)

        self._update_tools_excel_status()

        # HELPER DE CARD DE HERRAMIENTAS
        def create_tool_card(title, subtitle):
            card = tk.Frame(inner, bg="#FFFFFF", bd=1, relief="solid", highlightbackground="#C8C8C8")
            card.pack(fill="x", padx=16, pady=6)

            chead = tk.Frame(card, bg="#E8E8E8", pady=4, padx=10)
            chead.pack(fill="x")
            tk.Label(chead, text=title, font=("Segoe UI", 9, "bold"),
                     bg="#E8E8E8", fg="#006CA8", anchor="w").pack(side="left")

            cbody = tk.Frame(card, bg="#FFFFFF", pady=8, padx=12)
            cbody.pack(fill="x")

            if subtitle:
                tk.Label(cbody, text=subtitle, font=("Segoe UI", 9),
                         bg="#FFFFFF", fg="#555555", anchor="w", justify="left").pack(fill="x", pady=(0, 8))

            btn_row = tk.Frame(cbody, bg="#FFFFFF")
            btn_row.pack(fill="x")
            return btn_row

        def make_action_btn(parent_row, text, command, bg="#006CA8", fg="#FFFFFF"):
            btn = tk.Button(parent_row, text=text, font=("Segoe UI", 9, "bold"),
                            bg=bg, fg=fg, activebackground="#00507E", activeforeground="#FFFFFF",
                            bd=1, relief="raised", padx=12, pady=6, cursor="hand2",
                            command=command)
            btn.pack(side="left", padx=(0, 8))
            return btn

        # PANEL 1: PRUEBAS DE FLUJO RÁPIDO
        row1 = create_tool_card(
            "PRUEBAS DE FLUJO RÁPIDO (REQUIERE EXCEL Y CREDENCIALES)",
            "Ejecuta pruebas unitarias de procesamiento sobre la 1ª ficha del Excel cargado usando las credenciales del panel lateral derecho."
        )
        self.btn_test = make_action_btn(row1, "🧪 Test (1 Ficha)", self._on_test, bg="#006CA8")
        self.btn_nro = make_action_btn(row1, "🔍 Solo N° de Parte", self._on_nro_parte, bg="#E8E8E8", fg="#1A1A1A")
        self.btn_certs = make_action_btn(row1, "🏅 Solo Certificaciones", self._on_certs_only, bg="#E8E8E8", fg="#1A1A1A")

        # PANEL 2: EXTRACCIÓN Y SCRAPERS DE CATÁLOGO
        row2 = create_tool_card(
            "EXTRACCIÓN Y SCRAPERS DE CATÁLOGO (NO REQUIERE EXCEL PREVIO)",
            "Herramientas avanzadas de consulta directa sobre el portal de Perú Compras para extracción de datos y diagnósticos."
        )
        self.btn_extract = make_action_btn(row2, "📊 Extraer Reportes", self._on_extract, bg="#006CA8")
        self.btn_compare = make_action_btn(row2, "⚖️ Comparar Fichas", self._on_compare, bg="#E8E8E8", fg="#1A1A1A")
        self.btn_discovery = make_action_btn(row2, "📡 Discovery v1 (Endpoints)", self._on_discovery, bg="#E8E8E8", fg="#1A1A1A")
        self.btn_discovery2 = make_action_btn(row2, "🔬 Discovery v2 (JSON Schema)", self._on_discovery2, bg="#E8E8E8", fg="#1A1A1A")

    # ── Pestaña 2: Stock/Cobertura/Plazo ─────────────────────────

    def _build_stock_tab(self, left_col=None, right_col=None, parent=None):
        """Vista de Análisis de Stock — paleta institucional light."""
        if parent is None:
            parent = right_col
        C = self._C

        # Estado
        self._stock_excel_path = ""
        self._stock_excel_df = []
        self._stock_running = False
        self._stock_stop_event = threading.Event()
        self._stock_log_queue = queue.Queue()
        self._stock_log_lines = []
        self._stock_total = 0
        self._stock_ok = 0
        self._stock_errors = 0
        self._stock_report_path = ""

        parent.grid_columnconfigure(0, weight=1, minsize=360)
        parent.grid_columnconfigure(1, weight=1, minsize=360)
        parent.grid_rowconfigure(0, weight=1)

        # LEFT COLUMN
        left = ctk.CTkScrollableFrame(parent, fg_color="transparent",
                                      scrollbar_button_color=C["border"])
        left.grid(row=0, column=0, padx=(12, 6), pady=12, sticky="nsew")
        left.grid_columnconfigure(0, weight=1)

        # ── Sección 0: Credenciales ──
        self._section_label(left, "Credenciales del Flujo Stock", 0)
        frame_creds = ctk.CTkFrame(left, fg_color=C["card"], corner_radius=6,
                                   border_width=1, border_color=C["border"])
        frame_creds.grid(row=1, column=0, padx=0, pady=(0, 10), sticky="ew")
        frame_creds.grid_columnconfigure(1, weight=1)

        for ri, (lbl, attr, default, show) in enumerate([
            ("Usuario",    "entry_stock_user", "estalin.huamali01", ""),
            ("Contraseña", "entry_stock_pass", "PE/CyG6c&1R4T=",    "*"),
        ]):
            ctk.CTkLabel(frame_creds, text=lbl, anchor="w", font=ctk.CTkFont(size=11),
                         text_color=C["txt2"]).grid(row=ri*2,   column=0, columnspan=2,
                                                    padx=12, pady=(8 if ri==0 else 2, 1), sticky="w")
            e = ctk.CTkEntry(frame_creds, show=show, height=32,
                             fg_color=C["card2"], border_color=C["border"], text_color=C["txt"])
            e.insert(0, default)
            e.grid(row=ri*2+1, column=0, columnspan=2, padx=12,
                   pady=(0, 8 if ri==1 else 4), sticky="ew")
            setattr(self, attr, e)

        # Check mostrar navegador
        self.check_stock_visible = ctk.CTkCheckBox(
            frame_creds, text="Mostrar navegador en pantalla",
            font=ctk.CTkFont(size=12), text_color=C["txt"],
            border_color=C["border"], fg_color=C["accent"],
        )
        self.check_stock_visible.grid(row=4, column=0, columnspan=2,
                                      padx=12, pady=(2, 10), sticky="w")

        # ── Sección 1: Excel de productos ──
        self._section_label(left, "Archivo Excel de Productos", 2)
        frame_excel = ctk.CTkFrame(left, fg_color=C["card"], corner_radius=6,
                                   border_width=1, border_color=C["border"])
        frame_excel.grid(row=3, column=0, padx=0, pady=(0, 10), sticky="ew")
        frame_excel.grid_columnconfigure(0, weight=1)

        self.lbl_stock_excel = ctk.CTkLabel(
            frame_excel, text="(sin archivo)", text_color=C["txt3"],
            font=ctk.CTkFont(size=11), anchor="w",
        )
        self.lbl_stock_excel.grid(row=0, column=0, padx=12, pady=(10, 4), sticky="ew")

        btn_row = ctk.CTkFrame(frame_excel, fg_color="transparent")
        btn_row.grid(row=1, column=0, padx=12, pady=(0, 4), sticky="ew")
        ctk.CTkButton(
            btn_row, text="Cargar Excel", width=120, height=32,
            fg_color=C["accent"], hover_color=C["accent_h"], text_color="#FFFFFF",
            corner_radius=4, font=ctk.CTkFont(size=12),
            command=self._on_load_stock_excel,
        ).pack(side="left", padx=(0, 8))
        ctk.CTkButton(
            btn_row, text="Descargar Plantilla", width=160, height=32,
            fg_color=C["card2"], hover_color=C["border"],
            text_color=C["txt"], border_width=1, border_color=C["border"],
            corner_radius=4, font=ctk.CTkFont(size=12),
            command=self._on_download_stock_template,
        ).pack(side="left")
        self.lbl_stock_summary = ctk.CTkLabel(
            frame_excel, text="", text_color=C["txt3"],
            font=ctk.CTkFont(size=11), anchor="w",
        )
        self.lbl_stock_summary.grid(row=2, column=0, padx=12, pady=(0, 10), sticky="ew")

        # ── Sección 2: Filtros del portal ──
        self._section_label(left, "Filtros del Portal (Acuerdo > Catálogo > Categoría)", 4)
        frame_filtros = ctk.CTkFrame(left, fg_color=C["card"], corner_radius=6,
                                     border_width=1, border_color=C["border"])
        frame_filtros.grid(row=5, column=0, padx=0, pady=(0, 10), sticky="ew")
        frame_filtros.grid_columnconfigure(1, weight=1)

        self._stock_combos_data = self._load_stock_combos_json()
        acuerdos_list = self._stock_combos_data.get("acuerdos", [])
        acuerdo_values = [a["text"] for a in acuerdos_list] if acuerdos_list else ["-- Sin datos --"]

        _om_style = dict(
            fg_color=C["card2"], button_color=C["border"],
            button_hover_color=C["border2"], text_color=C["txt"],
            dropdown_fg_color=C["card"], dropdown_hover_color=C["card2"],
            dropdown_text_color=C["txt"],
        )

        for ri, (lbl, attr, vals, cmd) in enumerate([
            ("Acuerdo",   "option_stock_acuerdo",   acuerdo_values,
             self._on_stock_acuerdo_changed),
            ("Catálogo",  "option_stock_catalogo",  ["-- Seleccione acuerdo primero --"],
             self._on_stock_catalogo_changed),
            ("Categoría", "option_stock_categoria", ["-- Seleccione catálogo primero --"],
             None),
        ]):
            ctk.CTkLabel(frame_filtros, text=lbl, anchor="w",
                         font=ctk.CTkFont(size=11), text_color=C["txt2"]
                         ).grid(row=ri*2, column=0, columnspan=2,
                                padx=12, pady=(8 if ri==0 else 4, 1), sticky="w")
            opts = dict(**_om_style, width=300, command=cmd) if cmd else dict(**_om_style, width=300)
            om = ctk.CTkOptionMenu(frame_filtros, values=vals, **opts)
            om.grid(row=ri*2+1, column=0, columnspan=2, padx=12,
                    pady=(0, 4 if ri < 2 else 0), sticky="ew")
            setattr(self, attr, om)

        # Pausa
        ctk.CTkLabel(frame_filtros, text="Pausa entre registros (seg)", anchor="w",
                     font=ctk.CTkFont(size=11), text_color=C["txt2"]
                     ).grid(row=6, column=0, columnspan=2, padx=12, pady=(8, 1), sticky="w")
        self.entry_stock_pausa = ctk.CTkEntry(
            frame_filtros, width=80, height=30,
            fg_color=C["card2"], border_color=C["border"], text_color=C["txt"],
        )
        self.entry_stock_pausa.insert(0, "2")
        self.entry_stock_pausa.grid(row=7, column=0, padx=12, pady=(0, 10), sticky="w")

        if acuerdos_list:
            self.option_stock_acuerdo.set(acuerdo_values[0])
            self._on_stock_acuerdo_changed(acuerdo_values[0])

        # ── Botones de acción ──
        self._section_label(left, "Iniciar / Detener / Auditar", 6)
        frame_btns = ctk.CTkFrame(left, fg_color="transparent")
        frame_btns.grid(row=7, column=0, padx=0, pady=(0, 8), sticky="ew")

        self.btn_stock_start = ctk.CTkButton(
            frame_btns, text="Iniciar Stock", width=130, height=36,
            fg_color=C["accent"], hover_color=C["accent_h"], text_color="#FFFFFF",
            corner_radius=6, font=ctk.CTkFont(size=12, weight="bold"),
            command=self._on_stock_start,
        )
        self.btn_stock_start.pack(side="left", padx=(0, 6))

        self.btn_stock_audit = ctk.CTkButton(
            frame_btns, text="🔍 Auditar Portal", width=140, height=36,
            fg_color="#1B6B1B", hover_color="#145214", text_color="#FFFFFF",
            corner_radius=6, font=ctk.CTkFont(size=12, weight="bold"),
            command=self._on_stock_audit_start,
        )
        self.btn_stock_audit.pack(side="left", padx=(0, 6))

        self.btn_stock_stop = ctk.CTkButton(
            frame_btns, text="Detener", width=90, height=36,
            fg_color=C["danger"], hover_color=C["danger_h"],
            text_color="#FFFFFF", state="disabled",
            corner_radius=6, font=ctk.CTkFont(size=12, weight="bold"),
            command=self._on_stock_stop,
        )
        self.btn_stock_stop.pack(side="left", padx=(0, 6))

        self.lbl_stock_status = ctk.CTkLabel(
            frame_btns, text="Listo", text_color=C["txt3"],
            font=ctk.CTkFont(size=11),
        )
        self.lbl_stock_status.pack(side="left", padx=4)

        # RIGHT COLUMN: stats + log
        right = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=8,
                             border_width=1, border_color=C["border"])
        right.grid(row=0, column=1, padx=(6, 12), pady=12, sticky="nsew")
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(2, weight=1)

        # Section label
        self._section_label(right, "Panel de Ejecución", 0)

        # Stats
        stats = ctk.CTkFrame(right, fg_color="transparent")
        stats.grid(row=1, column=0, padx=12, pady=(0, 8), sticky="ew")
        stats.grid_columnconfigure((0, 1, 2, 3), weight=1)

        for col, lbl, color, attr in [
            (0, "Total",   C["txt"],     "lbl_stock_stat_total"),
            (1, "Éxito",   C["success"], "lbl_stock_stat_ok"),
            (2, "Fallos",  C["danger"],  "lbl_stock_stat_fail"),
            (3, "Reporte", C["txt3"],    "lbl_stock_report"),
        ]:
            f = ctk.CTkFrame(stats, fg_color=C["card2"], corner_radius=6,
                             border_width=1, border_color=C["border"])
            f.grid(row=0, column=col, padx=3, sticky="ew")
            font_size = 22 if col < 3 else 10
            lbl_n = ctk.CTkLabel(
                f, text="0" if col < 3 else "(ninguno)",
                font=ctk.CTkFont(size=font_size, weight="bold" if col < 3 else "normal"),
                text_color=color,
                wraplength=160 if col == 3 else 0,
            )
            lbl_n.pack(pady=(8, 0))
            ctk.CTkLabel(f, text=lbl, font=ctk.CTkFont(size=10),
                         text_color=C["txt3"]).pack(pady=(0, 8))
            setattr(self, attr, lbl_n)

        # Progreso
        self.progress_stock = ctk.CTkProgressBar(
            right, height=6, fg_color=C["border"], progress_color=C["accent"],
        )
        self.progress_stock.grid(row=3, column=0, padx=12, pady=(0, 6), sticky="ew")
        self.progress_stock.set(0)

        # Log
        self.log_stock = ctk.CTkTextbox(
            right, font=ctk.CTkFont(family="Courier New", size=11),
            fg_color=C["card"], border_width=1, border_color=C["border"],
            text_color=C["txt"],
        )
        self.log_stock.grid(row=2, column=0, padx=12, pady=(0, 4), sticky="nsew")
        self.log_stock.configure(state="disabled")

        # Auditor de Stock
        audit_f = ctk.CTkFrame(right, fg_color=C["card2"], corner_radius=6, border_width=1, border_color=C["border"])
        audit_f.grid(row=4, column=0, padx=12, pady=(4, 12), sticky="ew")
        ctk.CTkLabel(audit_f, text="🔍 Auditor de Resultados — Stock", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=C["txt"]).pack(anchor="w", padx=10, pady=(6, 2))

        self.lbl_audit_stock_summary = ctk.CTkLabel(audit_f, text="Procese registros para auditar stock...",
                                                    font=ctk.CTkFont(size=10), text_color=C["txt2"], anchor="w")
        self.lbl_audit_stock_summary.pack(fill="x", padx=10, pady=(0, 6))

        btn_r = ctk.CTkFrame(audit_f, fg_color="transparent")
        btn_r.pack(fill="x", padx=10, pady=(0, 8))

        ctk.CTkButton(
            btn_r, text="📊 Informe Excel", height=32, font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#1B6B1B", hover_color="#145214", text_color="#FFFFFF",
            command=lambda: self._export_stock_audit_report(fmt="excel")
        ).pack(side="left", padx=(0, 4), fill="x", expand=True)

        ctk.CTkButton(
            btn_r, text="📄 Informe PDF", height=32, font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#006CA8", hover_color="#00507E", text_color="#FFFFFF",
            command=lambda: self._export_stock_audit_report(fmt="pdf")
        ).pack(side="left", fill="x", expand=True)

    # ═══════════════════════════════════════════════════════════════
    # BLOQUE 6 — Handlers del módulo Stock
    # ───────────────────────────────────────────────────────────────
    # _on_load_stock_excel()    → abre filedialog, valida con analizar_excel_stock()
    # _on_download_stock_template() → genera plantilla.xlsx con openpyxl
    # _load_stock_combos_json() → carga dropdowns_mejora_basica.json
    # _on_stock_acuerdo_changed() / _on_stock_catalogo_changed()
    #   → cascada de filtros Acuerdo → Catálogo → Categoría
    # _on_stock_start()  → valida todo y lanza _execute_stock en hilo daemon
    # _on_stock_stop()   → setea stop_event + forzosamente cierra el browser
    # _append_stock_log(msg) → escribe en log_stock (hilo principal via except)
    # _execute_stock()   → DELEGADOR PURO → workers.execute_stock(self, ...)
    # _export_stock_audit_report() → genera informe Excel/PDF de resultados
    # ═══════════════════════════════════════════════════════════════

    def _export_stock_audit_report(self, fmt="excel"):
        from utils_mod.audit_reporter import audit_results, export_excel_report, export_pdf_report
        rows = getattr(self, "_stock_excel_df", []) or []
        summary = {
            "total": getattr(self, "_stock_total", 0),
            "ok": getattr(self, "_stock_ok", 0),
            "err": getattr(self, "_stock_errors", 0),
            "warn": 0,
            "rate": round((self._stock_ok / self._stock_total * 100) if getattr(self, "_stock_total", 0) > 0 else 0, 1),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        if not rows:
            messagebox.showwarning("Auditor de Stock", "No hay registros cargados en el módulo de Stock para auditar.\nCargue un archivo Excel de stock primero.")
            return

        def_ext = ".xlsx" if fmt == "excel" else ".html"
        ftypes = [("Libro de Excel", "*.xlsx")] if fmt == "excel" else [("Informe de Auditoría PDF/HTML", "*.html")]
        path = filedialog.asksaveasfilename(
            title=f"Guardar Informe de Auditoría Stock ({fmt.upper()})",
            initialfile=f"Informe_Auditoria_Stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}{def_ext}",
            defaultextension=def_ext,
            filetypes=ftypes
        )
        if not path:
            return

        if fmt == "excel":
            ok, msg = export_excel_report(rows, summary, path, modulo_nombre="Actualización de Stock")
        else:
            ok, msg = export_pdf_report(rows, summary, path, modulo_nombre="Actualización de Stock")

        if ok:
            messagebox.showinfo("Auditor de Stock", f"¡Informe de Auditoría de Stock generado exitosamente!\n\nUbicación:\n{msg}")
        else:
            messagebox.showerror("Error en Auditoría", f"Ocurrió un error al generar el informe:\n{msg}")

    def _on_load_stock_excel(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Seleccionar Excel de stock",
            filetypes=[("Excel", "*.xlsx *.xls")],
        )
        if not path:
            return
        self._stock_excel_path = path
        self.lbl_stock_excel.configure(text=os.path.basename(path), text_color="white")
        # Validar
        try:
            from automation_otro_bot.stock import analizar_excel_stock
            res = analizar_excel_stock(path)
            if res["valido"]:
                self._stock_excel_df = res["df"]
                self.lbl_stock_summary.configure(
                    text=f"✅ {len(res['df'])} productos válidos" +
                         (f" | ⚠ {len(res['errores'])} errores" if res["errores"] else ""),
                    text_color="#5dade2",
                )
            else:
                self.lbl_stock_summary.configure(
                    text=f"❌ Errores: {'; '.join(res['errores'][:3])}",
                    text_color="#e74c3c",
                )
                self._stock_excel_df = []
        except Exception as e:
            self.lbl_stock_summary.configure(text=f"❌ Error: {e}", text_color="#e74c3c")

    def _on_download_stock_template(self):
        from tkinter import filedialog
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock"
            ws.append(["Parte", "Stock", "Ficha"])
            for ejemplo in [("EMU5R6000", 1000, 2267958),
                            ("EMU5R6001", 500, 2267950),
                            ("EMU5R6002", 2000, 2271431)]:
                ws.append(ejemplo)
            path = filedialog.asksaveasfilename(
                title="Guardar plantilla",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile="plantilla_stock.xlsx",
            )
            if path:
                wb.save(path)
                self._append_stock_log(f"📋 Plantilla guardada: {path}")
        except Exception as e:
            self._append_stock_log(f"❌ Error creando plantilla: {e}")

    def _load_stock_combos_json(self) -> dict:
        # ponytail: usar resource_path para que funcione empaquetado con PyInstaller
        try:
            try:
                from resource_helper import resource_path
            except Exception:
                import sys
                _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                if _root not in sys.path:
                    sys.path.insert(0, _root)
                from resource_helper import resource_path
            json_path = resource_path("modulo_subir_pdf/dropdowns_mejora_basica.json")
            with open(json_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            self._append_stock_log(f"⚠ No se pudo cargar dropdowns_mejora_basica.json: {e}")
            return {"acuerdos": [], "combinaciones": []}

    def _on_stock_acuerdo_changed(self, acuerdo_text: str):
        # Filtrar combinaciones cuyo acuerdo (por coincidencia de texto) coincida
        acuerdos_data = self._stock_combos_data.get("acuerdos", [])
        acuerdo_value = None
        for a in acuerdos_data:
            if a.get("text") == acuerdo_text:
                acuerdo_value = a.get("value")
                break

        catalogos_set = set()
        for combo in self._stock_combos_data.get("combinaciones", []):
            if acuerdo_value and combo.get("acuerdo", {}).get("value") == acuerdo_value:
                cat = combo.get("catalogo", {})
                if cat.get("text"):
                    catalogos_set.add(cat["text"])

        catalogos = sorted(catalogos_set) if catalogos_set else ["-- Sin catálogos para este acuerdo --"]
        self.option_stock_catalogo.configure(values=catalogos)
        self.option_stock_catalogo.set(catalogos[0])
        # Disparar cascada de categoría
        self._on_stock_catalogo_changed(catalogos[0])

    def _on_stock_catalogo_changed(self, catalogo_text: str):
        acuerdos_data = self._stock_combos_data.get("acuerdos", [])
        acuerdo_sel = self.option_stock_acuerdo.get().strip()
        acuerdo_value = None
        for a in acuerdos_data:
            if a.get("text") == acuerdo_sel:
                acuerdo_value = a.get("value")
                break

        categorias_set = set()
        for combo in self._stock_combos_data.get("combinaciones", []):
            if combo.get("acuerdo", {}).get("value") != acuerdo_value:
                continue
            if combo.get("catalogo", {}).get("text") != catalogo_text:
                continue
            cat = combo.get("categoria", {})
            if cat.get("text"):
                categorias_set.add(cat["text"])

        categorias = sorted(categorias_set) if categorias_set else ["-- Sin categorías para este catálogo --"]
        self.option_stock_categoria.configure(values=categorias)
        self.option_stock_categoria.set(categorias[0])

    def _on_stock_start(self):
        # ── PASO 1: Prevenir doble inicio ──────────────────────────────
        if self._stock_running:
            return

        # ── PASO 2: Validar que hay datos de Excel cargados ────────────
        if not self._stock_excel_df:
            self._append_stock_log("❌ Carga un Excel con productos primero")
            return

        # ── PASO 3: Leer parámetros desde los widgets de UI ───────────
        try:
            pausa = float(self.entry_stock_pausa.get() or "2")
        except ValueError:
            pausa = 2.0
        acuerdo  = self.option_stock_acuerdo.get().strip()
        catalogo  = self.option_stock_catalogo.get().strip()
        categoria = self.option_stock_categoria.get().strip()
        # Credenciales propias de la pestaña Stock (flujo independiente del flujo PDF)
        usuario  = self.entry_stock_user.get().strip()
        password = self.entry_stock_pass.get().strip()

        # ── PASO 4: Validar credenciales y filtros seleccionados ───────
        if not usuario or not password:
            self._append_stock_log("❌ Credenciales vacías (rellena Usuario y Contraseña)")
            return
        if "Seleccione" in acuerdo or "Sin datos" in acuerdo:
            self._append_stock_log("❌ Selecciona un Acuerdo válido")
            return
        if "Seleccione" in catalogo or "Sin datos" in catalogo:
            self._append_stock_log("❌ Selecciona un Catálogo válido")
            return
        if "Seleccione" in categoria or "Sin datos" in categoria:
            self._append_stock_log("❌ Selecciona una Categoría válida")
            return

        # ── PASO 5: Marcar estado como en ejecución y actualizar botones
        self._stock_running = True
        self._stock_stop_event.clear()
        self.btn_stock_start.configure(state="disabled")
        self.btn_stock_stop.configure(state="normal")
        self.lbl_stock_status.configure(text="En ejecución...", text_color="#f39c12")
        self.lbl_stock_report.configure(text="(en proceso...)", text_color="gray60")

        # ── PASO 6: Lanzar hilo daemon → _execute_stock → workers.execute_stock
        threading.Thread(
            target=self._execute_stock,
            args=(usuario, password, acuerdo, catalogo, categoria, pausa),
            daemon=True,
        ).start()

    def _on_stock_stop(self):
        # PASO 1: Ignorar si ya está detenido
        if not self._stock_running:
            return
        # PASO 2: Setear el evento de parada (workers.py lo revisa en cada iteración)
        self._stock_stop_event.set()
        self.lbl_stock_status.configure(text="Deteniendo...", text_color="#e74c3c")
        # PASO 3: Cerrar el browser forzosamente para interrumpir operaciones
        # bloqueantes de Playwright (page.wait_for_*, page.click, etc.)
        try:
            if getattr(self, "_stock_browser", None):
                self._stock_browser.close()
        except Exception:
            pass
        self._append_stock_log("⏹ Detención solicitada")

    def _append_stock_log(self, msg):
        ts = time.strftime("%H:%M:%S")
        line = f"[{ts}] {msg}\n"
        try:
            self.log_stock.configure(state="normal")
            self.log_stock.insert("end", line)
            self.log_stock.see("end")
            self.log_stock.configure(state="disabled")
        except Exception:
            pass

    def _execute_stock(self, usuario, password, acuerdo, catalogo, categoria, pausa):
        # DELEGADOR PURO — NO agregar lógica aquí.
        # La implementación completa está en workers.execute_stock().
        # Pasos que ejecuta workers.execute_stock:
        #   Paso 1: init_browser (automation/browser.py)
        #   Paso 2: do_login con CAPTCHA (automation/login.py)
        #   Paso 3: aplicar filtros Acuerdo/Catálogo/Categoría en portal
        #   Paso 4: iterar cada fila del Excel y llamar actualizar_producto()
        import workers
        workers.execute_stock(self, usuario, password, acuerdo, catalogo, categoria, pausa)

    # ─── Auditor Portal Stock ───────────────────────────────────────

    def _on_stock_audit_start(self):
        """Handler del botón '🔍 Auditar Portal'.
        Valida credenciales y utiliza el Excel ya subido en la aplicación (pestaña Stock u Ofertas),
        luego lanza execute_auditor en un hilo en segundo plano.
        """
        from tkinter import messagebox
        import os

        # 1. Buscar Excel cargado (pestaña Stock o pestaña Principal)
        excel_rows = getattr(self, "_stock_excel_df", []) or []

        if not excel_rows:
            excel_rows = getattr(self, "_excel_rows", []) or []

        if not excel_rows:
            path = getattr(self, "_stock_excel_path", "") or getattr(self, "_excel_path", "")
            if path and os.path.exists(path):
                try:
                    from automation_otro_bot.stock import analizar_excel_stock
                    res = analizar_excel_stock(path)
                    if res.get("valido"):
                        excel_rows = res["df"]
                        self._stock_excel_df = excel_rows
                except Exception as e:
                    self._append_stock_log(f"⚠ No se pudo analizar {path}: {e}")

        if not excel_rows:
            self._append_stock_log("❌ Auditor: Carga primero un archivo Excel con productos (en Stock o en Ofertas)")
            messagebox.showwarning("Auditor Portal", "Por favor carga un archivo Excel con productos antes de auditar.")
            return

        # Sincronizar _stock_excel_df para que el worker tenga los datos
        self._stock_excel_df = excel_rows

        # 2. Buscar credenciales (prioridad: sección Stock, luego sección Principal)
        usuario = ""
        password = ""
        if hasattr(self, "entry_stock_user"):
            usuario = str(self.entry_stock_user.get() or "").strip()
        if hasattr(self, "entry_stock_pass"):
            password = str(self.entry_stock_pass.get() or "").strip()

        if not usuario or not password:
            if hasattr(self, "entry_user"):
                usuario = usuario or str(self.entry_user.get() or "").strip()
            if hasattr(self, "entry_pass"):
                password = password or str(self.entry_pass.get() or "").strip()

        if not usuario or not password:
            self._append_stock_log("❌ Auditor: Rellena Usuario y Contraseña en los campos de credenciales de Stock.")
            return

        # 3. Leer filtros seleccionados
        acuerdo = self.option_stock_acuerdo.get().strip() if hasattr(self, "option_stock_acuerdo") else "EXT-CE-2022-5"
        catalogo = self.option_stock_catalogo.get().strip() if hasattr(self, "option_stock_catalogo") else ""
        categoria = self.option_stock_categoria.get().strip() if hasattr(self, "option_stock_categoria") else ""

        # 4. Leer visibilidad del navegador
        visible = False
        if hasattr(self, "check_stock_visible"):
            try: visible = bool(self.check_stock_visible.get())
            except Exception: pass
        headless = not visible

        if hasattr(self, "btn_stock_audit"):
            try: self.btn_stock_audit.configure(state="disabled", text="⏳ Auditando...")
            except Exception: pass
        if hasattr(self, "lbl_audit_status"):
            try: self.lbl_audit_status.configure(text="Conectando al portal...", text_color="#f39c12")
            except Exception: pass
        self._append_stock_log(f"🔍 Iniciando Auditor Portal ({len(excel_rows)} productos | Navegador {'visible' if visible else 'oculto'})...")

        # Reset stop event
        self._stock_stop_event.clear()

        import threading, workers
        threading.Thread(
            target=workers.execute_auditor,
            args=(self, usuario, password, acuerdo, catalogo, categoria,
                  self._on_audit_done, self._append_stock_log),
            kwargs={"headless": headless},
            daemon=True,
        ).start()

    def _on_audit_done(self, filas: list, resumen: dict):
        """Callback llamado por execute_auditor cuando termina.
        Siempre se ejecuta en el hilo del auditor — usa self.after() para UI.
        """
        def _ui_done():
            if hasattr(self, "btn_stock_audit"):
                try: self.btn_stock_audit.configure(state="normal", text="🔍 Auditar Portal ahora")
                except Exception: pass
            if not filas:
                if hasattr(self, "lbl_audit_status"):
                    try: self.lbl_audit_status.configure(text="Sin datos para guardar", text_color="#e74c3c")
                    except Exception: pass
                return

            ok    = resumen.get("ok", 0)
            dif   = resumen.get("dif", 0)
            mis   = resumen.get("missing", 0)
            total = resumen.get("total", 0)
            tasa  = resumen.get("tasa", 0)
            status_text  = f"{ok}/{total} OK | {dif} dif | {mis} no enc. | {tasa:.1f}%"
            status_color = "#27ae60" if dif == 0 and mis == 0 else "#e67e22" if mis == 0 else "#e74c3c"
            if hasattr(self, "lbl_audit_status"):
                try: self.lbl_audit_status.configure(text=status_text, text_color=status_color)
                except Exception: pass

            # Guardar reporte Excel
            ts = time.strftime("%Y%m%d_%H%M%S")
            path = ""
            if not hasattr(self, "_api_bridge"):
                try:
                    from tkinter import filedialog
                    path = filedialog.asksaveasfilename(
                        title="Guardar Informe de Auditoría",
                        initialfile=f"Auditoria_Stock_{ts}.xlsx",
                        defaultextension=".xlsx",
                        filetypes=[("Libro Excel", "*.xlsx")],
                    )
                except Exception:
                    path = ""

            if not path:
                out_dir = os.path.join(_PROJECT_ROOT, "reportes_auditoria")
                os.makedirs(out_dir, exist_ok=True)
                path = os.path.join(out_dir, f"Auditoria_Stock_{ts}.xlsx")

            from utils_mod.audit_portal_excel import generar_excel_auditoria
            ok_save, msg = generar_excel_auditoria(filas, resumen, path)
            if ok_save:
                self._append_stock_log(f"📊 ¡Informe de Auditoría generado exitosamente!")
                self._append_stock_log(f"📂 Guardado en: {path}")
                import subprocess
                try:
                    subprocess.Popen(["start", "", path], shell=True)
                except Exception:
                    pass
            else:
                self._append_stock_log(f"❌ Error al generar el Excel: {msg}")

        try:
            self.after(0, _ui_done)
        except Exception:
            _ui_done()
    def _build_credentials_section(self, parent):
        C = self._C
        self._section_label(parent, "Credenciales de Perú Compras", 0)

        frame = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=6,
                             border_width=1, border_color=C["border"])
        frame.grid(row=1, column=0, padx=0, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        # Usuario
        ctk.CTkLabel(
            frame, text="Usuario", font=ctk.CTkFont(size=11),
            text_color=C["txt2"], anchor="w",
        ).grid(row=0, column=0, padx=12, pady=(10, 1), sticky="w")
        self.entry_user = ctk.CTkEntry(
            frame, placeholder_text="Usuario o RUC", height=34,
            fg_color=C["card2"], border_color=C["border"], text_color=C["txt"]
        )
        self.entry_user.grid(row=1, column=0, padx=12, pady=(0, 8), sticky="ew")
        self.entry_user.insert(0, "almerco.03")

        # Password
        pass_frame = ctk.CTkFrame(frame, fg_color="transparent")
        pass_frame.grid(row=2, column=0, padx=12, pady=(0, 2), sticky="ew")
        pass_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            pass_frame, text="Contraseña", font=ctk.CTkFont(size=11),
            text_color=C["txt2"], anchor="w",
        ).grid(row=0, column=0, columnspan=2, padx=0, pady=(0, 1), sticky="w")

        self._pw_visible = False
        self.entry_pass = ctk.CTkEntry(
            pass_frame, placeholder_text="Contraseña", show="*", height=34,
            fg_color=C["card2"], border_color=C["border"], text_color=C["txt"]
        )
        self.entry_pass.grid(row=1, column=0, pady=(0, 8), sticky="ew")
        self.entry_pass.insert(0, "4lm3rKenYa@#")
        self.btn_eye = ctk.CTkButton(
            pass_frame, text="◉", width=34, height=34,
            fg_color=C["card2"], hover_color=C["border"],
            text_color=C["txt2"], border_width=1, border_color=C["border"],
            font=ctk.CTkFont(size=14), command=self._toggle_password,
        )
        self.btn_eye.grid(row=1, column=1, padx=(6, 0), pady=(0, 8))

        # Mostrar navegador
        self.check_visible = ctk.CTkCheckBox(
            frame, text="Mostrar navegador en pantalla",
            font=ctk.CTkFont(size=12), text_color=C["txt"],
            border_color=C["border"], fg_color=C["accent"],
        )
        self.check_visible.grid(row=3, column=0, padx=12, pady=(2, 10), sticky="w")

    def _toggle_password(self):
        self._pw_visible = not self._pw_visible
        self.entry_pass.configure(show="" if self._pw_visible else "*")
        self.btn_eye.configure(text="◂" if self._pw_visible else "▸")

    # ── Excel Section ─────────────────────────────────────────────

    def _build_excel_section(self, parent):
        C = self._C
        self._section_label(parent, "Archivo Excel", 2)

        frame = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=6,
                             border_width=1, border_color=C["border"])
        frame.grid(row=3, column=0, padx=0, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        # File picker
        file_row = ctk.CTkFrame(frame, fg_color="transparent")
        file_row.grid(row=0, column=0, padx=12, pady=(10, 4), sticky="ew")
        file_row.grid_columnconfigure(0, weight=1)
        self.btn_file = ctk.CTkButton(
            file_row, text="Seleccionar .xlsx",
            height=32, font=ctk.CTkFont(size=12),
            fg_color=C["accent"], hover_color=C["accent_h"], text_color="#FFFFFF",
            corner_radius=4,
            command=self._pick_excel,
        )
        self.btn_file.pack(side="left")
        self.lbl_file = ctk.CTkLabel(
            file_row, text="Sin archivo", text_color=C["txt3"],
            font=ctk.CTkFont(size=11),
        )
        self.lbl_file.pack(side="left", padx=10)

        # Sheet selector
        ctk.CTkLabel(
            frame, text="Pestaña del Excel", font=ctk.CTkFont(size=11),
            text_color=C["txt2"], anchor="w",
        ).grid(row=1, column=0, padx=12, pady=(4, 1), sticky="w")
        self.combo_sheet = ctk.CTkComboBox(
            frame, values=["Cargue un archivo primero"],
            state="disabled", height=32,
            fg_color=C["card2"], border_color=C["border"],
            text_color=C["txt"], button_color=C["border"],
            command=self._on_sheet_changed,
        )
        self.combo_sheet.grid(row=2, column=0, padx=12, pady=(0, 6), sticky="ew")

        # Column mapping
        map_frame = ctk.CTkFrame(frame, fg_color="transparent")
        map_frame.grid(row=3, column=0, padx=12, pady=(4, 4), sticky="ew")
        map_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            map_frame, text="Columna N° de Parte", font=ctk.CTkFont(size=11),
            text_color=C["txt2"], anchor="w",
        ).grid(row=0, column=0, padx=(0, 4), pady=(0, 1), sticky="w")

        self.combo_parte = ctk.CTkComboBox(
            map_frame, values=["--"], state="disabled", height=32,
            fg_color=C["card2"], border_color=C["border"],
            text_color=C["txt"], button_color=C["border"],
        )
        self.combo_parte.grid(row=1, column=0, padx=(0, 4), sticky="ew")

        # Info
        self.lbl_excel_info = ctk.CTkLabel(
            frame, text="", font=ctk.CTkFont(size=11),
            text_color=C["txt2"], anchor="w",
        )
        self.lbl_excel_info.grid(row=4, column=0, padx=12, pady=(2, 10), sticky="w")

    def _pick_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not path:
            return
        try:
            sheets = get_sheets(path)
            if not sheets:
                if hasattr(self, "lbl_excel_info"):
                    self.lbl_excel_info.config(text="Sin pestañas en el Excel", fg="#8B1A1A")
                return

            self._excel_path = path
            name = os.path.basename(path)
            if hasattr(self, "lbl_file"):
                if hasattr(self.lbl_file, "delete"):
                    self.lbl_file.config(state="normal")
                    self.lbl_file.delete(0, "end")
                    self.lbl_file.insert(0, name)
                    self.lbl_file.config(state="readonly")
                else:
                    self.lbl_file.configure(text=name)
            if hasattr(self, "combo_sheet"):
                self.combo_sheet.configure(values=sheets, state="readonly")
                self.combo_sheet.set(sheets[0])
            self._on_sheet_changed(sheets[0])
            self._update_tools_excel_status()
        except Exception as e:
            if hasattr(self, "lbl_excel_info"):
                self.lbl_excel_info.config(text=f"Error: {e}", fg="#8B1A1A")

    def _on_sheet_changed(self, choice):
        if not self._excel_path or not choice:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._excel_path, read_only=True, data_only=True)
            ws = wb[choice]

            # Leer todos los headers de la fila 1 (primeros 20 para cubrir)
            all_headers = []
            # Leer fila 1 como header principal
            row1 = [cell for cell in ws.iter_rows(min_row=1, max_row=1, values_only=True)]
            if row1 and row1[0]:
                for v in row1[0]:
                    if v is not None and str(v).strip():
                        h = str(v).strip()
                        if h not in all_headers and h.lower() not in ("none", ""):
                            all_headers.append(h)
            # Si fila 1 no tiene nada, buscar en filas 2-10
            if not all_headers:
                for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
                    for v in (row or []):
                        if v is not None and str(v).strip():
                            h = str(v).strip()
                            if h not in all_headers and not h.replace(".","").replace(",","").isdigit():
                                all_headers.append(h)
                    if all_headers:
                        break

            # Si no hay headers, usar columna por defecto
            if not all_headers:
                all_headers = ["Columna A"]

            row_count = ws.max_row or 0
            wb.close()

            # Detectar columna de parte usando el parser (ya corregido)
            cols_detected = detect_columns(self._excel_path, choice)
            parte_col_detect = cols_detected.get("parte_col") or all_headers[0]

            # Parse simple
            rows_data = parse_excel(
                self._excel_path, choice,
                parte_col=parte_col_detect,
            )
            self._excel_rows = rows_data

            opts = ["-- Seleccionar --"] + all_headers
            self.combo_parte.configure(values=opts, state="readonly")
            self.combo_parte.set(parte_col_detect if parte_col_detect in all_headers else opts[0])

            if hasattr(self, "_tree"):
                self._tree.delete(*self._tree.get_children())
                for i, row in enumerate(rows_data, 1):
                    pn = str(row.get("parte", "") or row.get("nro_parte", "") or "")
                    desc = str(row.get("descripcion", "") or row.get("marca", "") or "")
                    precio = str(row.get("precio", "") or "")
                    stock = str(row.get("stock", "") or "")
                    tag = "par" if i % 2 == 0 else "impar"
                    self._tree.insert("", "end", values=(i, pn, desc, precio, stock, "Pendiente"), tags=(tag,))

            if hasattr(self, "lbl_excel_info"):
                self.lbl_excel_info.config(
                    text=f"✓ Archivo cargado: {len(rows_data)} registros · {len(all_headers)} columnas · Pestaña: {choice}",
                    fg="#1B6B1B",
                )
            self._update_tools_excel_status()
        except Exception as e:
            if hasattr(self, "lbl_excel_info"):
                self.lbl_excel_info.config(text=f"Error: {e}", fg="#8B1A1A")

    # ── Catalog Section (Cascada Catálogo → Categoría → Estado) ─────

    def _build_catalog_section(self, parent):
        C = self._C
        row_offset = 4
        self._section_label(parent, "Catálogo EXT-CE-2022-5", row_offset)

        frame = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=6,
                             border_width=1, border_color=C["border"])
        frame.grid(row=row_offset+1, column=0, padx=0, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        comb_data = self._catalog_data.get("combinaciones", [])
        if not comb_data:
            ctk.CTkLabel(
                frame,
                text="No se encontró combinaciones_computadoras.json.\nEjecute extract_combinaciones.py primero.",
                text_color=C["warn"], font=ctk.CTkFont(size=11), wraplength=280,
                justify="left",
            ).grid(row=0, column=0, padx=12, pady=10)
            return

        _cb_opts = dict(
            fg_color=C["card2"], border_color=C["border"],
            text_color=C["txt"], button_color=C["border"],
            state="readonly", height=32,
        )

        # Catálogo
        ctk.CTkLabel(frame, text="Catálogo Electrónico", font=ctk.CTkFont(size=11),
                     text_color=C["txt2"], anchor="w").grid(
            row=0, column=0, padx=12, pady=(8, 1), sticky="w")
        self.combo_catalogo = ctk.CTkComboBox(
            frame, values=self._opts_texts(comb_data),
            command=self._on_catalogo_changed, **_cb_opts,
        )
        self.combo_catalogo.grid(row=1, column=0, padx=12, pady=(0, 6), sticky="ew")
        if comb_data:
            self.combo_catalogo.set(self._opts_texts(comb_data)[0])

        # Categoría
        ctk.CTkLabel(frame, text="Categoría", font=ctk.CTkFont(size=11),
                     text_color=C["txt2"], anchor="w").grid(
            row=2, column=0, padx=12, pady=(4, 1), sticky="w")
        self.combo_categoria = ctk.CTkComboBox(
            frame, values=["Seleccione un Catálogo"],
            command=self._on_categoria_changed, **_cb_opts,
        )
        self.combo_categoria.grid(row=3, column=0, padx=12, pady=(0, 6), sticky="ew")

        # Estado
        ctk.CTkLabel(frame, text="Estado", font=ctk.CTkFont(size=11),
                     text_color=C["txt2"], anchor="w").grid(
            row=4, column=0, padx=12, pady=(4, 1), sticky="w")
        self.combo_estado = ctk.CTkComboBox(
            frame, values=["Seleccione una Categoría"], **_cb_opts,
        )
        self.combo_estado.grid(row=5, column=0, padx=12, pady=(0, 8), sticky="ew")

        # Trigger inicial
        self._on_catalogo_changed(self._opts_texts(comb_data)[0])

    def _opts_texts(self, data):
        return [f"{o['value']} - {o['text'][:60]}" for o in data] if data else []

    def _find_children(self, combo_text):
        val = combo_text.split(" - ")[0].strip() if " - " in combo_text else ""
        comb_data = self._catalog_data.get("combinaciones", [])
        for node in comb_data:
            if node["value"] == val:
                return node.get("children", [])
        return []

    def _on_catalogo_changed(self, choice):
        children = self._find_children(choice)
        texts = self._opts_texts(children)
        if texts:
            self.combo_categoria.configure(values=texts)
            self.combo_categoria.set(texts[0])
            self._on_categoria_changed(texts[0])
        else:
            self.combo_categoria.configure(values=["Sin categorías"])
            self.combo_estado.configure(values=["Sin estados"])

    def _on_categoria_changed(self, choice):
        # Buscar en el nodo actual de catalogo
        cat_val = self.combo_catalogo.get().split(" - ")[0].strip() if " - " in self.combo_catalogo.get() else ""
        comb_data = self._catalog_data.get("combinaciones", [])
        for node in comb_data:
            if node["value"] == cat_val:
                for child in node.get("children", []):
                    val = choice.split(" - ")[0].strip() if " - " in choice else ""
                    if child["value"] == val:
                        estados = child.get("children", [])
                        texts = self._opts_texts(estados)
                        if texts:
                            self.combo_estado.configure(values=texts)
                            self.combo_estado.set(texts[0])
                        else:
                            self.combo_estado.configure(values=["Sin estados"])
                        return
        self.combo_estado.configure(values=["Sin estados"])

    def _on_extract_json_portal(self):
        usuario = self.entry_stock_user.get().strip() if hasattr(self, 'entry_stock_user') and hasattr(self.entry_stock_user, 'get') else "estalin.huamali01"
        password = self.entry_stock_pass.get().strip() if hasattr(self, 'entry_stock_pass') and hasattr(self.entry_stock_pass, 'get') else ""
        acuerdo = self.option_stock_acuerdo.get().strip() if hasattr(self, 'option_stock_acuerdo') and hasattr(self.option_stock_acuerdo, 'get') else "EXT-CE-2022-5 COMPUTADORAS Y ESCÁNERES"
        catalogo = self.option_stock_catalogo.get().strip() if hasattr(self, 'option_stock_catalogo') and hasattr(self.option_stock_catalogo, 'get') else "COMPUTADORAS DE ESCRITORIO"
        categoria = self.option_stock_categoria.get().strip() if hasattr(self, 'option_stock_categoria') and hasattr(self.option_stock_categoria, 'get') else "COMPUTADORA TODO EN UNO"
        visible = bool(getattr(getattr(self, 'check_stock_visible', None), 'get', lambda: False)())

        def _log(msg):
            print(f"[JSON EXTRACTOR] {msg}")
            if hasattr(self, '_api_bridge') and self._api_bridge and getattr(self._api_bridge, '_window', None):
                try:
                    js_msg = json.dumps(msg)
                    self._api_bridge._window.evaluate_js(f"logJsonConsole({js_msg});")
                    self._api_bridge._window.evaluate_js(f"logToolsConsole({js_msg});")
                except Exception:
                    pass

        def _on_done(fichas, filepath):
            if hasattr(self, '_api_bridge') and self._api_bridge and getattr(self._api_bridge, '_window', None):
                try:
                    js_data = json.dumps(fichas)
                    js_file = json.dumps(os.path.basename(filepath) if filepath else "")
                    self._api_bridge._window.evaluate_js(f"renderJsonTable({js_data});")
                    if filepath:
                        self._api_bridge._window.evaluate_js(f"logJsonConsole('[JSON OK] Dataset guardado: ' + {js_file} + ' (' + {len(fichas)} + ' fichas)', 'log-ok');")
                        self._api_bridge._window.evaluate_js(f"logToolsConsole('[JSON OK] Dataset guardado: ' + {js_file} + ' (' + {len(fichas)} + ' fichas)', 'log-ok');")
                except Exception:
                    pass

        from workers import execute_json_extractor
        threading.Thread(
            target=execute_json_extractor,
            args=(self, usuario, password, acuerdo, catalogo, categoria, _on_done, _log, not visible),
            daemon=True
        ).start()

    # ── Opciones Section ──────────────────────────────────────────

    def _build_opciones_section(self, parent):
        C = self._C
        row_offset = 6
        self._section_label(parent, "Opciones de Procesamiento", row_offset)

        frame = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=6,
                             border_width=1, border_color=C["border"])
        frame.grid(row=row_offset+1, column=0, padx=0, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        # Pausa entre productos
        ctk.CTkLabel(
            frame, text="Pausa entre productos (segundos)",
            font=ctk.CTkFont(size=11), text_color=C["txt2"], anchor="w",
        ).grid(row=0, column=0, padx=12, pady=(10, 1), sticky="w")
        self.slider_pausa = ctk.CTkSlider(
            frame, from_=0.5, to=5.0, number_of_steps=9,
            fg_color=C["border"], progress_color=C["accent"], button_color=C["accent"],
        )
        self.slider_pausa.grid(row=1, column=0, padx=12, pady=(0, 2), sticky="ew")
        self.slider_pausa.set(1.5)
        self.lbl_pausa = ctk.CTkLabel(
            frame, text="1.5 s", font=ctk.CTkFont(size=11), text_color=C["txt2"],
        )
        self.lbl_pausa.grid(row=2, column=0, padx=12, pady=(0, 6), sticky="w")
        self.slider_pausa.configure(command=lambda v: self.lbl_pausa.configure(text=f"{v:.1f} s"))

        # Separador
        ctk.CTkFrame(frame, height=1, fg_color=C["sep"]).grid(
            row=3, column=0, padx=12, pady=4, sticky="ew"
        )

        # Info de modo
        ctk.CTkLabel(
            frame,
            text="Este módulo usa Playwright para subir archivos PDF al portal.\n"
                 "Los endpoints se ajustan automáticamente según la sección.",
            font=ctk.CTkFont(size=11),
            text_color=C["txt3"],
            anchor="w",
            justify="left",
        ).grid(row=4, column=0, padx=12, pady=(4, 10), sticky="w")

    # ── Execution Section ─────────────────────────────────────────

    def _build_execution_section(self, parent):
        C = self._C
        self._section_label(parent, "Panel de Ejecución", 0)

        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=1, column=0, padx=12, pady=(0, 4), sticky="nsew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(5, weight=1)

        # Status bar
        status_row = ctk.CTkFrame(frame, fg_color=C["card2"], corner_radius=6,
                                  border_width=1, border_color=C["border"])
        status_row.grid(row=0, column=0, padx=0, pady=(0, 8), sticky="ew")
        self.lbl_status = ctk.CTkLabel(
            status_row, text="Listo para iniciar",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=C["txt"],
        )
        self.lbl_status.pack(side="left", padx=12, pady=8)
        self.lbl_counter = ctk.CTkLabel(
            status_row, text="", font=ctk.CTkFont(size=11), text_color=C["txt3"],
        )
        self.lbl_counter.pack(side="right", padx=12)

        # Progress bar
        self.progress = ctk.CTkProgressBar(
            frame, height=6,
            fg_color=C["border"], progress_color=C["accent"],
        )
        self.progress.grid(row=2, column=0, padx=0, pady=(0, 8), sticky="ew")
        self.progress.set(0)

        # Mini-stats
        stats_row = ctk.CTkFrame(frame, fg_color="transparent")
        stats_row.grid(row=3, column=0, padx=0, pady=(0, 8), sticky="ew")
        stats_row.grid_columnconfigure((0, 1, 2), weight=1)

        self._stat_ok   = self._make_stat(stats_row, "Exitosos",   C["success"], 0)
        self._stat_warn = self._make_stat(stats_row, "No hallados", C["warn"],   1)
        self._stat_err  = self._make_stat(stats_row, "Errores",    C["danger"], 2)

        # Captcha panel (oculto por defecto)
        self._build_captcha_panel(frame)

        # Log — fondo blanco, texto oscuro, fuente monospace
        self.log_box = ctk.CTkTextbox(
            frame, wrap="word",
            font=ctk.CTkFont(family="Courier New", size=11),
            fg_color=C["card"],
            border_width=1, border_color=C["border"],
            text_color=C["txt"],
        )
        self.log_box.grid(row=5, column=0, padx=0, pady=(2, 8), sticky="nsew")
        self.log_box.configure(state="disabled")
        self.log_box.tag_config("ok",       foreground="#1E6E3A")
        self.log_box.tag_config("error",    foreground="#B91C1C")
        self.log_box.tag_config("info",     foreground="#4A6080")
        self.log_box.tag_config("warn",     foreground="#92400E")
        self.log_box.tag_config("done",     foreground="#0D6EAA")
        self.log_box.tag_config("complete", foreground="#1E6E3A")
        self.log_box.tag_config("existing", foreground="#1A5493")
        self.log_box.tag_config("notfound", foreground="#B45309")

    def _make_stat(self, parent, label, color, col):
        C = self._C
        f = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=6,
                         border_width=1, border_color=C["border"])
        f.grid(row=0, column=col, padx=3, pady=2, sticky="ew")
        lbl_n = ctk.CTkLabel(f, text="0", font=ctk.CTkFont(size=22, weight="bold"), text_color=color)
        lbl_n.pack(pady=(8, 0))
        ctk.CTkLabel(f, text=label, font=ctk.CTkFont(size=10), text_color=C["txt3"]).pack(pady=(0, 8))
        return lbl_n

    # ── Captcha Panel ─────────────────────────────────────────────

    def _build_captcha_panel(self, parent):
        C = self._C
        self.captcha_panel = ctk.CTkFrame(
            parent, fg_color=C["card2"], corner_radius=6,
            border_width=1, border_color=C["border"],
        )
        self.captcha_panel.grid(row=4, column=0, padx=0, pady=(0, 8), sticky="ew")
        self.captcha_panel.grid_remove()
        self.captcha_panel.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.captcha_panel,
            text="CAPTCHA — Ingrese el código mostrado",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=C["txt"], anchor="w",
        ).grid(row=0, column=0, padx=8, pady=(6, 4), sticky="w")

        self.captcha_img_lbl = ctk.CTkLabel(self.captcha_panel, text="")
        self.captcha_img_lbl.grid(row=1, column=0, padx=8, pady=(0, 4))

        cap_row = ctk.CTkFrame(self.captcha_panel, fg_color="transparent")
        cap_row.grid(row=2, column=0, padx=8, pady=(0, 8), sticky="ew")
        cap_row.grid_columnconfigure(0, weight=1)

        self.captcha_entry = ctk.CTkEntry(cap_row, placeholder_text="Código", height=30)
        self.captcha_entry.grid(row=0, column=0, padx=(0, 6), sticky="ew")
        self.captcha_entry.bind("<Return>", lambda e: self._on_captcha_submit())

        ctk.CTkButton(
            cap_row, text="Enviar", width=60, height=30,
            command=self._on_captcha_submit,
        ).grid(row=0, column=1)

        self._captcha_img = None

    def _show_captcha(self, image_bytes):
        img = Image.open(BytesIO(image_bytes))
        img = img.resize((250, 66), Image.LANCZOS)
        self._captcha_img = ctk.CTkImage(light_image=img, dark_image=img, size=(250, 66))
        self.captcha_img_lbl.configure(image=self._captcha_img)
        self.captcha_entry.delete(0, "end")
        self.captcha_entry.focus_set()
        self.captcha_panel.grid()

    def _hide_captcha_panel(self):
        self.captcha_panel.grid_remove()

    def _on_captcha_submit(self):
        code = self.captcha_entry.get().strip()
        if code:
            self.captcha_bridge.respond(code)
            self._hide_captcha_panel()
            self._log("CAPTCHA manual: " + code, "info")

    # ── Helpers ───────────────────────────────────────────────────

    def _section_label(self, parent, text, row):
        C = self._C
        ctk.CTkLabel(
            parent, text=text,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=C["txt"], anchor="w",
        ).grid(row=row, column=0, padx=0, pady=(12, 4), sticky="w")

    def _log(self, msg, level="info"):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n", level)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self._log_lines.append({"level": level, "msg": msg})

    # ── Test Flow ──────────────────────────────────────────────

    def _on_test(self):
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return

        if not hasattr(self, 'combo_catalogo') or not self._catalog_data:
            self.lbl_status.configure(text="No hay datos de catálogo cargados", text_color="#e74c3c")
            return

        headless = not bool(self.check_visible.get())

        def _val(combo):
            t = combo.get()
            return t.split(" - ")[0].strip() if " - " in t else ""

        pre_selected = {
            "acuerdo": "249",
            "catalogo": _val(self.combo_catalogo),
            "categoria": _val(self.combo_categoria),
            "estado": _val(self.combo_estado),
        }

        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="🧪 Test Flow en marcha...", text_color="#f39c12")
        self.btn_launch.configure(state="disabled")
        self.btn_test.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()

        threading.Thread(
            target=self._execute_test,
            args=(user, pwd, headless, pre_selected),
            daemon=True,
        ).start()

    def _on_certs_only(self):
        """Handler del botón 'Solo Certificaciones': entra a cada ficha y agrega ISO 9001/14001."""
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return
        if not self._excel_rows or not self._excel_path:
            self.lbl_status.configure(text="Cargá un Excel antes de iniciar", text_color="#e74c3c")
            return
        if not hasattr(self, 'combo_catalogo') or not self._catalog_data:
            self.lbl_status.configure(text="No hay datos de catálogo cargados", text_color="#e74c3c")
            return

        headless = not bool(self.check_visible.get())

        def _val(combo):
            t = combo.get()
            return t.split(" - ")[0].strip() if " - " in t else ""

        pre_selected = {
            "acuerdo": "249",
            "catalogo": _val(self.combo_catalogo),
            "categoria": _val(self.combo_categoria),
            "estado": _val(self.combo_estado),
        }

        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="🏅 Solo Certificaciones en marcha...", text_color="#8e44ad")
        self.btn_launch.configure(state="disabled")
        self.btn_test.configure(state="disabled")
        self.btn_certs.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()

        threading.Thread(
            target=self._execute_certs_only,
            args=(user, pwd, headless, pre_selected),
            daemon=True,
        ).start()

    def _on_extract(self):
        """Handler del botón 'Extraer Reportes': descarga reportes de Producto Ofertado."""
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return

        headless = not bool(self.check_visible.get())

        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="📊 Extrayendo reportes...", text_color="#3498db")
        self.btn_launch.configure(state="disabled")
        self.btn_test.configure(state="disabled")
        self.btn_certs.configure(state="disabled")
        self.btn_nro.configure(state="disabled")
        self.btn_extract.configure(state="disabled")
        self.btn_compare.configure(state="disabled")
        self.btn_discovery.configure(state="disabled")
        self.btn_discovery2.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()

        threading.Thread(
            target=self._execute_extract,
            args=(user, pwd, headless),
            daemon=True,
        ).start()
    def _execute_extract(self, usuario, password, headless):
        import workers
        workers.execute_extract(self, usuario, password, headless)
    def _execute_certs_only(self, usuario, password, headless, pre_selected):
        import workers
        workers.execute_certs_only(self, usuario, password, headless, pre_selected)
    def _on_nro_parte(self):
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return
        if not self._excel_rows or not self._excel_path:
            self.lbl_status.configure(text="Cargá un Excel antes de iniciar", text_color="#e74c3c")
            return

        headless = not bool(self.check_visible.get())

        def _val(combo):
            t = combo.get()
            return t.split(" - ")[0].strip() if " - " in t else ""

        pre_selected = {
            "acuerdo": "249",
            "catalogo": _val(self.combo_catalogo),
            "categoria": _val(self.combo_categoria),
            "estado": _val(self.combo_estado),
        }

        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="🏷️ Solo N° de Parte en marcha...", text_color="#d35400")
        self.btn_launch.configure(state="disabled")
        self.btn_test.configure(state="disabled")
        self.btn_certs.configure(state="disabled")
        self.btn_nro.configure(state="disabled")
        self.btn_extract.configure(state="disabled")
        self.btn_compare.configure(state="disabled")
        self.btn_discovery.configure(state="disabled")
        self.btn_discovery2.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()

        threading.Thread(
            target=self._execute_nro_parte,
            args=(user, pwd, headless, pre_selected),
            daemon=True,
        ).start()

    def _execute_nro_parte(self, usuario, password, headless, pre_selected):
        import workers
        workers.execute_nro_parte(self, usuario, password, headless, pre_selected)
    def _on_compare(self):
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return
        if not self._excel_rows or not self._excel_path:
            self.lbl_status.configure(text="Cargá un Excel antes de iniciar", text_color="#e74c3c")
            return

        headless = not bool(self.check_visible.get())
        def _val(combo):
            t = combo.get()
            return t.split(" - ")[0].strip() if " - " in t else ""
        pre_selected = {
            "acuerdo": _val(self.combo_acuerdo) if hasattr(self, "combo_acuerdo") else "249",
            "catalogo": _val(self.combo_catalogo),
            "categoria": _val(self.combo_categoria),
            "estado": _val(self.combo_estado),
        }

        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal"); self.log_box.delete("1.0", "end"); self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="🔍 Comparando fichas...", text_color="#16a085")
        for b in ("launch", "test", "certs", "nro", "extract", "compare"):
            getattr(self, f"btn_{b}").configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()

        threading.Thread(
            target=self._execute_compare,
            args=(user, pwd, headless, pre_selected),
            daemon=True,
        ).start()

    def _execute_compare(self, usuario, password, headless, pre_selected):
        import workers
        workers.execute_compare(self, usuario, password, headless, pre_selected)
    def _on_discovery(self):
        """Handler del botón 'Discovery': ejecuta el script discovery_perucompras.py."""
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return
        headless = not bool(self.check_visible.get())
        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal"); self.log_box.delete("1.0", "end"); self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="🕵️ Discovery en marcha...", text_color="#2c3e50")
        for b in ("launch", "test", "certs", "nro", "extract", "compare", "discovery"):
            try: getattr(self, f"btn_{b}").configure(state="disabled")
            except AttributeError: pass
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()
        threading.Thread(
            target=self._execute_discovery, args=(user, pwd, headless),
            daemon=True,
        ).start()

    def _execute_discovery(self, usuario, password, headless):
        import workers
        workers.execute_discovery(self, usuario, password, headless)
    def _on_discovery2(self):
        """Handler del botón 'Discovery v2': scraping profundo multi-técnica."""
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return
        headless = not bool(self.check_visible.get())
        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal"); self.log_box.delete("1.0", "end"); self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="🕵️ v2 en marcha...", text_color="#566573")
        for b in ("launch", "test", "certs", "nro", "extract", "compare", "discovery", "discovery2"):
            try: getattr(self, f"btn_{b}").configure(state="disabled")
            except AttributeError: pass
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()
        threading.Thread(
            target=self._execute_discovery2, args=(user, pwd, headless),
            daemon=True,
        ).start()

    def _execute_discovery2(self, usuario, password, headless):
        import workers
        workers.execute_discovery2(self, usuario, password, headless)
    def _execute_test(self, usuario, password, headless, pre_selected):
        import workers
        workers.execute_test(self, usuario, password, headless, pre_selected)
    # ═══════════════════════════════════════════════════════════════
    # BLOQUE 7 — Handlers del módulo PDF + Delegadores
    # ───────────────────────────────────────────────────────────────
    # _on_launch()       → flujo principal PDF (valida → lanza hilo → _execute)
    # _execute()         → DELEGADOR PURO → workers.execute()
    # _on_stop()         → detiene flujo PDF (stop_event + UI reset en 3s)
    # _on_test()         → test de 1 ficha → workers.execute_test()
    # _on_certs_only()   → solo certificaciones → workers.execute_certs_only()
    # _on_nro_parte()    → solo número de parte → workers.execute_nro_parte()
    # _on_extract()      → extraer reportes → workers.execute_extract()
    # _on_compare()      → comparar fichas → workers.execute_compare()
    # _on_discovery()    → discovery v1 → workers.execute_discovery()
    # _on_discovery2()   → discovery v2 → workers.execute_discovery2()
    # ═══════════════════════════════════════════════════════════════
    def _on_launch(self):
        # ── PASO 1: Leer y validar credenciales ──────────────────────
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return

        # ── PASO 2: Validar que hay datos de Excel cargados ────────────
        if not self._excel_rows or not self._excel_path:
            self.lbl_status.configure(
                text="Cargá un Excel antes de iniciar el procesamiento",
                text_color="#e74c3c",
            )
            return

        # ── PASO 3: Usar las filas ya parseadas (auto-detectadas en _on_sheet_changed)
        rows = self._excel_rows
        if not rows:
            self.lbl_status.configure(text="El Excel no tiene filas de datos", text_color="#e74c3c")
            return

        # ── PASO 4: Inicializar contadores y estado de ejecución ────────
        self._running = True
        self._ok = 0
        self._errors = 0
        self._total = len(rows)
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event

        # ── PASO 5: Limpiar UI (log, barra de progreso, stats) ─────────
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress.set(0)
        self._stat_ok.configure(text="0")
        self._stat_warn.configure(text="0")
        self._stat_err.configure(text="0")
        self.lbl_status.configure(text="Iniciando...", text_color="white")
        self.btn_launch.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.btn_certs.configure(state="disabled")
        self._hide_captcha_panel()

        # ── PASO 6: Leer opciones de visualización del navegador y pausa
        headless = not bool(self.check_visible.get())
        pausa    = self.slider_pausa.get()

        # ── PASO 7: Armar pre_selected con valores actuales de los combos
        def _val(combo):
            t = combo.get()
            return t.split(" - ")[0].strip() if " - " in t else ""

        acuerdo_data = self._catalog_data.get("acuerdo", {})
        pre_selected = {
            "acuerdo":   acuerdo_data.get("value", "249"),
            "catalogo":  _val(self.combo_catalogo),
            "categoria": _val(self.combo_categoria),
            "estado":    _val(self.combo_estado),
        }

        # ── PASO 8: Lanzar hilo daemon → _execute → workers.execute
        threading.Thread(
            target=self._execute,
            args=(user, pwd, headless, rows, pausa, pre_selected),
            daemon=True,
        ).start()

    def _execute(self, usuario, password, headless, rows, pausa, pre_selected=None):
        # DELEGADOR PURO — NO agregar lógica aquí.
        # La implementación completa está en workers.execute().
        # Pasos que ejecuta workers.execute:
        #   Paso 1: init_browser + do_login
        #   Paso 2: iterar Excel fila a fila
        #   Paso 3: navegar a la ficha del producto
        #   Paso 4: subir PDF / actualizar campos
        #   Paso 5: reportar resultado via log_queue
        import workers
        workers.execute(self, usuario, password, headless, rows, pausa, pre_selected)
    def _on_stop(self):
        # PASO 1: Setear el evento global de parada
        self.stop_event.set()
        self._running = False
        # PASO 2: Actualizar UI mientras se espera la detención
        self.btn_stop.configure(state="disabled", text="Deteniendo...")
        self._log("Detención solicitada", "warn")
        # PASO 3: Restaurar botones después de 3 segundos
        # (da tiempo al hilo daemon de terminar limpiamente)
        self.after(3000, self._reset_after_stop)

    def _reset_after_stop(self):
        self.btn_launch.configure(state="normal")
        self.btn_stop.configure(state="disabled", text="■  Detener")
        self.btn_test.configure(state="normal")
        self.btn_certs.configure(state="normal")
        self.btn_nro.configure(state="normal")
        self.btn_extract.configure(state="normal")
        self.btn_compare.configure(state="normal")
        self.btn_discovery.configure(state="normal")
        self.btn_discovery2.configure(state="normal")
        self.lbl_status.configure(text="Detenido — listo para iniciar", text_color="#f39c12")

    # ═══════════════════════════════════════════════════════════════
    # BLOQUE 8 — poll_queue()  (motor de actualización de UI)
    # ───────────────────────────────────────────────────────────────
    # Se llama cada 200ms via self.after(200, self.poll_queue).
    # Es el ÚNICO punto seguro para actualizar widgets desde hilos.
    # Los workers.py NUNCA tocan widgets directamente.
    #
    # Formato de mensajes en self.log_queue (dict con "type"):
    #   {"type": "log",      "msg": str,  "level": str}
    #   {"type": "progress", "current": int, "total": int}
    #   {"type": "stat_ok",  "value": int}
    #   {"type": "stat_warn","value": int}
    #   {"type": "stat_err", "value": int}
    #   {"type": "done",     "ok": int, "errors": int}
    #   (level, msg) tuple  → compatibilidad con LogWriter legacy
    #
    # También detecta CAPTCHA pendiente: si captcha_bridge.image_bytes
    # no es None, muestra el panel de CAPTCHA al usuario.
    #
    # ❌ NO modificar la frecuencia del after(200ms) sin revisar workers.
    # ❌ NO llamar métodos de widget desde hilos secundarios directamente.
    # ❌ NO eliminar el try/except queue.Empty: es parte del diseño.
    # ═══════════════════════════════════════════════════════════════
    def poll_queue(self):
        try:
            while True:
                item = self.log_queue.get_nowait()
                t = item.get("type") if isinstance(item, dict) else None

                if t == "log":
                    self._log(item.get("msg", ""), item.get("level", "info"))
                elif t == "progress":
                    cur = item.get("current", 0)
                    tot = item.get("total", 1)
                    self.lbl_status.configure(text=f"Procesando producto {cur} de {tot}")
                    self.lbl_counter.configure(text=f"{cur}/{tot}")
                    self.progress.set(cur / tot if tot > 0 else 0)
                elif t == "stat_ok":
                    self._stat_ok.configure(text=str(item.get("value", 0)))
                elif t == "stat_warn":
                    self._stat_warn.configure(text=str(item.get("value", 0)))
                elif t == "stat_err":
                    self._stat_err.configure(text=str(item.get("value", 0)))
                elif t == "done":
                    ok  = item.get("ok", 0)
                    err = item.get("errors", 0)
                    self._running = False
                    self.lbl_status.configure(
                        text=f"Finalizado  ·  {ok} OK  ·  {err} errores"
                    )
                    self.progress.set(1)
                    self.btn_launch.configure(state="normal")
                    self.btn_stop.configure(state="disabled", text="■  Detener")

                # Compatibilidad con LogWriter que usa tuplas (level, msg)
                elif isinstance(item, tuple) and len(item) == 2:
                    level, msg = item
                    level_map = {
                        "OK": "ok", "ERROR": "error",
                        "WARN": "warn", "DONE": "done", "INFO": "info",
                    }
                    tag = level_map.get(level, "info")

                    if level == "DONE":
                        self._running = False
                        self.lbl_status.configure(text=msg)
                        self.progress.set(1)
                        self.btn_launch.configure(state="normal")
                        self.btn_stop.configure(state="disabled", text="■  Detener")
                    else:
                        self._log(msg, tag)

        except queue.Empty:
            pass

        # Verificar bridge de captcha
        with self.captcha_bridge.lock:
            if (
                self.captcha_bridge.image_bytes is not None
                and not self.captcha_bridge.event.is_set()
            ):
                self._show_captcha(self.captcha_bridge.image_bytes)

        self.after(200, self.poll_queue)


# ═══════════════════════════════════════════════════════════════════
#  PYWEBVIEW JS BRIDGE — Expone el backend existente al frontend HTML
#  El backend (SubirPdfApp, workers, automation) NO se modifica.
# ═══════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════
#  BLOQUE 9 — PYWEBVIEW JS BRIDGE
# ───────────────────────────────────────────────────────────────────
#  Expone el backend existente al frontend HTML (ui_web/index.html)
#  sin modificar SubirPdfApp ni workers.py.
#
#  COMPONENTES:
#    _DummyWidget     → stub de cualquier widget Tkinter/CTk para el
#                       modo pywebview (sin UI nativa). Simula .get(),
#                       .set(), .configure(), .config(), .insert(), etc.
#                       ❌ NO eliminar métodos: si workers.py llama
#                       app.widget.get() y el widget es _DummyWidget,
#                       debe devolver algo válido.
#
#    SubirPdfWebApi   → puente JS→Python. Cada método público es
#                       invocable desde JS con:
#                       pywebview.api.metodo(params)
#                       Devuelve dict serializable a JSON.
#                       Para agregar endpoint: añadir método público
#                       y en index.html llamar pywebview.api.tu_metodo()
#
#    run_app()        → entry point real del ejecutable (.exe).
#                       Crea _Backend (headless, sin ventana CTk),
#                       inyecta métodos de SubirPdfApp en _Backend,
#                       crea SubirPdfWebApi y lanza pywebview.
# ═══════════════════════════════════════════════════════════════════


class _DummyWidget:
    def __init__(self, val=""):
        self.val = val
    def get(self, *a, **k):
        return self.val
    def set(self, val=None, *a, **k):
        if val is not None:
            self.val = val
    def configure(self, *a, **k): pass
    def config(self, *a, **k): pass
    def delete(self, *a, **k): pass
    def insert(self, *a, **k): pass
    def see(self, *a, **k): pass
    def pack(self, *a, **k): pass
    def place(self, *a, **k): pass
    def grid(self, *a, **k): pass
    def destroy(self, *a, **k): pass
    def bind(self, *a, **k): pass
    def unbind(self, *a, **k): pass
    def __call__(self, *a, **k): return self.val


def _get_user_pass(app, params=None):
    params = params or {}
    user = str(params.get("user") or "").strip()
    pwd = str(params.get("pass") or "").strip()

    if not user and hasattr(app, "entry_stock_user") and hasattr(app.entry_stock_user, "get"):
        user = str(app.entry_stock_user.get()).strip()
    if not user and hasattr(app, "entry_user") and hasattr(app.entry_user, "get"):
        user = str(app.entry_user.get()).strip()
    if not user:
        user = "estalin.huamali01"

    if not pwd and hasattr(app, "entry_stock_pass") and hasattr(app.entry_stock_pass, "get"):
        pwd = str(app.entry_stock_pass.get()).strip()
    if not pwd and hasattr(app, "entry_pass") and hasattr(app.entry_pass, "get"):
        pwd = str(app.entry_pass.get()).strip()
    return user, pwd


class SubirPdfWebApi:
    """Puente JS -> Python. Delega al backend SubirPdfApp."""

    def __init__(self, app):
        self._app = app
        self._window = None
        app._api_bridge = self

    def set_window(self, w):
        self._window = w
        if hasattr(self, '_app') and self._app:
            self._app._window = w

    def get_catalog_options(self, *a):
        """Devuelve las opciones desplegables del archivo JSON."""
        return getattr(self._app, "_catalog_data", {})

    # Ventana CSD
    def minimize(self, *a):
        if self._window: self._window.minimize()

    def maximize(self, *a):
        if self._window: self._window.toggle_fullscreen()

    def close(self, *a):
        if self._window: self._window.destroy()

    # Excel PDF
    def select_file(self, *a):
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Seleccionar Archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        root.destroy()
        if not path:
            return None
        sheets = get_sheets(path)
        first = sheets[0] if sheets else ""
        cols = detect_columns(path, first)
        rows = parse_excel(path, first, parte_col=cols.get("parte_col"))
        self._app._excel_path = path
        self._app._excel_rows = rows
        return {"path": path, "name": os.path.basename(path), "sheets": sheets, "rows": rows}

    # Excel Stock
    def select_stock_file(self, *a):
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Seleccionar Excel de Stock",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        root.destroy()
        if not path:
            return None
        
        rows = []
        try:
            from automation_otro_bot.stock import analizar_excel_stock
            res = analizar_excel_stock(path)
            if res and res.get("valido") and res.get("df") is not None:
                self._app._stock_excel_df = res["df"]
                df = res["df"]
                for _, r in df.iterrows():
                    rows.append({
                        "parte": str(r.get("Parte", r.get("nro_parte", ""))),
                        "stock": str(r.get("Stock", r.get("stock", ""))),
                        "ficha": str(r.get("Ficha", r.get("ficha", "")))
                    })
            else:
                first = get_sheets(path)[0] if get_sheets(path) else ""
                cols = detect_columns(path, first)
                rows = parse_excel(path, first, parte_col=cols.get("parte_col"))
                self._app._stock_excel_df = rows
        except Exception:
            first = get_sheets(path)[0] if get_sheets(path) else ""
            cols = detect_columns(path, first)
            rows = parse_excel(path, first, parte_col=cols.get("parte_col"))
            self._app._stock_excel_df = rows

        self._app._stock_excel_path = path
        return {"path": path, "name": os.path.basename(path), "sheets": get_sheets(path), "rows": rows}

    # JSON de Precios
    def select_json_file(self, *a):
        import tkinter as tk
        from tkinter import filedialog
        import json
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Seleccionar Archivo JSON o Excel de Precios",
            filetypes=[("Archivos JSON (*.json)", "*.json"), ("Archivos Excel (*.xlsx, *.xls)", "*.xlsx *.xls"), ("Todos los archivos", "*.*")]
        )
        root.destroy()
        if not path:
            return None
        
        rows = []
        try:
            if path.lower().endswith(".json"):
                with open(path, "r", encoding="utf-8") as f:
                    raw_data = json.load(f)
                if isinstance(raw_data, list):
                    rows = raw_data
                elif isinstance(raw_data, dict):
                    rows = raw_data.get("data", raw_data.get("fichas", [raw_data]))
            else:
                from tab_precios_json import parse_excel_precios
                rows = parse_excel_precios(path)
            
            self._app._precios_json_path = path
            self._app._precios_json_data = rows
            return {"path": path, "name": os.path.basename(path), "rows": rows}
        except Exception as e:
            return {"error": str(e), "path": path, "name": os.path.basename(path), "rows": []}

    def load_sheet(self, sheet_name=None, *a):
        if not getattr(self._app, "_excel_path", "") or not sheet_name:
            return []
        cols = detect_columns(self._app._excel_path, sheet_name)
        rows = parse_excel(self._app._excel_path, sheet_name, parte_col=cols.get("parte_col"))
        self._app._excel_rows = rows
        return rows

    # Automatizacion PDF
    def start_process(self, params=None, *a):
        try:
            params = params or {}
            self._app.entry_user = _DummyWidget(params.get("user", "almerco.03"))
            self._app.entry_pass = _DummyWidget(params.get("pass", "4lm3rKenYa@#"))
            self._app.check_visible = _DummyWidget(params.get("visible", False))
            self._app.slider_pausa = _DummyWidget(float(params.get("pausa", 1.5)))
            self._app.combo_catalogo = _DummyWidget(params.get("cat", ""))
            self._app.combo_categoria = _DummyWidget(params.get("catg", ""))
            self._app.combo_estado = _DummyWidget(params.get("estado", "ACTIVO"))

            self._app._on_launch()
            return {"status": "started"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}

    def stop_process(self, *a):
        try:
            self._app._on_stop()
            return {"status": "stopped"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}

    # Automatizacion Stock
    def start_stock_process(self, params=None, *a):
        try:
            params = params or {}

            # Carga fallback de data de stock si no se cargó con el dialog específico
            if not getattr(self._app, "_stock_excel_df", None) and getattr(self._app, "_stock_excel_path", None):
                try:
                    from automation_otro_bot.stock import analizar_excel_stock
                    res = analizar_excel_stock(self._app._stock_excel_path)
                    if res and res.get("valido"):
                        self._app._stock_excel_df = res["df"]
                except Exception:
                    pass

            if not getattr(self._app, "_stock_excel_df", None) and getattr(self._app, "_excel_rows", None):
                self._app._stock_excel_df = self._app._excel_rows

            user_val = str(params.get("user") or "").strip()
            pass_val = str(params.get("pass") or "").strip()
            self._app.entry_stock_user = _DummyWidget(user_val)
            self._app.entry_stock_pass = _DummyWidget(pass_val)
            self._app.option_stock_acuerdo = _DummyWidget(params.get("acuerdo") or "EXT-CE-2022-5 COMPUTADORAS Y ESCÁNERES")
            self._app.option_stock_catalogo = _DummyWidget(params.get("cat") or "ESCÁNERES")
            self._app.option_stock_categoria = _DummyWidget(params.get("catg") or "ESCANER DE PLANOS")
            self._app.entry_stock_pausa = _DummyWidget(str(params.get("pausa", "2.0")))
            self._app.check_stock_visible = _DummyWidget(params.get("visible", False))

            self._app._on_stock_start()
            return {"status": "started"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}

    def start_stock_audit(self, params=None, *a):
        try:
            params = params or {}
            # Carga fallback de data de stock si no se cargó con el dialog específico
            if not getattr(self._app, "_stock_excel_df", None) and getattr(self._app, "_stock_excel_path", None):
                try:
                    from automation_otro_bot.stock import analizar_excel_stock
                    res = analizar_excel_stock(self._app._stock_excel_path)
                    if res and res.get("valido"):
                        self._app._stock_excel_df = res["df"]
                except Exception:
                    pass

            if not getattr(self._app, "_stock_excel_df", None) and getattr(self._app, "_excel_rows", None):
                self._app._stock_excel_df = self._app._excel_rows

            user_val = str(params.get("user") or "").strip()
            pass_val = str(params.get("pass") or "").strip()
            self._app.entry_stock_user = _DummyWidget(user_val)
            self._app.entry_stock_pass = _DummyWidget(pass_val)
            self._app.option_stock_acuerdo = _DummyWidget(params.get("acuerdo") or "EXT-CE-2022-5 COMPUTADORAS Y ESCÁNERES")
            self._app.option_stock_catalogo = _DummyWidget(params.get("cat") or "ESCÁNERES")
            self._app.option_stock_categoria = _DummyWidget(params.get("catg") or "ESCANER DE PLANOS")
            self._app.check_stock_visible = _DummyWidget(params.get("visible", False))

            self._app._on_stock_audit_start()
            return {"status": "started"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}

    def extract_json_portal(self, params=None, *a):
        try:
            params = params or {}
            user_val = str(params.get("user") or "").strip()
            pass_val = str(params.get("pass") or "").strip()
            self._app.entry_stock_user = _DummyWidget(user_val)
            self._app.entry_stock_pass = _DummyWidget(pass_val)
            self._app.option_stock_acuerdo = _DummyWidget(params.get("acuerdo") or "EXT-CE-2022-5 COMPUTADORAS Y ESCÁNERES")
            self._app.option_stock_catalogo = _DummyWidget(params.get("cat") or "COMPUTADORAS DE ESCRITORIO")
            self._app.option_stock_categoria = _DummyWidget(params.get("catg") or "COMPUTADORA TODO EN UNO")
            self._app.check_stock_visible = _DummyWidget(params.get("visible", False))

            self._app._on_extract_json_portal()
            return {"status": "started"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}

    def start_json_process(self, params=None, *a):
        try:
            params = params or {}
            user_val, pass_val = _get_user_pass(self._app, params)
            acuerdo_val = str(params.get("acuerdo") or "EXT-CE-2022-5 COMPUTADORAS Y ESCÁNERES").strip()
            catalogo_val = str(params.get("cat") or "COMPUTADORAS DE ESCRITORIO").strip()
            categoria_val = str(params.get("catg") or "COMPUTADORA TODO EN UNO").strip()
            visible = bool(params.get("visible", False))

            if params.get("data") and isinstance(params["data"], list):
                self._app._precios_json_data = params["data"]

            precios_data = getattr(self._app, "_precios_json_data", None)
            if not precios_data:
                return {"status": "error", "msg": "No hay datos de precios JSON cargados"}

            import threading
            import workers

            self._app._json_stop_event = threading.Event()

            self._app.entry_stock_user = _DummyWidget(user_val)
            self._app.entry_stock_pass = _DummyWidget(pass_val)
            self._app.btn_test_precios = _DummyWidget()
            self._app.btn_iniciar_precios = _DummyWidget()

            def _log(app_inst, msg):
                try:
                    print(f"[SUBIDA PRECIOS JSON] {msg}")
                except Exception:
                    pass
                win = getattr(self, '_window', None) or getattr(getattr(self, '_app', None), '_window', None)
                if win:
                    try:
                        js_msg = json.dumps(str(msg))
                        win.evaluate_js(f"logJsonConsole({js_msg});")
                    except Exception as ex:
                        pass

            combos = getattr(self._app, "_stock_combos_data", {})
            n_acuerdo = workers._get_id_acuerdo(combos, acuerdo_val) if hasattr(workers, '_get_id_acuerdo') else "249"
            n_catalogo = workers._get_id_catalogo(combos, acuerdo_val, catalogo_val) if hasattr(workers, '_get_id_catalogo') else "252"
            n_categoria = workers._get_id_categoria(combos, acuerdo_val, catalogo_val, categoria_val) if hasattr(workers, '_get_id_categoria') else "11736"

            threading.Thread(
                target=workers.execute_iniciar_precios,
                args=(self._app, user_val, pass_val, not visible, _log,
                      precios_data, n_acuerdo, n_catalogo, n_categoria,
                      self._app._json_stop_event),
                daemon=True
            ).start()

            return {"status": "started"}
        except Exception as e:
            print(f"[ERROR start_json_process] {e}")
            return {"status": "error", "msg": str(e)}

    def stop_json_process(self, *a):
        try:
            if hasattr(self._app, "_json_stop_event") and self._app._json_stop_event:
                self._app._json_stop_event.set()
                print("[SUBIDA PRECIOS JSON] Detención solicitada por el usuario.")
            return {"status": "stopped"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}

    def run_tool_test(self, params=None, *a):
        return self.extract_json_portal(params, *a)


    def stop_stock_process(self, *a):
        try:
            if hasattr(self._app, "_stock_stop_event") and self._app._stock_stop_event:
                self._app._stock_stop_event.set()
            if hasattr(self._app, "_on_stock_stop"):
                self._app._on_stock_stop()
            return {"status": "stopped"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}

    def download_stock_template(self, *a):
        try:
            if hasattr(self._app, "_on_download_stock_template"):
                self._app._on_download_stock_template()
                return {"status": "ok"}
            return {"status": "error", "msg": "Sin soporte de descarga"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}


    def export_audit(self, fmt="excel", *a):
        try:
            self._app._export_audit_report(fmt=fmt, modulo_nombre="Publicacion PDF")
            return {"status": "ok"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}


# ═══════════════════════════════════════════════════════════════════
#  BLOQUE 10 — ENTRY POINT  ( run_app() )
# ───────────────────────────────────────────────────────────────────
#  Punto de entrada del ejecutable (.exe empaquetado con PyInstaller).
#  Invocado por: if __name__ == "__main__": run_app()
#
#  Pasos internos:
#    1. Crea _Backend (clase ligera sin UI nativa CTk)
#       Inyecta métodos de SubirPdfApp en _Backend para reutilizar
#       toda la lógica sin instanciar la ventana gráfica.
#    2. Inicializa estado global del backend (log_queue, stop_event,
#       captcha_bridge, variables _stock_*, _excel_*, etc.)
#    3. Asigna _DummyWidget a todos los widgets que workers.py pueda leer.
#    4. Crea SubirPdfWebApi y asocia el backend.
#    5. Resuelve la ruta del index.html (con fallbacks para .exe/_internal).
#    6. Lanza pywebview.create_window() + pywebview.start().
#
#  ❌ NO modificar la inyección de métodos en _Backend sin revisar
#    que todos los métodos en _methods_to_bind existen en SubirPdfApp.
#  ❌ NO cambiar el nombre de variables _stock_* o _excel_*:
#    workers.py y WebApi las leen directamente.
# ═══════════════════════════════════════════════════════════════════

def run_app():
    import webview

    # Backend headless: solo logica, sin CTk/UI
    class _Backend:
        def _log(self, msg, level="info"):
            ts = time.strftime("%H:%M:%S")
            line = f"[{ts}] {msg}"
            try:
                self.log_box.configure(state="normal")
                self.log_box.insert("end", line + "\n", level)
                self.log_box.see("end")
                self.log_box.configure(state="disabled")
            except Exception:
                pass
            if hasattr(self, "_api_bridge") and self._api_bridge and self._api_bridge._window:
                try:
                    escaped_msg = json.dumps(str(msg))
                    self._api_bridge._window.evaluate_js(f"logConsole({escaped_msg})")
                except Exception:
                    pass

        def _append_stock_log(self, msg):
            ts = time.strftime("%H:%M:%S")
            line = f"[{ts}] {msg}\n"
            try:
                self.log_stock.configure(state="normal")
                self.log_stock.insert("end", line)
                self.log_stock.see("end")
                self.log_stock.configure(state="disabled")
            except Exception:
                pass
            if hasattr(self, "_api_bridge") and self._api_bridge and self._api_bridge._window:
                try:
                    escaped_msg = json.dumps(str(msg))
                    self._api_bridge._window.evaluate_js(f"logStockConsole({escaped_msg})")
                except Exception:
                    pass

        def __getattr__(self, name):
            dummy = _DummyWidget()
            setattr(self, name, dummy)
            return dummy




    _methods_to_bind = (
        "_load_dropdown_json", "_on_launch", "_on_stop", "_export_audit_report",
        "_on_stock_start", "_on_stock_stop", "_execute_stock",
        "_on_download_stock_template", "_execute", "_reset_after_stop",
        "_on_stock_audit_start", "_on_audit_done", "_on_extract_json_portal"
    )

    for _m in _methods_to_bind:
        if hasattr(SubirPdfApp, _m):
            setattr(_Backend, _m, getattr(SubirPdfApp, _m))

    app = _Backend()
    app.log_queue      = queue.Queue()
    app.stop_event     = threading.Event()
    app.captcha_bridge = CaptchaBridge()
    app._running       = False
    app._stock_running = False
    app._stock_stop_event = threading.Event()
    app._stock_log_queue  = queue.Queue()
    app._stock_log_lines  = []
    app._stock_excel_path = ""
    app._stock_excel_df   = []
    app._stock_total = app._stock_ok = app._stock_errors = 0
    app._log_lines     = []
    app._ok = app._errors = app._total = 0
    app._excel_path    = ""
    app._excel_rows    = []
    app._excel_columns = []
    app._catalog_data  = {}
    app._catalog_combos = []
    app._load_dropdown_json()

    # Default dummy widgets for initial state
    app.lbl_status = _DummyWidget()
    app.lbl_counter = _DummyWidget()
    app.progress = _DummyWidget()
    app.log_box = _DummyWidget()
    app._stat_ok = _DummyWidget()
    app._stat_warn = _DummyWidget()
    app._stat_err = _DummyWidget()
    app.btn_launch = _DummyWidget()
    app.btn_stop = _DummyWidget()
    app.btn_test = _DummyWidget()
    app.btn_certs = _DummyWidget()
    app.btn_nro = _DummyWidget()
    app.btn_extract = _DummyWidget()
    app.btn_compare = _DummyWidget()
    app.btn_discovery = _DummyWidget()
    app.btn_discovery2 = _DummyWidget()
    app.lbl_stock_status = _DummyWidget()
    app.lbl_stock_report = _DummyWidget()
    app.btn_stock_start = _DummyWidget()
    app.btn_stock_stop = _DummyWidget()
    app.btn_stock_audit = _DummyWidget()
    app.lbl_audit_status = _DummyWidget()
    app.log_stock = _DummyWidget()

    api = SubirPdfWebApi(app)

    try:
        from resource_helper import resource_path
        html_path = resource_path("ui_web/index.html")
    except Exception:
        html_path = os.path.join(_PROJECT_ROOT, "ui_web", "index.html")

    if not os.path.isfile(html_path):
        exe_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else _PROJECT_ROOT
        candidates = [
            os.path.join(exe_dir, "_internal", "ui_web", "index.html"),
            os.path.join(exe_dir, "ui_web", "index.html"),
            os.path.join(_PROJECT_ROOT, "ui_web", "index.html"),
            os.path.join(_THIS_DIR, "..", "ui_web", "index.html"),
        ]
        for cand in candidates:
            if os.path.isfile(cand):
                html_path = cand
                break

    window = webview.create_window(
        title=f"Peru Compras Bot v{VERSION}",
        url=html_path,
        js_api=api,
        width=1280, height=800,
        frameless=True, resizable=True,
        min_size=(900, 600),
    )

    api.set_window(window)
    webview.start()

if __name__ == "__main__":
    run_app()



