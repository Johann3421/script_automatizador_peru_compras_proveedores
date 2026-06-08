import os
import datetime
import openpyxl
from openpyxl.styles import PatternFill
from utils.excel_parser import find_header_row

# Colores
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def write_results(source_path: str, sheet_name: str, results: list[dict]) -> str:
    wb = openpyxl.load_workbook(source_path)
    ws = wb[sheet_name]

    header_row = find_header_row(ws)
    data_start = header_row + 1

    for r in results:
        row_idx = r.get("index", 0)
        status = r.get("status", "")
        excel_row = data_start + row_idx

        if status == "ok":
            fill = GREEN
        elif status in ("no_encontrado", "sin_part_number"):
            fill = YELLOW
        elif status in ("excede", "supera", "fuera_rango"):
            fill = RED
        elif status in ("inferior", "menor", "minimo"):
            fill = BLUE
        else:
            fill = YELLOW

        for col_idx in range(1, ws.max_column + 1):
            try:
                ws.cell(row=excel_row, column=col_idx).fill = fill
            except Exception:
                pass

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(source_path)
    out_path = "%s_procesado_%s%s" % (base, ts, ext)
    wb.save(out_path)
    wb.close()
    return out_path
