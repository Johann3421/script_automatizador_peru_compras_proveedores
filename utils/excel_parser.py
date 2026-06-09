import openpyxl


def get_sheets(filepath: str) -> list[str]:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def find_header_row(ws, max_scan=15) -> int:
    best_row = 1
    best_texts = 0

    for row_idx in range(1, min(ws.max_row + 1, max_scan + 1)):
        texts = 0
        numbers = 0
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            if isinstance(v, str) and len(v.strip()) > 1:
                texts += 1
            elif isinstance(v, (int, float)):
                numbers += 1

        if texts >= 2 and texts >= numbers and texts > best_texts:
            best_texts = texts
            best_row = row_idx

    return best_row


def get_columns(filepath: str, sheet_name: str = None) -> list[str]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    header_row = find_header_row(ws)
    cols = []
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col_idx).value
        name = str(v).strip() if v is not None else "Columna %d" % col_idx
        if name in cols:
            name = "%s_%d" % (name, col_idx)
        cols.append(name)
    wb.close()
    return cols


def parse_excel(filepath: str, sheet_name: str = None,
                header_row: int = None) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    if header_row is None:
        header_row = find_header_row(ws)

    headers = []
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col_idx).value
        name = str(v).strip() if v is not None else "Columna %d" % col_idx
        if name in headers:
            name = "%s_%d" % (name, col_idx)
        headers.append(name)

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {}
        has_value = False
        for col_idx, header in enumerate(headers, start=1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                has_value = True
                if isinstance(v, float) and v == int(v):
                    v = int(v)
            row[header] = v
        if has_value:
            rows.append(row)

    wb.close()
    return rows
