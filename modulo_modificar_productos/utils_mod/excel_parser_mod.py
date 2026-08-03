"""
excel_parser_mod.py — Lector del Excel de entrada para modificación de productos.

Columnas esperadas (auto-detectadas por nombre):
  - N° de Parte / Código / Part Number  → clave de búsqueda
  - PDF / Archivo / Ficha               → ruta absoluta o relativa al PDF a subir
  - Certificaciones / Certs             → texto de certificaciones (opcional)

Retorna una lista de dicts con las claves normalizadas.
"""
import os
from typing import Optional

try:
    import openpyxl
except ImportError as e:
    raise ImportError("openpyxl es requerido: pip install openpyxl") from e


# Alias de nombres de columna reconocidos (insensible a mayúsculas/acentos)
_PARTE_ALIASES = [
    "n° de parte", "n° parte", "num parte", "numero parte",
    "código", "codigo", "cod", "part number", "partnumber",
    "código único", "codigo unico",
]
_PDF_ALIASES = [
    "pdf", "archivo", "ficha", "ruta pdf", "ruta_pdf",
    "ficha técnica", "ficha tecnica", "documento",
]
_CERT_ALIASES = [
    "certificaciones", "certificacion", "certs", "cert",
    "certif", "certificado",
]


def _normalize(text: str) -> str:
    """Normaliza texto para comparación: minúsculas, sin acentos."""
    text = str(text).lower().strip()
    replacements = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}
    for a, b in replacements.items():
        text = text.replace(a, b)
    return text


def _match_col(header: str, aliases: list[str]) -> bool:
    norm = _normalize(header)
    return any(alias in norm or norm in alias for alias in aliases)


def get_sheets(path: str) -> list[str]:
    """Retorna la lista de hojas del Excel."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def detect_columns(path: str, sheet: str) -> dict:
    """
    Detecta automáticamente las columnas de parte, PDF y certificaciones.
    Retorna un dict con las claves 'parte_col', 'pdf_col', 'cert_col' (puede ser None).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    parte_col = pdf_col = cert_col = None

    for row in ws.iter_rows(max_row=10):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value)
            if parte_col is None and _match_col(val, _PARTE_ALIASES):
                parte_col = val
            if pdf_col is None and _match_col(val, _PDF_ALIASES):
                pdf_col = val
            if cert_col is None and _match_col(val, _CERT_ALIASES):
                cert_col = val

    wb.close()
    return {"parte_col": parte_col, "pdf_col": pdf_col, "cert_col": cert_col}


def parse_excel(
    path: str,
    sheet: str,
    parte_col: str,
    pdf_col: Optional[str] = None,
    cert_col: Optional[str] = None,
) -> list[dict]:
    """
    Lee el Excel y retorna una lista de dicts con:
      {
        "parte":   str,         # N° de Parte
        "pdf":     str | None,  # Ruta al PDF (o None)
        "certs":   str | None,  # Texto de certificaciones (o None)
        "_row_idx": int,        # Índice de fila (0-based, desde header+1)
      }
    Filas con parte vacía son ignoradas.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw:
        return []

    # Encontrar header row — usar la primera fila con contenido
    header_row_idx = 0
    headers = []
    for i, row in enumerate(rows_raw[:10]):
        vals = [str(v) if v is not None else "" for v in row]
        if any(v.strip() for v in vals):
            header_row_idx = i
            headers = vals
            break
    else:
        return []

    # Mapear columna de parte: match exacto contra parte_col o primer columna
    col_idx: dict[str, int] = {}
    parte_found = False
    if parte_col:
        for j, h in enumerate(headers):
            if h.strip() == parte_col.strip():
                col_idx["parte"] = j
                parte_found = True
                break
    if not parte_found:
        # Fallback: buscar por alias
        for j, h in enumerate(headers):
            if h and _match_col(h.strip(), _PARTE_ALIASES):
                col_idx["parte"] = j
                parte_found = True
                break
    if not parte_found:
        # Fallback extremo: columna 0
        col_idx["parte"] = 0

    # Mapear columnas opcionales
    for col_name, aliases in [("pdf", _PDF_ALIASES), ("certs", _CERT_ALIASES)]:
        for j, h in enumerate(headers):
            if h and _match_col(h.strip(), aliases):
                col_idx[col_name] = j
                break

    # Parsear filas de datos
    results = []
    for i, row in enumerate(rows_raw[header_row_idx + 1:]):
        if all(v is None for v in row):
            continue

        parte_val = ""
        pdf_val = None
        certs_val = None

        idx = col_idx.get("parte", 0)
        if idx < len(row):
            v = row[idx]
            parte_val = str(v).strip() if v is not None else ""

        if not parte_val:
            continue

        if col_idx.get("pdf") is not None and col_idx["pdf"] < len(row):
            v = row[col_idx["pdf"]]
            if v is not None and str(v).strip():
                pdf_val = str(v).strip()

        if col_idx.get("certs") is not None and col_idx["certs"] < len(row):
            v = row[col_idx["certs"]]
            if v is not None and str(v).strip():
                certs_val = str(v).strip()

        results.append({
            "parte": parte_val,
            "pdf": pdf_val,
            "certs": certs_val,
            "_row_idx": i,
        })

    return results
