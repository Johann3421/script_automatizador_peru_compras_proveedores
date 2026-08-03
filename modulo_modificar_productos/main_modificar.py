import sys, os, queue, threading, time, json
from io import BytesIO
from tkinter import filedialog

VERSION = "1.0"

# ── Paths ─────────────────────────────────────────────────────────
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.abspath(os.path.join(_THIS_DIR, ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)

from automation.browser import init_browser, close_browser
from automation.login import do_login
from utils.logger import LogWriter

# Imports locales del módulo (renombrados para evitar conflicto con el proyecto raíz)
from utils_mod.excel_parser_mod import get_sheets, detect_columns, parse_excel
from utils_mod.logger_mod import LogWriter as LocalLogWriter

try:
    import customtkinter as ctk
    from PIL import Image
except ImportError:
    print("Error: instala customtkinter y Pillow: pip install customtkinter pillow")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════
#  BRIDGE — Comunicación UI ↔ Thread (CAPTCHA manual)
# ═══════════════════════════════════════════════════════════════════

class CaptchaBridge:
    def __init__(self):
        self.lock = threading.Lock()
        self.event = threading.Event()
        self.image_bytes = None
        self.user_code = ""
        self.stop_event = None

    def request(self, img):
        with self.lock:
            self.image_bytes = img
            self.user_code = ""
            self.event.clear()
        while True:
            if self.stop_event and self.stop_event.is_set():
                self.event.set()
                return ""
            if self.event.wait(timeout=0.5):
                break
        with self.lock:
            return self.user_code

    def respond(self, code):
        with self.lock:
            self.user_code = code
            self.image_bytes = None
        self.event.set()


# ═══════════════════════════════════════════════════════════════════
#  APP
# ═══════════════════════════════════════════════════════════════════

class ModificarProductosApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"Peru Compras — Modificar Productos v{VERSION}")
        self.geometry("960x760")
        self.minsize(820, 640)
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.log_queue = queue.Queue()
        self.stop_event = threading.Event()
        self.captcha_bridge = CaptchaBridge()
        self._running = False
        self._log_lines = []
        self._ok = 0
        self._errors = 0
        self._total = 0

        self._excel_path = ""
        self._excel_rows = []
        self._excel_columns = []

        self._catalog_data = {}
        self._catalog_combos = []
        self._load_dropdown_json()

        self._build_ui()
        self.poll_queue()

    def _load_dropdown_json(self):
        json_path = os.path.join(_THIS_DIR, "combinaciones_computadoras.json")
        if not os.path.isfile(json_path):
            json_path = os.path.join(_THIS_DIR, "dropdown_options_modificar.json")
        if not os.path.isfile(json_path):
            return
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                self._catalog_data = json.load(f)
        except Exception:
            pass

    # ── Build UI ──────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=0)

        # ── HEADER ──
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, padx=20, pady=(16, 4), sticky="ew")
        ctk.CTkLabel(
            header,
            text=f"Modificar Productos v{VERSION}",
            font=ctk.CTkFont(size=22, weight="bold"),
        ).pack(side="left")
        ctk.CTkLabel(
            header,
            text="Subida de PDF + Certificaciones",
            font=ctk.CTkFont(size=12),
            text_color="gray60",
        ).pack(side="left", padx=12)

        # ── BADGE de módulo ──
        ctk.CTkLabel(
            header,
            text="  MÓDULO INDEPENDIENTE  ",
            font=ctk.CTkFont(size=10, weight="bold"),
            fg_color="#1a4a1a",
            text_color="#5dade2",
            corner_radius=6,
        ).pack(side="right")

        # ── MAIN CONTENT ──
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.grid(row=1, column=0, padx=20, pady=(0, 8), sticky="nsew")
        content.grid_columnconfigure(0, weight=1, minsize=340)
        content.grid_columnconfigure(1, weight=1, minsize=340)
        content.grid_rowconfigure(0, weight=1)

        # LEFT COLUMN
        left = ctk.CTkScrollableFrame(content, fg_color="transparent")
        left.grid(row=0, column=0, padx=(0, 6), sticky="nsew")
        left.grid_columnconfigure(0, weight=1)

        self._build_credentials_section(left)
        self._build_excel_section(left)
        self._build_catalog_section(left)
        self._build_opciones_section(left)

        # RIGHT COLUMN
        right = ctk.CTkFrame(content)
        right.grid(row=0, column=1, padx=(6, 0), sticky="nsew")
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=1)

        self._build_execution_section(right)

        # ── FOOTER ──
        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.grid(row=2, column=0, padx=20, pady=(4, 12), sticky="ew")
        footer.grid_columnconfigure(0, weight=1)

        self.btn_test = ctk.CTkButton(
            footer,
            text="🔬  Test Flow",
            height=42,
            fg_color="#2c3e50",
            hover_color="#34495e",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._on_test,
        )
        self.btn_test.pack(side="right", padx=(8, 0))

        self.btn_launch = ctk.CTkButton(
            footer,
            text="▶  Iniciar Procesamiento",
            height=42,
            font=ctk.CTkFont(size=15, weight="bold"),
            command=self._on_launch,
        )
        self.btn_launch.pack(side="right", padx=(8, 0))

        self.btn_stop = ctk.CTkButton(
            footer,
            text="■  Detener",
            height=42,
            fg_color="#c0392b",
            hover_color="#96281b",
            state="disabled",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._on_stop,
        )
        self.btn_stop.pack(side="right", padx=8)

    # ── Credentials Section ───────────────────────────────────────

    def _build_credentials_section(self, parent):
        self._section_label(parent, "Credenciales de Peru Compras", 0)

        frame = ctk.CTkFrame(parent, fg_color="gray10")
        frame.grid(row=1, column=0, padx=12, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        # Usuario
        ctk.CTkLabel(
            frame, text="Usuario", font=ctk.CTkFont(size=11),
            text_color="gray60", anchor="w",
        ).grid(row=0, column=0, padx=12, pady=(10, 1), sticky="w")
        self.entry_user = ctk.CTkEntry(frame, placeholder_text="Usuario o RUC", height=34)
        self.entry_user.grid(row=1, column=0, padx=12, pady=(0, 8), sticky="ew")
        self.entry_user.insert(0, "almerco.03")

        # Password
        pass_frame = ctk.CTkFrame(frame, fg_color="transparent")
        pass_frame.grid(row=2, column=0, padx=12, pady=(0, 2), sticky="ew")
        pass_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            pass_frame, text="Contraseña", font=ctk.CTkFont(size=11),
            text_color="gray60", anchor="w",
        ).grid(row=0, column=0, columnspan=2, padx=0, pady=(0, 1), sticky="w")

        self._pw_visible = False
        self.entry_pass = ctk.CTkEntry(
            pass_frame, placeholder_text="Contraseña", show="*", height=34
        )
        self.entry_pass.grid(row=1, column=0, pady=(0, 8), sticky="ew")
        self.entry_pass.insert(0, "4lm3rKenYa@#")
        self.btn_eye = ctk.CTkButton(
            pass_frame, text="▸", width=34, height=34,
            font=ctk.CTkFont(size=14), command=self._toggle_password,
        )
        self.btn_eye.grid(row=1, column=1, padx=(6, 0), pady=(0, 8))

        # Mostrar navegador
        self.check_visible = ctk.CTkCheckBox(
            frame, text="Mostrar navegador en pantalla",
            font=ctk.CTkFont(size=12),
        )
        self.check_visible.grid(row=3, column=0, padx=12, pady=(2, 10), sticky="w")

    def _toggle_password(self):
        self._pw_visible = not self._pw_visible
        self.entry_pass.configure(show="" if self._pw_visible else "*")
        self.btn_eye.configure(text="◂" if self._pw_visible else "▸")

    # ── Excel Section ─────────────────────────────────────────────

    def _build_excel_section(self, parent):
        self._section_label(parent, "Archivo Excel", 2)

        frame = ctk.CTkFrame(parent, fg_color="gray10")
        frame.grid(row=3, column=0, padx=12, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        # File picker
        file_row = ctk.CTkFrame(frame, fg_color="transparent")
        file_row.grid(row=0, column=0, padx=12, pady=(10, 4), sticky="ew")
        file_row.grid_columnconfigure(0, weight=1)
        self.btn_file = ctk.CTkButton(
            file_row, text="Seleccionar archivo .xlsx",
            height=32, font=ctk.CTkFont(size=12),
            command=self._pick_excel,
        )
        self.btn_file.pack(side="left")
        self.lbl_file = ctk.CTkLabel(
            file_row, text="Sin archivo", text_color="gray60",
            font=ctk.CTkFont(size=11),
        )
        self.lbl_file.pack(side="left", padx=10)

        # Sheet selector
        ctk.CTkLabel(
            frame, text="Pestaña del Excel", font=ctk.CTkFont(size=11),
            text_color="gray60", anchor="w",
        ).grid(row=1, column=0, padx=12, pady=(4, 1), sticky="w")
        self.combo_sheet = ctk.CTkComboBox(
            frame, values=["Primero cargá un Excel"],
            state="disabled", height=32,
            command=self._on_sheet_changed,
        )
        self.combo_sheet.grid(row=2, column=0, padx=12, pady=(0, 6), sticky="ew")

        # Column mapping — solo N° de Parte
        map_frame = ctk.CTkFrame(frame, fg_color="transparent")
        map_frame.grid(row=3, column=0, padx=12, pady=(4, 4), sticky="ew")
        map_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            map_frame, text="Columna N° de Parte", font=ctk.CTkFont(size=11),
            text_color="gray60", anchor="w",
        ).grid(row=0, column=0, padx=(0, 4), pady=(0, 1), sticky="w")

        self.combo_parte = ctk.CTkComboBox(
            map_frame, values=["--"], state="disabled", height=32,
        )
        self.combo_parte.grid(row=1, column=0, padx=(0, 4), sticky="ew")

        # Info
        self.lbl_excel_info = ctk.CTkLabel(
            frame, text="", font=ctk.CTkFont(size=11),
            text_color="gray60", anchor="w",
        )
        self.lbl_excel_info.grid(row=4, column=0, padx=12, pady=(2, 10), sticky="w")

    def _pick_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not path:
            return
        try:
            sheets = get_sheets(path)
            if not sheets:
                self.lbl_file.configure(text="Sin pestañas en el Excel", text_color="#e74c3c")
                return

            self._excel_path = path
            name = os.path.basename(path)
            self.lbl_file.configure(text=name, text_color="#5dade2")
            self.combo_sheet.configure(values=sheets, state="readonly")
            self.combo_sheet.set(sheets[0])
            self._on_sheet_changed(sheets[0])
        except Exception as e:
            self.lbl_file.configure(text=f"Error: {e}", text_color="#e74c3c")

    def _on_sheet_changed(self, choice):
        if not self._excel_path or not choice:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._excel_path, read_only=True, data_only=True)
            ws = wb[choice]

            # Leer todos los headers de la fila 1 (primeros 20 para cubrir)
            all_headers = []
            # Leer fila 1 como header principal
            row1 = [cell for cell in ws.iter_rows(min_row=1, max_row=1, values_only=True)]
            if row1 and row1[0]:
                for v in row1[0]:
                    if v is not None and str(v).strip():
                        h = str(v).strip()
                        if h not in all_headers and h.lower() not in ("none", ""):
                            all_headers.append(h)
            # Si fila 1 no tiene nada, buscar en filas 2-10
            if not all_headers:
                for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
                    for v in (row or []):
                        if v is not None and str(v).strip():
                            h = str(v).strip()
                            if h not in all_headers and not h.replace(".","").replace(",","").isdigit():
                                all_headers.append(h)
                    if all_headers:
                        break

            # Si no hay headers, usar columna por defecto
            if not all_headers:
                all_headers = ["Columna A"]

            row_count = ws.max_row or 0
            wb.close()

            # Detectar columna de parte (match exacto o primer columna)
            parte_col_detect = None
            aliases = ["parte", "part number", "n°", "codigo", "cod", "id", "sku", "item"]
            for h in all_headers:
                hl = h.lower().replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("ñ","n")
                for a in aliases:
                    if a in hl or hl in a:
                        parte_col_detect = h
                        break
                if parte_col_detect:
                    break
            if not parte_col_detect:
                parte_col_detect = all_headers[0]

            # Parse simple
            rows_data = parse_excel(
                self._excel_path, choice,
                parte_col=parte_col_detect,
            )
            self._excel_rows = rows_data

            opts = ["-- Seleccionar --"] + all_headers
            self.combo_parte.configure(values=opts, state="readonly")
            self.combo_parte.set(parte_col_detect if parte_col_detect in all_headers else opts[0])

            self.lbl_excel_info.configure(
                text=f"{len(rows_data)} filas · {len(all_headers)} cols · Pestaña: {choice}",
                text_color="gray60",
            )
        except Exception as e:
            self.lbl_excel_info.configure(text=f"Error: {e}", text_color="#e74c3c")

    # ── Catalog Section (Cascada Catálogo → Categoría → Estado) ─────

    def _build_catalog_section(self, parent):
        row_offset = 4
        self._section_label(parent, "Configuración del Catálogo (EXT-CE-2022-5)", row_offset)

        frame = ctk.CTkFrame(parent, fg_color="gray10")
        frame.grid(row=row_offset+1, column=0, padx=12, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        comb_data = self._catalog_data.get("combinaciones", [])
        if not comb_data:
            ctk.CTkLabel(
                frame,
                text="combinaciones_computadoras.json no encontrado.\nEjecutá extract_combinaciones.py primero",
                text_color="#f39c12", font=ctk.CTkFont(size=11), wraplength=280,
                justify="left",
            ).grid(row=0, column=0, padx=12, pady=10)
            return

        # Catálogo
        ctk.CTkLabel(frame, text="Catálogo Electrónico", font=ctk.CTkFont(size=11),
                     text_color="gray60", anchor="w").grid(
            row=0, column=0, padx=12, pady=(8, 1), sticky="w")
        self.combo_catalogo = ctk.CTkComboBox(
            frame, values=self._opts_texts(comb_data), state="readonly", height=32,
            command=self._on_catalogo_changed,
        )
        self.combo_catalogo.grid(row=1, column=0, padx=12, pady=(0, 6), sticky="ew")
        if comb_data:
            self.combo_catalogo.set(self._opts_texts(comb_data)[0])

        # Categoría
        ctk.CTkLabel(frame, text="Categoría", font=ctk.CTkFont(size=11),
                     text_color="gray60", anchor="w").grid(
            row=2, column=0, padx=12, pady=(4, 1), sticky="w")
        self.combo_categoria = ctk.CTkComboBox(
            frame, values=["Seleccioná un Catálogo"], state="readonly", height=32,
            command=self._on_categoria_changed,
        )
        self.combo_categoria.grid(row=3, column=0, padx=12, pady=(0, 6), sticky="ew")

        # Estado
        ctk.CTkLabel(frame, text="Estado", font=ctk.CTkFont(size=11),
                     text_color="gray60", anchor="w").grid(
            row=4, column=0, padx=12, pady=(4, 1), sticky="w")
        self.combo_estado = ctk.CTkComboBox(
            frame, values=["Seleccioná una Categoría"], state="readonly", height=32,
        )
        self.combo_estado.grid(row=5, column=0, padx=12, pady=(0, 8), sticky="ew")

        # Trigger inicial
        self._on_catalogo_changed(self._opts_texts(comb_data)[0])

    def _opts_texts(self, data):
        return [f"{o['value']} - {o['text'][:60]}" for o in data] if data else []

    def _find_children(self, combo_text):
        val = combo_text.split(" - ")[0].strip() if " - " in combo_text else ""
        comb_data = self._catalog_data.get("combinaciones", [])
        for node in comb_data:
            if node["value"] == val:
                return node.get("children", [])
        return []

    def _on_catalogo_changed(self, choice):
        children = self._find_children(choice)
        texts = self._opts_texts(children)
        if texts:
            self.combo_categoria.configure(values=texts)
            self.combo_categoria.set(texts[0])
            self._on_categoria_changed(texts[0])
        else:
            self.combo_categoria.configure(values=["Sin categorías"])
            self.combo_estado.configure(values=["Sin estados"])

    def _on_categoria_changed(self, choice):
        # Buscar en el nodo actual de catalogo
        cat_val = self.combo_catalogo.get().split(" - ")[0].strip() if " - " in self.combo_catalogo.get() else ""
        comb_data = self._catalog_data.get("combinaciones", [])
        for node in comb_data:
            if node["value"] == cat_val:
                for child in node.get("children", []):
                    val = choice.split(" - ")[0].strip() if " - " in choice else ""
                    if child["value"] == val:
                        estados = child.get("children", [])
                        texts = self._opts_texts(estados)
                        if texts:
                            self.combo_estado.configure(values=texts)
                            self.combo_estado.set(texts[0])
                        else:
                            self.combo_estado.configure(values=["Sin estados"])
                        return
        self.combo_estado.configure(values=["Sin estados"])

    # ── Opciones Section ──────────────────────────────────────────

    def _build_opciones_section(self, parent):
        row_offset = 6
        self._section_label(parent, "Opciones de Procesamiento", row_offset)

        frame = ctk.CTkFrame(parent, fg_color="gray10")
        frame.grid(row=row_offset+1, column=0, padx=12, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        # Pausa entre productos
        ctk.CTkLabel(
            frame, text="Pausa entre productos (segundos)",
            font=ctk.CTkFont(size=11), text_color="gray60", anchor="w",
        ).grid(row=0, column=0, padx=12, pady=(10, 1), sticky="w")
        self.slider_pausa = ctk.CTkSlider(frame, from_=0.5, to=5.0, number_of_steps=9)
        self.slider_pausa.grid(row=1, column=0, padx=12, pady=(0, 2), sticky="ew")
        self.slider_pausa.set(1.5)
        self.lbl_pausa = ctk.CTkLabel(
            frame, text="1.5 s", font=ctk.CTkFont(size=11), text_color="gray60",
        )
        self.lbl_pausa.grid(row=2, column=0, padx=12, pady=(0, 6), sticky="w")
        self.slider_pausa.configure(command=lambda v: self.lbl_pausa.configure(text=f"{v:.1f} s"))

        # Separador
        ctk.CTkFrame(frame, height=1, fg_color="gray25").grid(
            row=3, column=0, padx=12, pady=4, sticky="ew"
        )

        # Info de modo
        ctk.CTkLabel(
            frame,
            text="ℹ  Este módulo usa Playwright para subir archivos.\n"
                 "   Los endpoints se ajustan según la sección del portal.",
            font=ctk.CTkFont(size=11),
            text_color="gray50",
            anchor="w",
            justify="left",
        ).grid(row=4, column=0, padx=12, pady=(4, 10), sticky="w")

    # ── Execution Section ─────────────────────────────────────────

    def _build_execution_section(self, parent):
        self._section_label(parent, "Ejecución", 0)

        frame = ctk.CTkFrame(parent, fg_color="gray10")
        frame.grid(row=1, column=0, padx=12, pady=(0, 4), sticky="nsew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(5, weight=1)

        # Status bar
        status_row = ctk.CTkFrame(frame, fg_color="transparent")
        status_row.grid(row=0, column=0, padx=12, pady=(8, 4), sticky="ew")
        self.lbl_status = ctk.CTkLabel(
            status_row, text="Listo para iniciar",
            font=ctk.CTkFont(size=13, weight="bold"),
        )
        self.lbl_status.pack(side="left")
        self.lbl_counter = ctk.CTkLabel(
            status_row, text="", font=ctk.CTkFont(size=12), text_color="gray60",
        )
        self.lbl_counter.pack(side="right")

        # Progress bar
        self.progress = ctk.CTkProgressBar(frame, height=8)
        self.progress.grid(row=2, column=0, padx=12, pady=(0, 4), sticky="ew")
        self.progress.set(0)

        # Resumen de resultados (mini-stats)
        stats_row = ctk.CTkFrame(frame, fg_color="transparent")
        stats_row.grid(row=3, column=0, padx=12, pady=(0, 4), sticky="ew")
        stats_row.grid_columnconfigure((0, 1, 2), weight=1)

        self._stat_ok = self._make_stat(stats_row, "✓  OK", "#2ecc71", 0)
        self._stat_warn = self._make_stat(stats_row, "⚠  No encontrado", "#f39c12", 1)
        self._stat_err = self._make_stat(stats_row, "✗  Error", "#e74c3c", 2)

        # Captcha panel (oculto por defecto)
        self._build_captcha_panel(frame)

        # Log
        self.log_box = ctk.CTkTextbox(frame, wrap="word", font=ctk.CTkFont(family="Courier New", size=11))
        self.log_box.grid(row=5, column=0, padx=12, pady=(2, 8), sticky="nsew")
        self.log_box.configure(state="disabled")
        self.log_box.tag_config("ok",       foreground="#f1c40f")
        self.log_box.tag_config("error",    foreground="#e74c3c")
        self.log_box.tag_config("info",     foreground="#95a5a6")
        self.log_box.tag_config("warn",     foreground="#f39c12")
        self.log_box.tag_config("done",     foreground="#5dade2")
        self.log_box.tag_config("complete", foreground="#2ecc71")
        self.log_box.tag_config("existing", foreground="#3498db")
        self.log_box.tag_config("notfound", foreground="#c0392b")

    def _make_stat(self, parent, label, color, col):
        f = ctk.CTkFrame(parent, fg_color="gray14", corner_radius=6)
        f.grid(row=0, column=col, padx=3, pady=2, sticky="ew")
        lbl_n = ctk.CTkLabel(f, text="0", font=ctk.CTkFont(size=20, weight="bold"), text_color=color)
        lbl_n.pack(pady=(6, 0))
        ctk.CTkLabel(f, text=label, font=ctk.CTkFont(size=10), text_color="gray60").pack(pady=(0, 6))
        return lbl_n

    # ── Captcha Panel ─────────────────────────────────────────────

    def _build_captcha_panel(self, parent):
        self.captcha_panel = ctk.CTkFrame(parent, fg_color="gray14")
        self.captcha_panel.grid(row=4, column=0, padx=12, pady=(0, 4), sticky="ew")
        self.captcha_panel.grid_remove()
        self.captcha_panel.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.captcha_panel,
            text="CAPTCHA — Ingresá el código",
            font=ctk.CTkFont(size=12, weight="bold"), anchor="w",
        ).grid(row=0, column=0, padx=8, pady=(6, 4), sticky="w")

        self.captcha_img_lbl = ctk.CTkLabel(self.captcha_panel, text="")
        self.captcha_img_lbl.grid(row=1, column=0, padx=8, pady=(0, 4))

        cap_row = ctk.CTkFrame(self.captcha_panel, fg_color="transparent")
        cap_row.grid(row=2, column=0, padx=8, pady=(0, 8), sticky="ew")
        cap_row.grid_columnconfigure(0, weight=1)

        self.captcha_entry = ctk.CTkEntry(cap_row, placeholder_text="Código", height=30)
        self.captcha_entry.grid(row=0, column=0, padx=(0, 6), sticky="ew")
        self.captcha_entry.bind("<Return>", lambda e: self._on_captcha_submit())

        ctk.CTkButton(
            cap_row, text="Enviar", width=60, height=30,
            command=self._on_captcha_submit,
        ).grid(row=0, column=1)

        self._captcha_img = None

    def _show_captcha(self, image_bytes):
        img = Image.open(BytesIO(image_bytes))
        img = img.resize((250, 66), Image.LANCZOS)
        self._captcha_img = ctk.CTkImage(light_image=img, dark_image=img, size=(250, 66))
        self.captcha_img_lbl.configure(image=self._captcha_img)
        self.captcha_entry.delete(0, "end")
        self.captcha_entry.focus_set()
        self.captcha_panel.grid()

    def _hide_captcha_panel(self):
        self.captcha_panel.grid_remove()

    def _on_captcha_submit(self):
        code = self.captcha_entry.get().strip()
        if code:
            self.captcha_bridge.respond(code)
            self._hide_captcha_panel()
            self._log("CAPTCHA manual: " + code, "info")

    # ── Helpers ───────────────────────────────────────────────────

    def _section_label(self, parent, text, row):
        ctk.CTkLabel(
            parent, text=text,
            font=ctk.CTkFont(size=14, weight="bold"), anchor="w",
        ).grid(row=row, column=0, padx=12, pady=(8, 4), sticky="w")

    def _log(self, msg, level="info"):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n", level)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self._log_lines.append({"level": level, "msg": msg})

    # ── Test Flow ──────────────────────────────────────────────

    def _on_test(self):
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return

        if not hasattr(self, 'combo_catalogo') or not self._catalog_data:
            self.lbl_status.configure(text="No hay datos de catálogo cargados", text_color="#e74c3c")
            return

        headless = not bool(self.check_visible.get())

        def _val(combo):
            t = combo.get()
            return t.split(" - ")[0].strip() if " - " in t else ""

        pre_selected = {
            "acuerdo": "249",
            "catalogo": _val(self.combo_catalogo),
            "categoria": _val(self.combo_categoria),
            "estado": _val(self.combo_estado),
        }

        self._running = True
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress.set(0)
        self.lbl_status.configure(text="🧪 Test Flow en marcha...", text_color="#f39c12")
        self.btn_launch.configure(state="disabled")
        self.btn_test.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()

        threading.Thread(
            target=self._execute_test,
            args=(user, pwd, headless, pre_selected),
            daemon=True,
        ).start()

    def _execute_test(self, usuario, password, headless, pre_selected):
        log = LogWriter(self.log_queue)
        stop = self.stop_event
        pw = browser = None
        try:
            log.info("🧪 Test Flow: iniciando navegador...")
            if stop.is_set(): return
            pw, browser, page = init_browser(headless=headless)
            if not headless:
                try:
                    page.set_viewport_size({"width": 1920, "height": 1080})
                except Exception:
                    pass
            log.info("🧪 Navegador listo (viewport 1920x1080)")

            if stop.is_set(): return
            ok = do_login(page, usuario, password, "", log, stop, self.captcha_bridge)
            if not ok or stop.is_set():
                if not ok: log.error("🧪 Login fallido.")
                return
            log.ok("🧪 Login exitoso")

            from automation_mod.navegacion_productos import (
                GESTION_URL, apply_dropdowns_and_search,
                buscar_por_parte, click_editar,
                subir_pdf_en_edicion, guardar_cambios,
                agregar_caracteristicas, volver_a_lista,
                process_single_product, ensure_logged_in_and_ready,
            )

            log.info("🧪 Navegando a t_CatalogoProductoMarca...")
            try:
                page.goto(GESTION_URL, wait_until="networkidle", timeout=60_000)
            except Exception as e:
                log.warn(f"🧪 Timeout en goto ({e})")

            # 1. Aplicar dropdowns y Buscar
            log.info("🧪 Aplicando dropdowns y Buscar...")
            result = apply_dropdowns_and_search(page, pre_selected, log, stop)
            if result["status"] != "ok":
                log.warn("🧪 La búsqueda con filtros no dio resultados")
                return

            if not self._excel_rows:
                log.warn("🧪 No hay filas en el Excel")
                return

            # 2. Iterar TODOS los productos del Excel
            pdf_dir = r"D:\SISTEMAS 02\Downloads\COMPUTADORAS\COMPUTADORAS"
            stats = {"ok": 0, "not_found": 0, "failed": 0, "existing": 0}
            all_results = []

            for idx, row in enumerate(self._excel_rows):
                if stop.is_set():
                    break
                parte = row["parte"]
                ruta_pdf = os.path.join(pdf_dir, f"{parte}.pdf")

                log.info(f"🧪 [{idx+1}/{len(self._excel_rows)}] Procesando: {parte}")

                # Verificar sesión antes de procesar
                if not ensure_logged_in_and_ready(
                    page, usuario, password, pre_selected, log, stop, self.captcha_bridge
                ):
                    log.error(f"🧪 [{idx+1}] No se pudo restablecer sesión, deteniendo")
                    all_results.append({"index": idx, "parte": parte, "status": "session_lost"})
                    break

                result = process_single_product(
                    page, parte, ruta_pdf, log, stop,
                    pre_selected=None,
                )
                status = result["status"]
                all_results.append({"index": idx, "parte": parte, "status": status})

                if status == "ok":
                    stats["ok"] += 1
                    self.log_queue.put({"type": "log", "msg": f"🧪 [{idx+1}/{len(self._excel_rows)}] ✓ {parte} — PDF + ISOs OK", "level": "complete"})
                elif status == "not_found":
                    stats["not_found"] += 1
                    self.log_queue.put({"type": "log", "msg": f"🧪 [{idx+1}/{len(self._excel_rows)}] ? {parte} — No encontrado", "level": "notfound"})
                elif status == "certs_already_exist":
                    stats["existing"] += 1
                    self.log_queue.put({"type": "log", "msg": f"🧪 [{idx+1}/{len(self._excel_rows)}] ~ {parte} — PDF OK, ISOs ya existían", "level": "existing"})
                else:
                    stats["failed"] += 1
                    self.log_queue.put({"type": "log", "msg": f"🧪 [{idx+1}/{len(self._excel_rows)}] ✗ {parte} — Falló: {status}", "level": "error"})

            # Resumen final
            log.info(f"🧪 {'='*40}")
            log.info(f"🧪 RESUMEN: {stats['ok']} OK | {stats['not_found']} No encontrados | "
                     f"{stats['existing']} Ya existían | {stats['failed']} Fallos")

            # Colorear Excel
            if all_results and self._excel_path:
                from utils_mod.excel_writer_mod import write_colored_results
                try:
                    sheet = self.combo_sheet.get() if hasattr(self, 'combo_sheet') else "Hoja1"
                    out = write_colored_results(self._excel_path, sheet, all_results)
                    log.ok(f"🧪 Excel coloreado guardado: {os.path.basename(out)}")
                except Exception as e:
                    log.warn(f"🧪 No se pudo colorear el Excel: {e}")

            log.info(f"🧪 {'='*40}")
            if stats["ok"] > 0:
                log.ok(f"🧪 Procesamiento completado exitosamente")

        except Exception as e:
            log.error(f"🧪 Error en test: {e}")
            import traceback
            log.error(traceback.format_exc())
        finally:
            if browser and pw:
                try:
                    close_browser(pw, browser)
                    log.info("🧪 Navegador cerrado")
                except Exception:
                    pass
            self.log_queue.put({"type": "done", "ok": 0, "errors": 0})
            self.after(0, lambda: self.btn_test.configure(state="normal"))

    # ── Launch ─────────────────────────────────────────────────

    def _on_launch(self):
        user = self.entry_user.get().strip()
        pwd  = self.entry_pass.get().strip()
        if not user or not pwd:
            self.lbl_status.configure(text="Ingresá usuario y contraseña", text_color="#e74c3c")
            return

        parte_col = self.combo_parte.get()
        pdf_col   = None
        cert_col  = None
        sheet     = self.combo_sheet.get()

        if parte_col in ("-- Seleccionar --", "--") or not self._excel_path:
            self.lbl_status.configure(
                text="Cargá un Excel y elegí al menos la columna N° de Parte",
                text_color="#e74c3c",
            )
            return

        # Re-parsear con las columnas seleccionadas
        try:
            rows = parse_excel(
                path=self._excel_path,
                sheet=sheet,
                parte_col=parte_col,
                pdf_col=None,
                cert_col=None,
            )
        except Exception as e:
            self.lbl_status.configure(text=f"Error leyendo Excel: {e}", text_color="#e74c3c")
            return

        if not rows:
            self.lbl_status.configure(text="El Excel no tiene filas de datos", text_color="#e74c3c")
            return

        self._running = True
        self._ok = 0
        self._errors = 0
        self._total = len(rows)
        self._log_lines.clear()
        self.stop_event.clear()
        self.captcha_bridge.stop_event = self.stop_event

        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

        self.progress.set(0)
        self._stat_ok.configure(text="0")
        self._stat_warn.configure(text="0")
        self._stat_err.configure(text="0")
        self.lbl_status.configure(text="Iniciando...", text_color="white")
        self.btn_launch.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._hide_captcha_panel()

        headless = not bool(self.check_visible.get())
        pausa    = self.slider_pausa.get()

        # Valores de los dropdowns en cascada
        def _val(combo):
            t = combo.get()
            return t.split(" - ")[0].strip() if " - " in t else ""

        acuerdo_data = self._catalog_data.get("acuerdo", {})
        pre_selected = {
            "acuerdo": acuerdo_data.get("value", "249"),
            "catalogo": _val(self.combo_catalogo),
            "categoria": _val(self.combo_categoria),
            "estado": _val(self.combo_estado),
        }

        threading.Thread(
            target=self._execute,
            args=(user, pwd, headless, rows, pausa, pre_selected),
            daemon=True,
        ).start()

    def _execute(self, usuario, password, headless, rows, pausa, pre_selected=None):
        log = LogWriter(self.log_queue)
        stop = self.stop_event
        pw = browser = None

        try:
            log.info("Iniciando navegador...")
            if stop.is_set():
                return
            pw, browser, page = init_browser(headless=headless)
            if not headless:
                try:
                    page.set_viewport_size({"width": 1920, "height": 1080})
                except Exception:
                    pass
            log.info("Navegador listo")

            if stop.is_set():
                return
            ok = do_login(page, usuario, password, "", log, stop, self.captcha_bridge)
            if not ok or stop.is_set():
                if not ok:
                    log.error("Login fallido.")
                return

            log.ok("Login exitoso. Iniciando procesamiento masivo...")

            from automation_mod.bulk_modificar import run_bulk_modificar, URL_MANAGEMENT
            import time as _time

            log.info("Navegando al catálogo de fichas...")
            try:
                page.goto(URL_MANAGEMENT, wait_until="networkidle", timeout=60_000)
                _time.sleep(2)
            except Exception:
                pass

            self._total = len(rows)

            results = run_bulk_modificar(
                page, rows, pre_selected or {}, log, stop, self.captcha_bridge,
                usuario=usuario, password=password,
            )

            all_results = results
            ok_count = sum(1 for r in results if r["status"] == "ok")
            warn_count = sum(1 for r in results if r["status"] in ("not_found",))
            err_count = len(results) - ok_count - warn_count

            # Guardar Excel coloreado
            if all_results and self._excel_path:
                from utils_mod.excel_writer_mod import write_colored_results
                try:
                    sheet = self.combo_sheet.get() if hasattr(self, 'combo_sheet') else "Hoja1"
                    out = write_colored_results(self._excel_path, sheet, all_results)
                    log.ok(f"Excel coloreado guardado: {os.path.basename(out)}")
                except Exception as e:
                    log.warn(f"No se pudo colorear el Excel: {e}")

            self.log_queue.put({"type": "stat_ok", "value": ok_count})
            self.log_queue.put({"type": "stat_warn", "value": warn_count})
            self.log_queue.put({"type": "stat_err", "value": err_count})

            log.ok(f"Completado: {ok_count} OK | {warn_count} No encontrados | {err_count} Errores")

        except Exception as e:
            log.error(f"Error fatal: {e}")
            import traceback
            log.error(traceback.format_exc())
        finally:
            if browser and pw:
                try:
                    close_browser(pw, browser)
                    log.info("Navegador cerrado")
                except Exception:
                    pass

    def _on_stop(self):
        self.stop_event.set()
        self._running = False
        self.btn_stop.configure(state="disabled", text="Deteniendo...")
        self._log("Detención solicitada", "warn")
        self.after(3000, self._reset_after_stop)

    def _reset_after_stop(self):
        self.btn_launch.configure(state="normal")
        self.btn_stop.configure(state="disabled", text="■  Detener")
        self.btn_test.configure(state="normal")
        self.lbl_status.configure(text="Detenido — listo para iniciar", text_color="#f39c12")

    # ── Queue Polling ─────────────────────────────────────────────

    def poll_queue(self):
        try:
            while True:
                item = self.log_queue.get_nowait()
                t = item.get("type") if isinstance(item, dict) else None

                if t == "log":
                    self._log(item.get("msg", ""), item.get("level", "info"))
                elif t == "progress":
                    cur = item.get("current", 0)
                    tot = item.get("total", 1)
                    self.lbl_status.configure(text=f"Procesando producto {cur} de {tot}")
                    self.lbl_counter.configure(text=f"{cur}/{tot}")
                    self.progress.set(cur / tot if tot > 0 else 0)
                elif t == "stat_ok":
                    self._stat_ok.configure(text=str(item.get("value", 0)))
                elif t == "stat_warn":
                    self._stat_warn.configure(text=str(item.get("value", 0)))
                elif t == "stat_err":
                    self._stat_err.configure(text=str(item.get("value", 0)))
                elif t == "done":
                    ok  = item.get("ok", 0)
                    err = item.get("errors", 0)
                    self._running = False
                    self.lbl_status.configure(
                        text=f"Finalizado  ·  {ok} OK  ·  {err} errores"
                    )
                    self.progress.set(1)
                    self.btn_launch.configure(state="normal")
                    self.btn_stop.configure(state="disabled", text="■  Detener")

                # Compatibilidad con LogWriter que usa tuplas (level, msg)
                elif isinstance(item, tuple) and len(item) == 2:
                    level, msg = item
                    level_map = {
                        "OK": "ok", "ERROR": "error",
                        "WARN": "warn", "DONE": "done", "INFO": "info",
                    }
                    tag = level_map.get(level, "info")

                    if level == "DONE":
                        self._running = False
                        self.lbl_status.configure(text=msg)
                        self.progress.set(1)
                        self.btn_launch.configure(state="normal")
                        self.btn_stop.configure(state="disabled", text="■  Detener")
                    else:
                        self._log(msg, tag)

        except queue.Empty:
            pass

        # Verificar bridge de captcha
        with self.captcha_bridge.lock:
            if (
                self.captcha_bridge.image_bytes is not None
                and not self.captcha_bridge.event.is_set()
            ):
                self._show_captcha(self.captcha_bridge.image_bytes)

        self.after(200, self.poll_queue)


# ═══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = ModificarProductosApp()
    app.mainloop()
